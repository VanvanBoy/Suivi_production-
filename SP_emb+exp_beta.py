# -*- coding: utf-8 -*-
"""
Created on Thu Aug 21 15:18:57 2025

@author: User
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os, re, shutil
import io
import tkinter as tk
from tkinter import ttk, messagebox
from ttkthemes import ThemedTk
from PIL import Image, ImageTk
import pandas as pd
import mysql.connector
import webbrowser
from datetime import datetime
from datetime import date
from collections import defaultdict
import qrcode
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as OpenpyxlImage


EXCEL_PATH = r"G:\Drive partagés\VoltR\11_Data\IHM\Suivi_prod_par_modele.xlsx"  

class DBManager:
    def __init__(self):
        
        self.config = {
            'user': 'Vanvan',
            'password': 'VoltR99!',
            'host': '34.77.226.40',
            'database': 'bdd_23102025',
            'auth_plugin': 'mysql_native_password'
        }
        
        """
        self.config = {
            'user': 'Vanvan',
            'password': 'VoltR99!',
            'host': '34.77.226.40',
            'database': 'cellules_batteries_cloud',
            'auth_plugin': 'mysql_native_password'
        }
        """
        """
        self.config = {
            'user': 'root',
            'password': 'VoltR99!',
            'host': '127.0.0.1',
            'database': 'bdd_29072025',
            'auth_plugin': 'mysql_native_password'
        }
        
        """
        
    def connect(self):
        try:
            return mysql.connector.connect(**self.config)
        except mysql.connector.Error as err:
            messagebox.showerror("Erreur DB", f"Erreur lors de la connexion : {err}")
            return None

class StockApp(ThemedTk):

    
    def __init__(self):
        super().__init__(theme="xpnative")
        self.title("Suivi de production")
        self.geometry("1000x650")
        
        self.refresh_ms = 10_000  # 10 secondes
        self._tab_refresh_job = None
        self._refreshing = False
        
        # annuler proprement à la fermeture
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
        self.focus_targets = {}  # stage -> widget qui doit recevoir le focu

        self.db_manager = DBManager()
        self.picking_file_path = None

        # styles
        self.style = ttk.Style()
        self.style.configure("TLabel", font=('Segoe UI', 11), padding=4)
        self.style.configure("TButton", font=('Segoe UI', 11), padding=6)
        self.style.configure("TEntry", font=('Segoe UI', 11))
        self.style.configure("TCombobox", font=('Segoe UI', 11))
        self.style.configure("Danger.TButton", foreground="black", background="#F06F65")
        self.style.map("Danger.TButton", background=[('active', '#e55b50')], foreground=[('disabled', 'black')])
        self.style.configure("Good.TButton", foreground="black", background="#0ED329")
        self.style.map("Good.TButton", background=[('active', '#0ED329')], foreground=[('disabled', 'black')])

        
        self.selected_model = None
        self.stage_order = None  
    
        self.STAGE_TO_DBCOL = {
            "picking":   "picking_tension",          
            "pack":      "soudure_pack",     
            "nappe":     "soudure_nappe",
            "bms":       "soudure_bms",
            "wrap":      "wrap",
            "fermeture_batt": "fermeture_batt",
            "capa":      "test_capa",        
            "emb":       "emballage",        
            "exp":       "expedition",       
            "recherche": "recherche",
            "recyclage": "recyclage"   
        }
        
        self.ALLOWED_STAGE_KEYS = set(self.STAGE_TO_DBCOL.keys())  # pour sécuriser
        
        self._show_model_selector_and_build()
    
    def on_close(self):
        self._cancel_tab_refresh()
        self.destroy()
        
    
    def _required_previous_dbcols(self,current_stage: str):
        """
        Retourne la liste des NOMS DE COLONNES MySQL (dans suivi_production)
        à vérifier (=1) avant de valider current_stage.
    
        On s'appuie sur self.stage_order (rangs > 0 = étape active)
        et sur STAGE_TO_DBCOL pour la correspondance.
        """
        if not hasattr(self, "stage_order") or not self.stage_order:
            return []
    
        if current_stage not in self.ALLOWED_STAGE_KEYS:
            return []
    
        cur_rank = self.stage_order.get(current_stage, 0)
        if cur_rank <= 0:
            return []
    
        prev_pairs = []
        for k, v in self.stage_order.items():
            if k in self.ALLOWED_STAGE_KEYS and v > 0 and v < cur_rank:
                dbcol = self.STAGE_TO_DBCOL.get(k)
                if dbcol:
                    prev_pairs.append((k, v, dbcol))
    
        # tri pour lisibilité/diagnostic
        prev_pairs.sort(key=lambda x: x[1])
    
        # on retourne uniquement les colonnes DB
        return [dbcol for (_, _, dbcol) in prev_pairs]
    
    
    def _check_prereqs_and_warn(self, num_batt: str, current_stage: str) -> bool:
        """
        Vérifie dans suivi_production que TOUTES les colonnes prérequis (=1)
        sont validées pour num_batt, d'après le mapping STAGE_TO_DBCOL.
        """
        prev_dbcols = self._required_previous_dbcols(current_stage)
        if not prev_dbcols:
            return True  # pas de prérequis
    
        conn = self.db_manager.connect()
        if not conn:
            return False
        try:
            cursor = conn.cursor(dictionary=True)
    
            # Sécurisation : on n'insère que des noms de colonnes issus de STAGE_TO_DBCOL
            cols_sql = ", ".join(prev_dbcols)
            sql = f"""
                SELECT {cols_sql}
                FROM suivi_production
                WHERE numero_serie_batterie = %s
                LIMIT 1
            """
            cursor.execute(sql, (num_batt,))
            row = cursor.fetchone()
            if not row:
                messagebox.showwarning("Vérification", "Sélectionner une batterie")
                return False
    
            missing = [c for c in prev_dbcols if (row.get(c) or 0) != 1]
            if missing:
                lis = ", ".join(missing)
                messagebox.showwarning(
                    "Pré-requis manquants",
                    f"Impossible de valider '{current_stage}'. Étape(s) non validée(s) : {lis}"
                )
                return False
    
            return True
        except Exception as e:
            messagebox.showerror("SQL", f"Erreur vérification prérequis:\n{e}")
            return False
        finally:
            try:
                cursor.close(); conn.close()
            except:
                pass
                
    def _show_model_selector_and_build(self):
        
        try:
            df = pd.read_excel(EXCEL_PATH,sheet_name="Flux test modele")
        except Exception as e:
            messagebox.showerror("Excel", f"Impossible de lire l'Excel:\n{e}")
            self.destroy()
            return

        if "nom_modele" not in df.columns:
            messagebox.showerror("Excel", "La colonne 'nom_modele' est absente du fichier.")
            self.destroy()
            return

        self.models = df["nom_modele"].dropna().astype(str).unique().tolist()

        
        self._create_widgets_with_order()
        
        self.funcs_to_run= [
            self.setup_emb,
            self.setup_exp,
            self.setup_recherche,
            self.setup_recyclage
        ]
        
        for f in self.funcs_to_run:
            print(f.__name__)


    def _create_widgets_with_order(self):
        
        self.title("Suivi de production emballage + expédition")
        stage_defs = {
            "emb":       ("Contrôle emballage", self.setup_emb),
            "exp":       ("Contrôle expédition", self.setup_exp),
            "recherche": ("Recherche de batterie", self.setup_recherche),
            "recyclage": ("Gestion recyclage", self.setup_recyclage)
        }
 
        ordered_keys = [
            "emb" ,    
            "exp",
            "recherche",
            "recyclage"
        ]

        if not ordered_keys:
            messagebox.showwarning("Configuration", "Aucun onglet actif pour cette référence.")
          
            ordered_keys = ["picking"]
        
        self.ordered_keys = ordered_keys

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill="both")
        
        self.tab_to_stage = {}   # tab_id (str) -> 'picking' | 'pack' | ...
        self.stage_refreshers = {}  # 'picking' -> fonction reload

        for key in ordered_keys:
            title, setup_fn = stage_defs[key]
            print(title)
            frame = ttk.Frame(self.notebook)
            self.notebook.add(frame, text=title)
            focus_widget=setup_fn(frame)
            if focus_widget is not None:
                self.register_focus_target(key, focus_widget)
            # mémorise le mapping tab -> stage
            self.tab_to_stage[str(frame)] = key
            
    
    
        # Bind: lorsque l’onglet change → reload immédiat + restart timer
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
    
        # Premier reload immédiat, puis démarrage du cycle 10 s
        self._refresh_active_tab_now()
        self._schedule_next_tab_tick()
        self.after(0, self._focus_active_tab)
    
    def register_focus_target(self, stage: str, widget):
        """Enregistre le widget qui doit recevoir le focus au changement d'onglet."""
        if not hasattr(self, "focus_targets"):
            self.focus_targets = {}
        if widget is not None:
            self.focus_targets[stage] = widget
            
    def make_tab_chain(self, widgets, submit_button=None, ring=True, enter_from_fields=False):
        """
        Définit l'ordre Tab explicite et fait en sorte que seul le bouton
        déclenche l'action quand il est focalisé.
        - widgets : liste ordonnée des widgets (Entry/Combobox/Button/…).
                    Si tu veux que Tab atteigne le bouton, mets-le en dernier.
        - submit_button : ttk.Button (optionnel). Seul le bouton recevra <Return>.
        - ring : True => Tab boucle sur le 1er widget
        - enter_from_fields : si True, <Return> sur les champs déclenchera submit_button
                              (NE PAS utiliser si dangereux). Par défaut False.
        """
        if not widgets:
            return
    
        # assure takefocus sur les widgets inclus
        for w in widgets:
            try:
                w['takefocus'] = True
            except Exception:
                pass
    
        def next_idx(i):
            return (i + 1) % len(widgets) if ring else min(i + 1, len(widgets) - 1)
        def prev_idx(i):
            return (i - 1) % len(widgets) if ring else max(i - 1, 0)
    
        for i, w in enumerate(widgets):
            # handlers capturant i (closure sûre grâce à i=i)
            def go_next(e, i=i):
                widgets[next_idx(i)].focus_set()
                return "break"
            def go_prev(e, i=i):
                widgets[prev_idx(i)].focus_set()
                return "break"
    
            # Bind Tab / Shift-Tab (généralement suffisants)
            w.bind("<Tab>", go_next)
            w.bind("<Shift-Tab>", go_prev)
    
            # Certains environnements linux/old-tk envoient ISO_Left_Tab pour Shift-Tab.
            # On essaye de binder mais on ignore proprement l'erreur si le keysym n'existe pas.
            try:
                w.bind("<ISO_Left_Tab>", go_prev)
            except tk.TclError:
                # keysym non supporté sur cette plateforme : on ignore silencieusement
                pass
    
            # IMPORTANT: on NE bind pas <Return> sur les champs par défaut (sécurité)
            if enter_from_fields and submit_button is not None:
                w.bind("<Return>", lambda e, b=submit_button: b.invoke())
    
        # Sur le bouton : Enter et Espace déclenchent l'action (mais seulement si le bouton a le focus)
        if submit_button is not None:
            try:
                submit_button['takefocus'] = True
            except Exception:
                pass
            # binding sûr sur le bouton lui-même
            submit_button.bind("<Return>", lambda e, b=submit_button: b.invoke())
            submit_button.bind("<space>", lambda e, b=submit_button: b.invoke())
        
    
    def _get_active_stage(self) -> str | None:
        if not hasattr(self, "notebook"):
            return None
        tab_id = self.notebook.select()
        return self.tab_to_stage.get(tab_id)
    
    def _refresh_active_tab_now(self):
        if self._refreshing:
            return
        stage = self._get_active_stage()
        if not stage:
            return
        ref_fn = self.stage_refreshers.get(stage)
        if not callable(ref_fn):
            return
        self._refreshing = True
        try:
            ref_fn()  # ta fonction qui va en BDD et recharge le Treeview de l’onglet
        except Exception as e:
            print(f"[refresh:{stage}] {e}")
        finally:
            self._refreshing = False
    
    def _on_tab_changed(self, event=None):
        self._cancel_tab_refresh()
        self._refresh_active_tab_now()   # reload immédiat en arrivant sur le nouvel onglet
        self._focus_active_tab()         # <<< donne le focus dans le nouvel ongle
        self._schedule_next_tab_tick()   # redémarre le timer
    
    def _tab_tick(self):
        self._refresh_active_tab_now()
        self._schedule_next_tab_tick()
    
    def _schedule_next_tab_tick(self):
        # petit jitter pour éviter que 10 postes tapent la DB exactement en même temps
        import random
        jitter = random.randint(0, 1500)
        self._tab_refresh_job = self.after(self.refresh_ms + jitter, self._tab_tick)
    
    def _cancel_tab_refresh(self):
        if self._tab_refresh_job is not None:
            try:
                self.after_cancel(self._tab_refresh_job)
            except Exception:
                pass
            self._tab_refresh_job = None
            
    def register_focus_target(self, stage: str, widget):
        """À appeler depuis tes setup_* pour déclarer le champ à focus par défaut."""
        if widget is not None:
            self.focus_targets[stage] = widget
    
    def _find_first_input(self, container):
        """Fallback: cherche récursivement le premier Entry/Combobox/Text dans un frame."""
        for w in container.winfo_children():
            # si c'est un champ saisissable
            if isinstance(w, (ttk.Entry, tk.Entry, ttk.Combobox, tk.Text)):
                return w
            # si c'est un conteneur, on descend
            if isinstance(w, (ttk.Frame, tk.Frame, ttk.LabelFrame, ttk.Labelframe, ttk.Panedwindow)):
                found = self._find_first_input(w)
                if found:
                    return found
        return None
    
    def _focus_active_tab(self):
        """Donne le focus au widget cible de l'onglet actif (ou fallback)."""
        stage = self._get_active_stage()
        if not stage or not hasattr(self, "notebook"):
            return
    
        tab_id = self.notebook.select()
        tab = self.nametowidget(tab_id)
    
        target = self.focus_targets.get(stage)
        if target is None or not target.winfo_exists():
            target = self._find_first_input(tab)
    
        if target and target.winfo_exists():
            # after(0) pour que l'onglet ait fini de s'afficher
            self.after(0, lambda: (
                target.focus_set(),
                # si Entry/Combobox: on met le curseur à la fin + on sélectionne tout (optionnel)
                (hasattr(target, "icursor") and target.icursor('end')),
                (hasattr(target, "selection_range") and target.selection_range(0, 'end'))
            ))

    
    def set_photo(self, label: tk.Label, chemin_image: str, size=(200, 200)):
        """Charge une image et l'affecte AU SEUL label donné."""
        try:
            img = Image.open(chemin_image)
            img = img.resize(size)
            photo = ImageTk.PhotoImage(img)
            label.config(image=photo, text="")
            label.image = photo  # garder une référence pour éviter GC
        except Exception as e:
            messagebox.showerror("Erreur image", f"Impossible de charger l'image : {e}")
            
    def verfier_coherence_ref(self,num_batt):
        modele=self.selected_model
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "Select reference_produit_voltr from produit_voltr where numero_serie_produit = %s "
            param = (num_batt,)
            cursor.execute(query, param)
            modele_act=cursor.fetchone()[0]
            if modele_act != modele:
                reponse=messagebox.askyesno("Modèle de batterie différent",f"Le mdodèle de batterie n'est pas coherent \n Passer du modèle {modele_act} au modèle {modele} pour la batterie {num_batt} ?")
                if not reponse:
                    return 'stop'
                else :
                    cursor.execute("UPDATE produit_voltr SET reference_produit_voltr = %s WHERE numero_serie_produit =%s",(modele,num_batt))
                    messagebox.showinfo("Nouveau modèle", f'la batterie {num_batt} est passé au modèle {modele}.')
                    conn.commit()
                    return 'next'
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()  
    
    def changer_ref_batterie(self,new_ref,num_batt):
        modele=self.selected_model
        if modele == new_ref:
            messagebox.showinfo("Réference identique","Le nouveau modele est identique au modele actuel")
            return 
        else :
            conn = self.db_manager.connect()
            if not conn:
                return
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE produit_voltr SET reference_produit_voltr = %s WHERE numero_serie_produit =%s",(new_ref,num_batt))
                messagebox.showinfo("Nouveau modèle", f'la batterie {num_batt} est passé au modèle {new_ref}.')
                conn.commit()
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
            finally:
                try:
                    cursor.close()
                except:
                    pass
                conn.close()  
                
    def build_stage_query(self, current_stage):
        # Trouver l’index de l’étape courante
        idx = self.ordered_keys.index(current_stage)
    
        # Étapes précédentes
        previous_stages = self.ordered_keys[:idx]
    
        # Colonnes SQL associées
        prev_cols = [self.STAGE_TO_DBCOL[s] for s in previous_stages]
        current_col = self.STAGE_TO_DBCOL[current_stage]
    
        # Condition : toutes les étapes précédentes doivent être validées (=1)
        prev_conditions = " AND ".join([f"sp.{col} = 1" for col in prev_cols]) if prev_cols else "1=1"
    
        # Condition : l’étape courante doit être non validée (=0 ou NULL)
        current_condition = f"(sp.{current_col} = 0 OR sp.{current_col} IS NULL)"
    
        if current_stage != 'exp' and current_stage != 'emb' :
            # Construire la requête
            query = f"""
                SELECT sp.numero_serie_batterie
                FROM suivi_production AS sp
                JOIN produit_voltr AS pv
                  ON sp.numero_serie_batterie = pv.numero_serie_produit
                WHERE {prev_conditions}
                  AND {current_condition}
                  AND pv.reference_produit_voltr = %s
            """
        else: 
            # Construire la requête (pas de filtre ref produit)
            query = f"""
                SELECT sp.numero_serie_batterie
                FROM suivi_production AS sp
                JOIN produit_voltr AS pv
                  ON sp.numero_serie_batterie = pv.numero_serie_produit
                WHERE {prev_conditions}
                  AND {current_condition}
            """
    
        return query

    #------------------------------ Onglet emballage -----------------------------------------------   
    
    def setup_emb(self, frame):
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        
        # --- Frame droite ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.emb_numero_serie_batt_entry = ttk.Entry(left_frame)
        self.emb_numero_serie_batt_entry.pack(pady=5)
        self.emb_numero_serie_batt_entry.bind("<KeyRelease>", self.emb_check_entry_length)
        
        ttk.Label(left_frame,text="Réference batterie:").pack(pady=5)
        cb_var_r=tk.StringVar()
        self.emb_ref_combo = ttk.Combobox(
            left_frame,
            textvariable=cb_var_r,
            values=self.models,
            #state="disabled",
            width=40)
        self.emb_ref_combo.pack(pady=5)
        self.emb_ref_combo.bind("<<ComboboxSelected>>", self.emb_on_select_ref)
        
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        
        # --- Bloc Listbox + Scrollbar ---
        listbox_frame = tk.Frame(left_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
        
        self.emb_listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10,
            selectmode='multiple',    # ← permet sélection multiple (Ctrl/Shift)
            exportselection=False     # ← conserve la sélection quand la listbox perd le focus
        )
        self.emb_listbox_batt.pack(side="left", fill="both", expand=True)
        
        
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.emb_listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
        
        self.emb_listbox_batt.config(yscrollcommand=scrollbar.set)
        #self.emb_listbox_batt.bind("<<ListboxSelect>>", self.emb_on_select_batt)
        # --- fin bloc listbox ---
        
        self.btn_emb_batt = ttk.Button(
            left_frame, text="Emballer les batteriess",
            command=self.emballer_batterie,
            style="TButton"
        )
        self.btn_emb_batt.pack(pady=5)
  
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie_emb,
            style="Danger.TButton"
        ).pack(pady=10)
          
        self.check_var = tk.BooleanVar(value=False)  # False = décoché, True = coché
        # --- Variable pour Radiobutton ---
        self.check = tk.Checkbutton(right_frame, text="emb individuel", variable=self.check_var)
        self.check.pack(pady=5)
        
        self.entry_var = tk.StringVar(value="")
        self.entry = tk.Entry(right_frame, textvariable=self.entry_var)#, state="readonly")  # readonly pour display only
        self.entry.pack(pady=5)
        
        listbox2_frame = tk.Frame(right_frame)
        listbox2_frame.pack(fill="both", expand=True, pady=5)
        
        ttk.Label(right_frame, text="Liste des batteries emballées:").pack(pady=5)
        # Treeview avec 2 colonnes : numéro_serie et reference_produit
        self.emb_tree_f = ttk.Treeview(
            listbox2_frame,
            columns=("numero_serie", "reference_produit"),
            show="headings",
            height=10
        )
        self.emb_tree_f.heading("numero_serie", text="N° série")
        self.emb_tree_f.heading("reference_produit", text="Réf produit")
        self.emb_tree_f.column("numero_serie", width=120, anchor="center")
        self.emb_tree_f.column("reference_produit", width=200, anchor="w")
        self.emb_tree_f.pack(side="left", fill="both", expand=True)
        self.emb_tree_f.bind("<Double-1>", self.on_double_click_delete)
        
        # Scrollbar verticale attachée au treeview
        scrollbar = tk.Scrollbar(listbox2_frame, orient="vertical", command=self.emb_tree_f.yview)
        scrollbar.pack(side="right", fill="y")
        self.emb_tree_f.configure(yscrollcommand=scrollbar.set)
        
        ttk.Button(
            right_frame, text="✅ Contrôle OK",
            command=self.valider_emballage,
            style="Good.TButton"
        ).pack(pady=10)
        
        self.creer_numero_emballage()
    
    def creer_numero_emballage(self):
        datej=str(date.today())
        conn= self.db_manager.connect()
        if not conn:
            messagebox.showerror("Erreur BDD", "Impossible de se connecter à la base de données.")
            return
        cursor=conn.cursor()
        query="""
        SELECT COUNT(DISTINCT num_emballage) AS nb_valeurs_diff
        FROM suivi_production
        WHERE num_emballage LIKE %s;
        """
        param=(datej+'%',)
        cursor.execute(query,param)
        num_jour=int(cursor.fetchone()[0])+1
        num_jour_d=str(num_jour).zfill(3)
        datej=str(date.today())
        code_emb=datej+'-emb'+num_jour_d
        
        self.entry.delete(0,tk.END)
        self.entry.insert(0,code_emb)
        
    def on_double_click_delete(self, event):
        # récupère l'item sous le curseur
        item_id = self.emb_tree_f.identify_row(event.y)
        if not item_id:
            return  # clic en dehors d'une ligne
        # supprime l'item
        self.emb_tree_f.delete(item_id)
        
    def emb_on_select_ref(self, event=None):
        self.emb_listbox_batt.delete(0, tk.END)
        ref_prod=self.emb_ref_combo.get()
        conn = self.db_manager.connect()
        if not conn:
            messagebox.showerror("Erreur BDD", "Impossible de se connecter à la base de données.")
            return
        try: 
            cursor = conn.cursor()
            query = "SELECT sp.numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie=pv.numero_serie_produit WHERE sp.fermeture_batt=1 and sp.emballage is null and sp.recyclage is null and pv.reference_produit_voltr= %s"
            param = (ref_prod,)
            cursor.execute(query, param)
            rows = cursor.fetchall()
            
            for r in rows:
                serial = r[0] if r and r[0] is not None else ""
                self.emb_listbox_batt.insert(tk.END, str(serial))
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la lecture BDD :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
        
    def emb_on_select_batt(self):
        print('batt')
        
        
    def group_numero_by_ref(self):
        """
        Retourne un dict { reference_produit: [numero_serie, ...], ... }
        Ne garde que les lignes qui ont des values valides.
        """
        grouped = defaultdict(list)
        for iid in self.emb_tree_f.get_children():
            vals = self.emb_tree_f.item(iid, "values")
            if not vals:
                continue
            # supposons vals = (numero_serie, reference_produit, ...)
            numero = vals[0]
            ref = vals[1] if len(vals) > 1 else None
            if not ref:
                continue
            # nettoyage simple
            numero = str(numero).strip()
            ref = str(ref).strip()
            if numero == "" or ref == "":
                continue
            # évite doublons dans une même ref
            if numero not in grouped[ref]:
                grouped[ref].append(numero)
        return dict(grouped)
    
    def count_per_ref(self):
        grouped = self.group_numero_by_ref()
        return {ref: len(nums) for ref, nums in grouped.items()}
        
    def valider_emballage(self):
        
        def create_packaging_excel(self, numero_emballage: str = None, save_dir: str = None):
            """
            Crée un Excel récapitulant les batteries emballées.
            - numero_emballage: nom du fichier (sans extension). Si None, on génère un nom timestamp.
            - save_dir: dossier où enregistrer le fichier (par défaut working dir).
            Retourne le chemin complet du fichier créé, ou None si pas de données.
            """
            # Récupération des données depuis le treeview
            # attend que self.group_numero_by_ref() et self.count_per_ref() existent
            grouped = getattr(self, "group_numero_by_ref", None)
            count_fn = getattr(self, "count_per_ref", None)
            if grouped is None or count_fn is None:
                raise RuntimeError("Méthodes group_numero_by_ref() et count_per_ref() requises dans la classe.")
        
            grouped_dict = self.group_numero_by_ref()  # {ref: [numero_serie,...], ...}
            counts = self.count_per_ref()              # {ref: count, ...}
        
            if not grouped_dict:
                # Rien à exporter
                return None
        
            # Nom du fichier
            if not numero_emballage:
                numero_emballage = datetime.now().strftime("emballage_%Y%m%d_%H%M%S")
            filename = f"{numero_emballage}.xlsx"
            if save_dir:
                os.makedirs(save_dir, exist_ok=True)
                filepath = os.path.join(save_dir, filename)
            else:
                filepath = os.path.join(os.getcwd(), filename)
        
            # --- Préparer les DataFrames ---
            # Récap : ref, quantité
            recap_df = pd.DataFrame(
                [(ref, int(counts.get(ref, 0))) for ref in sorted(counts.keys())],
                columns=["reference_produit", "quantité"]
            )
        
            # Détails : une ligne par numéro de série
            detail_rows = []
            for ref, numeros in grouped_dict.items():
                for num in numeros:
                    detail_rows.append({"reference_produit": ref, "numero_serie": num})
            details_df = pd.DataFrame(detail_rows)
        
            # Date et méta
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
            # --- Créer un classeur openpyxl et écrire les DataFrames ---
            wb = Workbook()
            # Supprime la feuille par défaut et en recrée une proprement
            default_sheet = wb.active
            default_sheet.title = "Récap"
            ws_recap = default_sheet
        
            # Insérer en-têtes d'info en haut
            ws_recap["A1"] = "N° emballage"
            ws_recap["B1"] = numero_emballage
            ws_recap["A2"] = "Date"
            ws_recap["B2"] = date_str
        
            # Ecrire le recap_df à partir de la ligne 4
            start_row = 4
            for r_idx, row in enumerate(dataframe_to_rows(recap_df, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    ws_recap.cell(row=r_idx, column=c_idx, value=value)
        
            # --- Générer QR code (numéro d'emballage) et l'insérer ---
            try:
                qr = qrcode.QRCode(box_size=6, border=2)
                qr.add_data(numero_emballage)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")
                # passer par BytesIO pour insertion
                bio = io.BytesIO()
                qr_img.save(bio, format="PNG")
                bio.seek(0)
                img = OpenpyxlImage(bio)
                # placer le QR dans une cellule (par ex. colonne D, ligne 1)
                img.anchor = "D1"
                ws_recap.add_image(img)
            except Exception as e:
                # QR non critique : on ignore en cas d'erreur
                print("Erreur génération QR:", e)
        
            # Ajouter feuille Détails
            ws_details = wb.create_sheet(title="Détails")
            # écrire le details_df avec headers
            for r_idx, row in enumerate(dataframe_to_rows(details_df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws_details.cell(row=r_idx, column=c_idx, value=value)

            # Optionnel: feuille complète CSV-style par référence (liste des nums en une cellule)
            # Exemple si tu veux aussi une feuille avec "ref" + "liste_numero_serie_concat"
            ws_list_by_ref = wb.create_sheet(title="Liste par Ref")
            ws_list_by_ref.append(["reference_produit", "numero_series_concat"])
            for ref in sorted(grouped_dict.keys()):
                nums = grouped_dict[ref]
                concat = ", ".join(nums)
                ws_list_by_ref.append([ref, concat])

            # Sauvegarder
            wb.save(filepath)
            return filepath
        
        
        def show_confirmation_from_counts(self):
            """
            Récupère le dict {ref: count} via self.count_per_ref()
            et affiche un askyesno du type :
            "Vous avez emballé 12 REF_A, 10 REF_B. Valider ?"
            Retourne True si l'utilisateur clique 'Oui', False sinon.
            """
            counts = self.count_per_ref()  # suppose que cette méthode existe et retourne {ref: count}
            if not counts:
                messagebox.showinfo("Récapitulatif", "Aucune référence trouvée.")
                return False
        
            # Option : trier les références pour affichage stable
            items = sorted(counts.items())
        
            # Construire les fragments "12 REF_A" ou "1 REF_A" (on peut ajouter une mise en forme si besoin)
            parts = []
            for ref, cnt in items:
                # nettoyage simple des valeurs
                ref_str = str(ref).strip()
                cnt_int = int(cnt) if isinstance(cnt, (int, float, str)) else cnt
                parts.append(f"{cnt_int} {ref_str}")
        
            # Joindre par ", " puis construire le message
            recap = ", ".join(parts)
            msg = f"Vous avez emballé {recap}. Valider ?"
        
            # Option: si tu veux limiter la longueur dans la boite de dialogue :
            # if len(msg) > 200: msg = msg[:197] + "..."
        
            answer = messagebox.askyesno("Confirmer emballage", msg)
            if answer:
                # ici, place l'action à effectuer si l'utilisateur valide
                # par ex. self.finalize_emballage()
                pass
            else:
                # action si annulation (facultatif)
                pass
        
            return answer

        def get_all_numero_serie(self):
            numeros = []
            for iid in self.emb_tree_f.get_children():
                vals = self.emb_tree_f.item(iid, "values")
                if vals:                      # sécurité si values = ()
                    numeros.append(vals[0])   # 0 = première colonne (numero_serie)
            return numeros
        
        
        emb_batts=get_all_numero_serie(self)
        if not emb_batts:
            messagebox.showerror('Pas de batterie emballées','Veuillez emballer des batteries')
        
        reponse=show_confirmation_from_counts(self)
        
        if reponse==True:
            conn=self.db_manager.connect()
            cursor=conn.cursor()
            for batt_emb in emb_batts:
                if self.check_var.get()==True:
                    query="Update suivi_production set emballage=1, date_emballage=NOW() where numero_serie_batterie= %s"
                    param=(batt_emb,)
                    cursor.execute(query,param)
                    
                else :
                    num_emb=self.entry.get()
                    query="Update suivi_production set emballage=1,date_emballage=NOW(),num_emballage=%s where numero_serie_batterie= %s"
                    param=(num_emb,batt_emb)

                    cursor.execute(query,param)
            
            if self.check_var.get()==False:
                path = create_packaging_excel(self,numero_emballage=num_emb, save_dir=r"G:\Drive partagés\VoltR\4_Production\8_Picking\Suivi_prod_emballage")
                if path:
                    messagebox.showinfo("Export OK", f"Fichier créé :\n{path}")
                    
                else:
                    messagebox.showwarning("Aucun", "Aucune batterie à exporter.")
                    return
            
            conn.commit()
            cursor.close()
            conn.close()
            
            self.creer_numero_emballage() 
            
            for item in self.emb_tree_f.get_children():
                self.emb_tree_f.delete(item)
        
    def emballer_batterie(self):
        
        def get_selected_batts(self):
            idxs = self.emb_listbox_batt.curselection()  # tuple d'indices sélectionnés
            return [self.emb_listbox_batt.get(i) for i in idxs]
        
        def delete_selected_batts(self):
            for i in reversed(self.emb_listbox_batt.curselection()):
                self.emb_listbox_batt.delete(i)
                
        batt_emballees=get_selected_batts(self)
        for batt in batt_emballees:
            conn = self.db_manager.connect()
            if not conn:
                messagebox.showerror("Erreur BDD", "Impossible de se connecter à la base de données.")
                return
        
            try:
                cursor = conn.cursor()
                query = "SELECT reference_produit_voltr FROM produit_voltr WHERE numero_serie_produit = %s"
                param = (batt,)
                cursor.execute(query, param)
                row = cursor.fetchone()
        
                if not row:
                    # pas trouvé : avertir et ne pas insérer
                    messagebox.showwarning("Non trouvé", f"Le N° série {batt} n'existe pas en base.")
                    self.emb_numero_serie_batt_entry.delete(0, tk.END)
                    return
        
                reference_produit = row[0] if row[0] is not None else ""
        
                # --- vérifier doublon dans le treeview ---
                # on utilise le numéro de série comme iid pour empêcher les doublons
                if self.emb_tree_f.exists(batt):
                    # déjà présent -> on peut jouer un beep ou message discret
                    # par ex. on met la ligne en surbrillance et on fait un petit retour
                    self.emb_tree_f.selection_set(batt)
                    self.emb_tree_f.see(batt)
                    messagebox.showinfo("Doublon", f"Le N° série {batt} est déjà dans la liste.")
                    self.emb_numero_serie_batt_entry.delete(0, tk.END)
                    return
        
                # --- insérer la nouvelle ligne ---
                self.emb_tree_f.insert("", "end", iid=batt, values=(batt, reference_produit))
        
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la lecture BDD :\n{e}")
            finally:
                try:
                    cursor.close()
                except:
                    pass
                conn.close()
                
        delete_selected_batts(self)
            
    def emb_check_entry_length(self, event=None):
        """
        Déclenchée quand le contenu de self.emb_numero_serie_batt_entry change.
        Si la chaîne contient exactement 9 chiffres, on cherche la référence en BDD
        et on ajoute la ligne dans le Treeview (sans doublon).
        """
        numero_serie_cell = self.emb_numero_serie_batt_entry.get().strip()
    
        # Ne rien faire si longueur différente de 9
        if len(numero_serie_cell) != 9:
            return
    
        # --- connexion BDD ---
        conn = self.db_manager.connect()
        if not conn:
            messagebox.showerror("Erreur BDD", "Impossible de se connecter à la base de données.")
            return
    
        try:
            cursor = conn.cursor()
            query = "SELECT reference_produit_voltr FROM produit_voltr WHERE numero_serie_produit = %s"
            param = (numero_serie_cell,)
            cursor.execute(query, param)
            row = cursor.fetchone()
    
            if not row:
                # pas trouvé : avertir et ne pas insérer
                messagebox.showwarning("Non trouvé", f"Le N° série {numero_serie_cell} n'existe pas en base.")
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
                return
    
            reference_produit = row[0] if row[0] is not None else ""
    
            # --- vérifier doublon dans le treeview ---
            # on utilise le numéro de série comme iid pour empêcher les doublons
            if self.emb_tree_f.exists(numero_serie_cell):
                # déjà présent -> on peut jouer un beep ou message discret
                # par ex. on met la ligne en surbrillance et on fait un petit retour
                self.emb_tree_f.selection_set(numero_serie_cell)
                self.emb_tree_f.see(numero_serie_cell)
                messagebox.showinfo("Doublon", f"Le N° série {numero_serie_cell} est déjà dans la liste.")
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
                return
    
            # --- insérer la nouvelle ligne ---
            self.emb_tree_f.insert("", "end", iid=numero_serie_cell, values=(numero_serie_cell, reference_produit))
    
            # optionnel : vider l'Entry après insertion
            self.emb_numero_serie_batt_entry.delete(0, tk.END)
    
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la lecture BDD :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
    def add_non_conf_batterie_emb(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
                
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
        
        self.emb_numero_serie_batt_entry.delete(0, tk.END)
        
    #------------------------------ Onglet recherche -----------------------------------------------   
        
    def setup_recherche(self, frame):
        # ----- Layout principal : gauche | boutons | droite -----
        container = ttk.Frame(frame); container.pack(fill="both", expand=True, padx=12, pady=12)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=0)
        container.columnconfigure(2, weight=2)
        container.rowconfigure(0, weight=1)
    
        # ------- Colonne gauche : entrées + combobox + listbox -------
        left = ttk.LabelFrame(container, text="Recherche", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
       
    
        ttk.Label(left, text="N° série cellule").grid(row=0, column=0, sticky="w")
        self.rech_entry_cell = ttk.Entry(left, width=28)
        self.rech_entry_cell.grid(row=1, column=0, sticky="we", pady=(0,8))
        # Remplissage auto du n° batterie quand l'entry cellule atteint 12 chars
        self.rech_entry_cell.bind("<KeyRelease>", self._rech_on_cell_entry)
    
        ttk.Label(left, text="N° série batterie").grid(row=2, column=0, sticky="w")
        self.rech_entry_batt = ttk.Entry(left, width=28)
        self.rech_entry_batt.grid(row=3, column=0, sticky="we", pady=(0,8))
    
        ttk.Label(left, text="Référence batterie").grid(row=4, column=0, sticky="w")
        # valeur par défaut nulle (vide)
        self.rech_model_var = tk.StringVar(value="")
        self.rech_combo = ttk.Combobox(left, textvariable=self.rech_model_var,
                                       values=(self.models or []), state="readonly", width=30)
        self.rech_combo.grid(row=5, column=0, sticky="we", pady=(0,8))
        self.rech_combo.bind("<<ComboboxSelected>>", lambda e: self._rech_on_model_change())
    
        ttk.Label(left, text="Liste batterie").grid(row=6, column=0, sticky="w")
        lb_frame = ttk.Frame(left); lb_frame.grid(row=7, column=0, sticky="nsew")
        left.rowconfigure(7, weight=1)
    
        # multi-sélection
        self.rech_listbox = tk.Listbox(lb_frame, height=10, activestyle="dotbox", selectmode="extended")
        yscroll = ttk.Scrollbar(lb_frame, orient="vertical", command=self.rech_listbox.yview)
        self.rech_listbox.configure(yscrollcommand=yscroll.set)
        self.rech_listbox.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
    
        # Double-clic => déplacer à droite
        self.rech_listbox.bind("<Double-1>", lambda e: self._rech_move_right())
    
        # --------- Colonne boutons centraux ----------
        mid = ttk.Frame(container); mid.grid(row=0, column=1, sticky="ns")
        for i in range(3): mid.rowconfigure(i, weight=1)
        ttk.Button(mid, text="→", width=3, command=self._rech_move_right).grid(row=0, column=0, pady=4)
        ttk.Button(mid, text="←", width=3, command=self._rech_remove_selected_right).grid(row=1, column=0, pady=4)
    
        # --------- Colonne droite : table dynamique ----------
        right = ttk.LabelFrame(container, text="Sélection / Détails", padding=10)
        right.grid(row=0, column=2, sticky="nsew", padx=(8,0))
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        
        # largeur fixe (par ex. 500 px, ajuste comme tu veux)
        right.configure(width=800)
        right.grid_propagate(False)   # bloque l’expansion auto
        
        self.rech_right_title = ttk.Label(right, text="", font=("", 10, "bold"))
        self.rech_right_title.grid(row=0, column=0, sticky="w", pady=(0,6))
    
        tv_frame = ttk.Frame(right); tv_frame.grid(row=1, column=0, sticky="nsew")
        # + scroll vertical & horizontal
        self.rech_tree = ttk.Treeview(tv_frame, columns=(), show="headings", selectmode="extended")
        y2 = ttk.Scrollbar(tv_frame, orient="vertical", command=self.rech_tree.yview)
        x2 = ttk.Scrollbar(tv_frame, orient="horizontal", command=self.rech_tree.xview)
        self.rech_tree.configure(yscrollcommand=y2.set, xscrollcommand=x2.set)
        self.rech_tree.grid(row=0, column=0, sticky="nsew")
        y2.grid(row=0, column=1, sticky="ns")
        x2.grid(row=1, column=0, sticky="ew")
        tv_frame.rowconfigure(0, weight=1); tv_frame.columnconfigure(0, weight=1)
    
        # set utilisé pour éviter les doublons à droite (clé = numero_serie_batterie)
        self._rech_right_keys = set()
        
    def _rech_on_cell_entry(self, event=None):
        """Quand l'entry cellule atteint 12 chars, on cherche la batterie associée et on remplit l'entry batterie."""
        numero_serie_cell = self.rech_entry_cell.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # affectation_produit = numero_serie_batterie (selon ta logique existante)
            cur.execute("SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s", (numero_serie_cell,))
            row = cur.fetchone()
            self.rech_entry_batt.delete(0, tk.END)
            if row and row[0]:
                self.rech_entry_batt.insert(0, str(row[0]))
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Lookup cellule→batterie impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_on_model_change(self):
        """Quand on choisit un modèle, on alimente la liste des n° batteries via la jointure demandée."""
        ref = self.rech_model_var.get().strip()
        self.rech_listbox.delete(0, tk.END)
        if not ref:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # Liste des NUMÉROS DE SÉRIE BATTERIE pour la référence choisie
            # sp = suivi_production / p = produit
            sql = ("""
                SELECT DISTINCT sp.numero_serie_batterie
                FROM suivi_production sp
                JOIN produit_voltr p
                  ON sp.numero_serie_batterie = p.numero_serie_produit
                WHERE p.reference_produit_voltr = %s
                ORDER BY sp.numero_serie_batterie
            """)
            cur.execute(sql, (ref,))
            for (num_batt,) in cur.fetchall():
                if num_batt:
                    self.rech_listbox.insert(tk.END, str(num_batt))
            self.rech_right_title.config(text=f"Modèle sélectionné : {ref}")
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement liste batteries impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_move_right(self):
        """Ajoute à droite : 1) le n° saisi dans l’entry batterie (si présent),
        2) tous les n° sélectionnés dans la liste. Charge les LIGNES de suivi_production correspondantes."""
        # rassembler les cibles
        targets = set()
        batt_from_entry = self.rech_entry_batt.get().strip()
        if batt_from_entry:
            targets.add(batt_from_entry)
        for idx in self.rech_listbox.curselection():
            targets.add(self.rech_listbox.get(idx))
    
        if not targets:
            return
    
        # Charger les lignes depuis suivi_production
        rows, colnames = self._rech_load_suivi_rows(list(targets))
        if not rows:
            return
    
        # Configurer le tree si c'est la 1ère fois ou si colonnes différentes
        self._rech_configure_tree_for_columns(colnames)
    
        # Insérer sans doublon (clé = numero_serie_batterie)
        try:
            k_idx = colnames.index("numero_serie_batterie")  # l’utilisateur veut dédupliquer sur ce champ
        except ValueError:
            # si absent (peu probable), on déduplique sur la 1ère colonne
            k_idx = 0
    
        added = 0
        for r in rows:
            key = str(r[k_idx]) if r[k_idx] is not None else ""
            if key and key not in self._rech_right_keys:
                self.rech_tree.insert("", tk.END, values=[("" if v is None else v) for v in r])
                self._rech_right_keys.add(key)
                added += 1
    
        if added == 0:
            self.rech_right_title.config(text="Aucun nouvel élément (déduplication active)")
    
    def _rech_remove_selected_right(self):
        """Supprimer les lignes sélectionnées dans la table de droite."""
        sel = self.rech_tree.selection()
        if not sel:
            return
        # identifier l'index de numero_serie_batterie pour nettoyer le set
        cols = self.rech_tree.cget("columns")
        try:
            k_idx = cols.index("numero_serie_batterie")
        except ValueError:
            k_idx = 0
        for iid in sel:
            vals = self.rech_tree.item(iid, "values")
            # protéger si vide
            if vals:
                key = str(vals[k_idx])
                if key in self._rech_right_keys:
                    self._rech_right_keys.remove(key)
            self.rech_tree.delete(iid)
    
    def _rech_load_suivi_rows(self, numero_serie_batteries):
        """Retourne (rows, colnames) pour les lignes de suivi_production filtrées par n° série batterie."""
        if not numero_serie_batteries:
            return [], []
        conn = self.db_manager.connect()
        if not conn:
            return [], []
        try:
            cur = conn.cursor()
            # On charge TOUTE la ligne de suivi_production (c’est ce que tu as demandé)
            # IN sécurisé : on fabrique la liste de %s
            placeholders = ", ".join(["%s"] * len(numero_serie_batteries))
            sql = f"SELECT * FROM suivi_production WHERE numero_serie_batterie IN ({placeholders})"
            cur.execute(sql, tuple(numero_serie_batteries))
            rows = cur.fetchall()
            colnames = [d[0] for d in cur.description]
            return rows, colnames
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement suivi_production impossible :\n{e}")
            return [], []
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_configure_tree_for_columns(self, colnames):
        """Configure le Treeview (colonnes, largeur, ancrage, headings) + stretch + scroll horiz/vert déjà mis côté UI."""
        # si colonnes déjà identiques, ne rien faire
        current = self.rech_tree.cget("columns")
        if tuple(current) == tuple(colnames):
            return
    
        # reset
        for c in current:
            try: self.rech_tree.heading(c, text="")
            except: pass
        self.rech_tree.delete(*self.rech_tree.get_children())
        self._rech_right_keys.clear()
    
        self.rech_tree["columns"] = colnames
        # headings & options
        for c in colnames:
            self.rech_tree.heading(c, text=c)
            # largeur heuristique : min 100, sinon ~len*9 px
            width = max(100, min(380, int(len(c) * 9)))
            self.rech_tree.column(c, width=width, minwidth=80, stretch=True, anchor="w")

    
    #------------------------------ Onglet expedition -----------------------------------------------   

    def setup_exp(self, frame):
        # ---------- Layout principal : gauche | droite ----------
        container = ttk.Frame(frame); container.pack(fill="both", expand=True, padx=12, pady=12)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(0, weight=1)
    
        # ========== Colonne gauche : entrées & combos ==========
        left = ttk.LabelFrame(container, text="Expédition", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
        
        ttk.Label(left,text="Choisir le mode d'emballage").pack(pady=(8,4), anchor="w")
        
        cb_var_mode=tk.StringVar()
        self.ex_mode_combo = ttk.Combobox(
            left,
            textvariable=cb_var_mode,
            values=["Numero serie batterie","Numero emballage"],
            width=40)
        self.ex_mode_combo.pack(fill="x")
    
        # N° série produit + ajout auto si 9 chiffres
        ttk.Label(left, text="N° série produit ou N° emballage:").pack(pady=(0,4), anchor="w")
        self.exp_numero_serie_batt_entry = ttk.Entry(left, width=30)
        self.exp_numero_serie_batt_entry.pack(pady=(0,8), fill="x")
        self.exp_numero_serie_batt_entry.bind("<KeyRelease>", self._exp_on_entry_change)
        
        ttk.Label(left,text="Réference batterie:").pack(pady=(8,4), anchor="w")
        valeur_ref=self.models
        valeur_ref.append("Toutes")
        valeur_ref_all=valeur_ref
        cb_var_ex=tk.StringVar()
        self.ex_ref_combo = ttk.Combobox(
            left,
            textvariable=cb_var_ex,
            values=valeur_ref_all,
            width=40)
        self.ex_ref_combo.pack(fill="x")
        self.ex_ref_combo.bind("<<ComboboxSelected>>", self.ex_on_select_ref)
    
        #ttk.Button(left, text="❌ Non conforme", command=self.add_non_conf_batterie, style="Danger.TButton").pack(pady=6, anchor="w")
    
        # Client
        ttk.Label(left, text="Client").pack(pady=(8,4), anchor="w")
        cb_var_client = tk.StringVar()
        self.cb_cl = ttk.Combobox(left, textvariable=cb_var_client, state="readonly", width=40)
        self.cb_cl.pack(fill="x")
    
        # Alimentation des clients
        conn = self.db_manager.connect()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT nom_client FROM client ORDER BY nom_client")
                mots = [row[0].replace(" ", "-") for row in cursor.fetchall()]
                self.cb_cl["values"] = mots
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Chargement clients :\n{e}")
            finally:
                try: cursor.close()
                except: pass
                conn.close()
    
        # Projet
        ttk.Label(left, text="Projet").pack(pady=(8,4), anchor="w")
        self.cb_pr = ttk.Combobox(left, state="readonly", width=40)
        self.cb_pr.pack(fill="x")
        
        # Alimentation des clients
        conn = self.db_manager.connect()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT nom_projet FROM projet ORDER BY nom_projet")
                mots = [row[0] for row in cursor.fetchall()]
                self.cb_pr["values"] = mots
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Chargement projet :\n{e}")
            finally:
                try: cursor.close()
                except: pass
                conn.close()
    
        # Commentaire
        ttk.Label(left, text="Commentaire").pack(pady=(8,4), anchor="w")
        self.exp_comm_entry = ttk.Entry(left)
        self.exp_comm_entry.pack(fill="x")
    
        # ========= Colonne droite : 2 Listbox (disponibles | sélectionnées) =========
        right = ttk.Frame(container)
        right.grid(row=0, column=1, sticky="nsew", padx=(8,0))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)
        right.rowconfigure(4, weight=1)
    
        # Liste des batteries disponibles (emballées)
        ttk.Label(right, text="Batteries emballées (disponibles)").grid(row=0, column=0, sticky="w", pady=(0,4))
        self.exp_listbox_batt = tk.Listbox(right, font=('Segoe UI', 11), height=10)
        self.exp_listbox_batt.grid(row=1, column=0, sticky="nsew")
        self.exp_listbox_batt.bind("<<ListboxSelect>>", lambda e: None)  # neutre
        self.exp_listbox_batt.bind("<Double-Button-1>", self._exp_on_available_double_click)
    
        ttk.Separator(right, orient="horizontal").grid(row=2, column=0, sticky="ew", pady=10)
    
        # Sélection (avec compteur)
        header_sel = ttk.Frame(right); header_sel.grid(row=3, column=0, sticky="ew", pady=(0,4))
        ttk.Label(header_sel, text="Batteries sélectionnées").pack(side="left")
        ttk.Label(header_sel, text="Quantité:").pack(side="right")
        self._exp_count_var = tk.IntVar(value=0)
        self._exp_count_lbl = ttk.Label(header_sel, textvariable=self._exp_count_var)
        self._exp_count_lbl.pack(side="right", padx=(0,8))
    
        self.send_listbox_batt = ttk.Treeview(
            right,
            columns=("num_b","num_emb"),
            show="headings",      # n'affiche pas la colonne d'arborescence, juste les valeurs
            selectmode="browse",  # comme la Listbox d'origine : une seule sélection à la fois
            height=10
        )
        # En-têtes
        self.send_listbox_batt.heading("num_b", text="Numero serie batterie")
        self.send_listbox_batt.heading("num_emb", text="Num emballage")
        
        # Colonnes (largeur + alignement)
        self.send_listbox_batt.column("num_b", anchor="w", width=220)
        self.send_listbox_batt.column("num_emb", anchor="center", width=120)
        
        # Scrollbar verticale
        vsb = ttk.Scrollbar(right, orient="vertical", command=self.send_listbox_batt.yview)
        self.send_listbox_batt.configure(yscrollcommand=vsb.set)
        
        # Placement dans la grille (ajuster si tu as déjà des colonnes)
        self.send_listbox_batt.grid(row=4, column=0, sticky="nsew")
        vsb.grid(row=4, column=1, sticky="ns")
        
        # Bind du double-clic pour suppression
        self.send_listbox_batt.bind("<Double-Button-1>", self._exp_on_selected_double_click)
    
        ttk.Button(right, text="✅ Contrôle OK", command=self.valider_exp, style="Good.TButton").grid(
            row=5, column=0, pady=10, sticky="e"
        )
    
        # Charge la liste des disponibles
        #self.display_model_list_exp()
        
    
    def ex_on_select_ref(self,event=None):
        
        self.exp_listbox_batt.delete(0, tk.END)
        mode=self.ex_mode_combo.get()
        if not mode:
            messagebox.showerror("Pas de mode","Veuillez selectionner un mode")
         
        ref=self.ex_ref_combo.get()    
        conn=self.db_manager.connect() 
        cursor=conn.cursor()
        list_for_box=[]
        if mode == "Numero serie batterie":
            if ref=="Toutes":
                query="select numero_serie_batterie where emballage = 1 and (expedition = 0 or expedition is null) and sp.recyclage is null "
                cursor.execute(query)        
            else :
                query="Select pv.numero_serie_produit from produit_voltr as pv join suivi_production as sp on pv.numero_serie_produit=sp.numero_serie_batterie where pv.reference_produit_voltr =%s and (sp.expedition = 0 or sp.expedition is null) and sp.recyclage is null"
                param=(ref,)
                cursor.execute(query,param)
            
            rows=cursor.fetchall()
            if not rows:
                return
            for row in rows :
                self.exp_listbox_batt.insert(tk.END,row[0])
                
        elif mode == "Numero emballage":
            if ref=="Toutes":
                query="SELECT distinct(num_emballage) FROM suivi_production where num_emballage is not null and (sp.expedition = 0 or sp.expedition is null) and sp.recyclage is null "
                cursor.execute(query)        
            else :
                query="SELECT distinct sp.num_emballage from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie=pv.numero_serie_produit where pv.reference_produit_voltr=%s and num_emballage is not null and (sp.expedition = 0 or sp.expedition is null) and sp.recyclage is null"
                param=(ref,)
                cursor.execute(query,param)
            
            rows=cursor.fetchall()
            if not rows:
                return
            for row in rows :
                self.exp_listbox_batt.insert(tk.END,row[0])
                    
                
                
            query="Select pv.numero_serie_produit from produit_voltr as pv join suivi_production as sp on pv.numero_serie_produit=sp.numero_serie_batterie where pv.reference_produit_voltr =%s"
            #param=
        
        elif mode=="Numero emballage":
            query="SELECT distinct(num_emballage) FROM suivi_production where num_emballage is not null"
            
    def add_non_conf_batterie(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if reponse:
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 

        num_batt=str(self.numero_serie_batt_entry.get())
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension_fail = picking_tension_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.numero_serie_batt_entry.delete(0, tk.END)
    
    """
    def display_model_list_exp(self):
        
        conn = self.db_manager.connect()
        stage_act='exp'
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query=self.build_stage_query(stage_act)
            cursor.execute(query)
            rows = cursor.fetchall()
            liste_batteries = [str(r[0]) for r in rows]
            self.exp_listbox_batt.delete(0, tk.END)
            for batt in liste_batteries:
                self.exp_listbox_batt.insert(tk.END, batt)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try: cursor.close()
            except: pass
            conn.close()
    """   
    #------------------------------ Onglet recyclage -----------------------------------------------   
    def setup_recyclage(self,frame):

        left = ttk.Frame(frame)
        left.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        
        right = ttk.Frame(frame)
        right.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        
        ttk.Label(left, text="N° série cellule").pack(pady=5)
        self.r_entry_cell = ttk.Entry(left, width=28)
        self.r_entry_cell.pack(pady=5)
        # Remplissage auto du n° batterie quand l'entry cellule atteint 12 chars
        self.r_entry_cell.bind("<KeyRelease>", self.r_on_cell_entry)
    
        ttk.Label(left, text="N° série batterie").pack(pady=5)
        self.r_entry_batt = ttk.Entry(left, width=28)
        self.r_entry_batt.pack(pady=5)
    
        ttk.Label(left, text="Référence batterie").pack(pady=5)
        # valeur par défaut nulle (vide)
        self.r_model_var = tk.StringVar(value="")
        self.r_combo = ttk.Combobox(left, textvariable=self.r_model_var,
                                       values=(self.models or []), state="readonly", width=30)
        self.r_combo.pack(pady=5)
        self.r_combo.bind("<<ComboboxSelected>>", lambda e: self.r_on_model_change())
        
        ttk.Label(left, text="Cause recyclage:").pack(pady=5)
        
        self.cause_var = tk.StringVar(value="")
        self.cause_combo = ttk.Combobox(left, textvariable=self.cause_var,
                                       values=("erreur_soudure","choc","autres"), state="readonly", width=30)
        self.cause_combo.pack(pady=5)
        
        ttk.Button(
            left, text="🔄 Recyler la batterie",
            command=self.recycle_batterie_sp, style="Danger.TButton"
        ).pack(pady=10)
    
        ttk.Label(right, text="Liste batterie").pack(pady=5)
        
        self.r_listbox = tk.Listbox(right, height=10, activestyle="dotbox", selectmode="extended")
        yscroll = ttk.Scrollbar(right, orient="vertical", command=self.r_listbox.yview)
        self.r_listbox.configure(yscrollcommand=yscroll.set)
        self.r_listbox.pack(side="right", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        self.r_listbox.bind("<<ListboxSelect>>", self.on_r_listbox_select)
        
    def on_r_listbox_select(self, event):
        # obtenir indices sélectionnés (peut être plusieurs si selectmode="extended")
        sel = self.r_listbox.curselection()
        if not sel:
            return
        # on prend le premier sélectionné
        idx = sel[0]
        text = self.r_listbox.get(idx)
    
        # --- si text est déjà le numero de série simple ---
        # numero = text.strip()
    
        # --- OU : si text contient d'autres champs, extraire le numéro ---
        # Exemples d'extraction (choisis celle qui correspond à ton format)
        # 1) format "123456789012" => direct
        # 2) format "1;123456789012;moduleA" => split par ';' et prendre le 2ème
        # 3) format "1 | 123456789012 | module A" => split par '|' et strip
        numero = None
        if ";" in text:
            parts = [p.strip() for p in text.split(";")]
            # si le numéro est en 2e position
            if len(parts) >= 2:
                numero = parts[1]
        elif "|" in text:
            parts = [p.strip() for p in text.split("|")]
            # chercher la première partie qui ressemble à un n° (ex: longueur 12, chiffres)
            for p in parts:
                if p and any(ch.isdigit() for ch in p):
                    numero = p
                    break
        else:
            # par défaut on prend toute la chaîne
            numero = text.strip()
    
        # si tu utilises une StringVar pour l'entry, mets-la ; sinon delete/insert
        try:
            self.r_entry_batt.delete(0, "end")
            if numero:
                self.r_entry_batt.insert(0, numero)
        except Exception as e:
            # fallback si self.r_entry_batt est une StringVar
            if hasattr(self, "r_model_var") and isinstance(self.r_model_var, tk.StringVar):
                self.r_model_var.set(numero or "")
            else:
                raise
    
    def r_on_cell_entry(self, event=None):
        """Quand l'entry cellule atteint 12 chars, on cherche la batterie associée et on remplit l'entry batterie."""
        numero_serie_cell = self.r_entry_cell.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # affectation_produit = numero_serie_batterie (selon ta logique existante)
            cur.execute("SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s", (numero_serie_cell,))
            row = cur.fetchone()
            self.r_entry_batt.delete(0, tk.END)
            if row and row[0]:
                self.rech_entry_batt.insert(0, str(row[0]))
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Lookup cellule→batterie impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
        
    def r_on_model_change(self):
        """Quand on choisit un modèle, on alimente la liste des n° batteries via la jointure demandée."""
        ref = self.r_model_var.get().strip()
        self.r_listbox.delete(0, tk.END)
        if not ref:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # Liste des NUMÉROS DE SÉRIE BATTERIE pour la référence choisie
            # sp = suivi_production / p = produit
            sql = ("""
                SELECT DISTINCT sp.numero_serie_batterie
                FROM suivi_production sp
                JOIN produit_voltr p
                  ON sp.numero_serie_batterie = p.numero_serie_produit
                WHERE p.reference_produit_voltr = %s
                AND recyclage is null 
                ORDER BY sp.numero_serie_batterie
            """)
            cur.execute(sql, (ref,))
            for (num_batt,) in cur.fetchall():
                if num_batt:
                    self.r_listbox.insert(tk.END, str(num_batt))
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement liste batteries impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def recycle_batterie_sp(self):
        type_obj = "batterie"
        cause = self.cause_combo.get()
        if not cause:
            messagebox.showerror("Pas de cause", "Veuillez sélectionner une cause !")
            return
    
        numero_serie_batt = self.r_entry_batt.get().strip()
        if not numero_serie_batt:
            messagebox.showerror("Erreur", "Veuillez renseigner le numéro de série de la batterie.")
            return
    
        conn = None
        cursor = None
        try:
            conn = self.db_manager.connect()
            cursor = conn.cursor()
    
            # 1) Récupérer référence et poids de la batterie
            query = "SELECT pv.reference_produit_voltr, rv.poids FROM produit_voltr as pv join ref_batterie_voltr as rv on pv.reference_produit_voltr=rv.reference_batterie_voltr WHERE numero_serie_produit = %s"
            cursor.execute(query, (numero_serie_batt,))
            row_prod = cursor.fetchone()
            if row_prod is None:
                messagebox.showerror("Non trouvé", f"Aucune batterie trouvée pour le n° {numero_serie_batt}")
                return
    
            reference_batt = row_prod[0]
            poids_batt = row_prod[1] or 0
    
            # 2) Lire la feuille Excel et trouver le dest_recyclage
            df_cyclage = pd.read_excel(EXCEL_PATH, sheet_name="Cyclage", header=1)
            sel = df_cyclage[df_cyclage["Nom_modele"] == reference_batt]
            if sel.empty:
                messagebox.showerror("Erreur Excel", f"Modèle {reference_batt} introuvable dans {EXCEL_PATH} sheet Cyclage.")
                return
            seuils = sel.iloc[0]
            dest_recyclage = str(seuils.get("Recyclage", "")).strip()
            if not dest_recyclage:
                messagebox.showerror("Erreur", f"Pas de destination de recyclage définie pour {reference_batt}.")
                return
            type_fut = dest_recyclage
    
            # 3) Chercher un fut ouvert pour ce type
            cursor.execute(
                "SELECT id_fut, poids FROM fut_recyclage WHERE exutoire = %s AND etat_fut = %s LIMIT 1",
                (type_fut, "en cours")
            )
            fut_row = cursor.fetchone()
            if fut_row is None:
                messagebox.showerror("Erreur !", "Aucun fut d'exutoire eo_org_mtl n'est ouvert")
                return
    
            id_fut, poids_fut = fut_row[0], fut_row[1] or 0
    
            # 4) Mettre à jour le poids du fut
            poids_tot = poids_fut + poids_batt
            cursor.execute("UPDATE fut_recyclage SET poids = %s WHERE id_fut = %s", (poids_tot, id_fut))
    
            # 5) Insérer la ligne de recyclage (remarquer VALUES (...) et les colonnes explicitement)
            query_recy = """
                INSERT INTO recyclage
                    (numero_serie, type_objet, id_fut, sur_site, date_rebut, cause)
                VALUES (%s, %s, %s, %s, NOW(), %s)
            """
            param_recy = (numero_serie_batt, type_obj, id_fut, "oui", cause)
            cursor.execute(query_recy, param_recy)
    
            # 6) Mettre à jour le suivi_production
            query_sp = "UPDATE suivi_production SET recyclage = 1, date_recyclage = NOW() WHERE numero_serie_batterie = %s"
            cursor.execute(query_sp, (numero_serie_batt,))
    
            # 7) Commit et message utilisateur
            conn.commit()
            emplacement = f"fut {id_fut}"
            messagebox.showinfo("Recyclage réussi",
                                f"La batterie {numero_serie_batt} recyclée dans un fut {type_fut} : {emplacement}")
    
        except Exception as e:
            # rollback en cas d'erreur et message
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
            messagebox.showerror("Erreur BDD", f"Erreur lors du recyclage : {e}")
        finally:
            if cursor:
                try:
                    cursor.close()
                except Exception:
                    pass
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
                
    # ===================== Handlers & utilitaires =====================
    
    def _exp_on_entry_change(self, event):
        """Ajoute la batterie dans 'sélectionnées' quand l'entrée atteint 9 chiffres."""
        mode=self.ex_mode_combo.get()
        txt = self.exp_numero_serie_batt_entry.get().strip()
        
        if not mode:
            messagebox.showerror("Pas de mode", "veuillez selectionner un mode")
            return
        
        if mode == 'Numero serie batterie':
            if len(txt) == 9:
                self._exp_add_to_selection(txt)
                self.exp_numero_serie_batt_entry.delete(0, tk.END)
        
        elif mode =='Numero emballage':
            if len(txt) == 17 :
                self._exp_add_to_selection(txt)
                self.exp_numero_serie_batt_entry.delete(0, tk.END)
                
                
    def _exp_on_available_double_click(self, event):
        """Double-clic sur une batterie disponible -> ajoute à la sélection."""
        sel = self.exp_listbox_batt.curselection()
        if not sel:
            return
        value = self.exp_listbox_batt.get(sel[0])
        self._exp_add_to_selection(value)
    
    def _exp_on_selected_double_click(self, event):
        """Double-clic sur une batterie sélectionnée -> la retire de la sélection."""
        sel = self.send_listbox_batt.selection()
        if not sel:
            return
        item_id = sel[0]
        # récupérer la valeur (première colonne)
        vals = self.send_listbox_batt.item(item_id, "values")
        if vals:
            value = vals[0]
        else:
            value = None
    
        # supprimer l'item
        self.send_listbox_batt.delete(item_id)
    
        # met à jour le compteur (ta fonction existante)
        try:
            self._exp_update_counter()
        except AttributeError:
            # Si tu n'as pas encore _exp_update_counter, on met à jour le compteur ici
            count = int(self.send_listbox_batt.index('end'))
            self._exp_count_var.set(count)
    
    def _exp_add_to_selection(self, numero):
        """Ajoute sans doublon à la listbox sélectionnée puis met à jour le compteur."""
        # Anti-doublon
        mode=self.ex_mode_combo.get()
        if mode=="Numero emballage":
            conn=self.db_manager.connect()
            cursor=conn.cursor()
            query="select numero_serie_batterie from suivi_production where num_emballage= %s"
            param=(numero,)
            cursor.execute(query,param)
            rows=cursor.fetchall()
            cursor.close()
            conn.close()
            
            for row in rows:
                numero_=row[0]
                current = set([self.send_listbox_batt.item(item)['values'][0] for item in self.send_listbox_batt.get_children()])
                if numero_ in current:
                    return
                self.send_listbox_batt.insert("", "end", values=(numero_, numero))
                self._exp_update_counter()
                
        elif mode =="Numero serie batterie":
            
            
            current = set([self.send_listbox_batt.item(item)['values'][0] for item in self.send_listbox_batt.get_children()])
            if numero in current:
                return
            self.send_listbox_batt.insert("", "end", values=(numero, "None"))
            self._exp_update_counter()
    
    def _exp_update_counter(self):
        """Mise à jour du compteur de batteries sélectionnées."""
        self._exp_count_var.set(len(self.send_listbox_batt.get_children()))
        
    def _exp_get_selected_batteries(self): 
        """Retourne toutes les batteries présentes dans la listbox 'sélectionnées'."""
        return list(self.send_listbox_batt.get(0, tk.END))
    
    def exp_add_non_conf_batterie(self):
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
                
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from)  
        
        self.exp_numero_serie_batt_entry.delete(0, tk.END)
        
        
    def valider_exp(self):
        
        # 0) Récup sélection
        #numeros = list(self.send_listbox_batt.get(0, tk.END))
        numeros=[self.send_listbox_batt.item(item)['values'][0] for item in self.send_listbox_batt.get_children()]
        if not numeros:
            messagebox.showwarning("Avertissement", "Aucune batterie sélectionnée.")
            return
        
        # 2) Lecture des références actuelles via IN (...)
        placeholders = ", ".join(["%s"] * len(numeros))
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
        
            
            comment = self.exp_comm_entry.get().strip()
            placeholders_final = ", ".join(["%s"] * len(numeros))
            projet=self.cb_pr.get().strip()
            nom_client=self.cb_cl.get().strip()
            
            conn2 = self.db_manager.connect()
            if conn2:
                try:
                    cursor = conn2.cursor()
                    cursor.execute("SELECT nom_client FROM client ORDER BY nom_client")
                    mots = [row[0].replace(" ", "-") for row in cursor.fetchall()]
                    self.cb_cl["values"] = mots
                except Exception as e:
                    messagebox.showerror("Erreur SQL", f"Chargement clients :\n{e}")
                finally:
                    try: cursor.close()
                    except: pass
                    conn2.close()
            
            #Obtenir id_client
            if nom_client:
                nom_c_bdd=nom_client.replace("-"," ")
                conn2 = self.db_manager.connect()
                if conn2:
                    try:
                        cursor = conn2.cursor()
                        cursor.execute("SELECT id_client FROM client where nom_client = %s",(nom_c_bdd,))
                        id_client = [row[0] for row in cursor.fetchall()]
                    except Exception as e:
                        messagebox.showerror("Erreur SQL", f"Chargement id_client :\n{e}")
                    finally:
                        try: cursor.close()
                        except: pass
                        conn2.close()
    
            if comment:
                
                sql_sp = f"""
                    UPDATE suivi_production
                    SET expedition = 1,
                        date_expedition = NOW(),
                        commentaire = %s
                    WHERE numero_serie_batterie IN ({placeholders_final})
                """
                params_sp = (comment, *numeros)
                
            else:
                
                sql_sp = f"""
                    UPDATE suivi_production
                    SET expedition = 1,
                        date_expedition = NOW()
                    WHERE numero_serie_batterie IN ({placeholders_final})
                """
                params_sp = tuple(numeros)
    
            cur.execute(sql_sp, params_sp)
            
            set_parts = ["statut = 'expediee'"]
            params_mark = []
            
            if nom_client:
                if id_client:
                    set_parts.append("id_client = %s")
                    params_mark.append(id_client[0])
            
            if projet:
                set_parts.append("numero_projet = %s")
                params_mark.append(projet)
            
            sql_mark = f"""
                UPDATE produit_voltr
                SET {', '.join(set_parts)}
                WHERE numero_serie_produit IN ({placeholders_final})
            """
            
            params_mark = tuple(params_mark) + tuple(numeros)
            
            cur.execute(sql_mark, params_mark)
            
            conn.commit()
    
            # 10) UI : succès + refresh
            messagebox.showinfo("Succès", f"{len(numeros)} batterie(s) marquées expédiées.")
            self.send_listbox_batt.delete(*self.send_listbox_batt.get_children())      # vide la sélection
            self._exp_update_counter()                      # remet le compteur à jour          # recharge les "disponibles"
            self.exp_comm_entry.delete(0, tk.END)           # vide le commentaire
            

        except Exception as e:
            try:
                conn.rollback()
            except:
                pass
            messagebox.showerror("Erreur SQL", f"valider_exp :\n{e}")
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass
        
if __name__ == "__main__":
    app = StockApp()
    app.mainloop()
