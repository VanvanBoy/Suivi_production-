# -*- coding: utf-8 -*-
"""
Created on Thu Aug 21 15:18:57 2025

@author: User
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os, re, shutil
import tkinter as tk
from tkinter import ttk, messagebox
from ttkthemes import ThemedTk
from PIL import Image, ImageTk
import pandas as pd
import mysql.connector
import webbrowser
from datetime import datetime

EXCEL_PATH = r"G:\Drive partagés\VoltR\11_Data\IHM\Suivi_prod_par_modele.xlsx"  

class DBManager:
    def __init__(self):
        
        self.config = {
            'user': 'Vanvan',
            'password': 'VoltR99!',
            'host': '34.77.226.40',
            'database': 'bdd_29082025',
            'auth_plugin': 'mysql_native_password'
        }
        
        """
        self.config = {
            'user': 'Vanvan',
            'password': 'VoltR99!',
            'host': '34.77.226.40',
            'database': 'cellules_batteries_cloud',
            'auth_plugin': 'mysql_native_password'
        }
        
        
        self.config = {
            'user': 'root',
            'password': 'VoltR99!',
            'host': '127.0.0.1',
            'database': 'bdd_29072025',
            'auth_plugin': 'mysql_native_password'
        }
        
        """
        
    def connect(self):
        try:
            return mysql.connector.connect(**self.config)
        except mysql.connector.Error as err:
            messagebox.showerror("Erreur DB", f"Erreur lors de la connexion : {err}")
            return None

class StockApp(ThemedTk):

    
    def __init__(self):
        super().__init__(theme="xpnative")
        self.title("Suivi de production")
        self.geometry("1000x650")

        self.db_manager = DBManager()
        self.picking_file_path = None

        # styles
        self.style = ttk.Style()
        self.style.configure("TLabel", font=('Segoe UI', 11), padding=4)
        self.style.configure("TButton", font=('Segoe UI', 11), padding=6)
        self.style.configure("TEntry", font=('Segoe UI', 11))
        self.style.configure("TCombobox", font=('Segoe UI', 11))
        self.style.configure("Danger.TButton", foreground="black", background="#F06F65")
        self.style.map("Danger.TButton", background=[('active', '#e55b50')], foreground=[('disabled', 'black')])
        self.style.configure("Good.TButton", foreground="black", background="#0ED329")
        self.style.map("Good.TButton", background=[('active', '#0ED329')], foreground=[('disabled', 'black')])

        
        self.selected_model = None
        self.stage_order = None  

        self._show_model_selector_and_build()
        
        
        self.STAGE_TO_DBCOL = {
            "picking":   "picking_tension",          
            "pack":      "soudure_pack",     
            "nappe":     "soudure_nappe",
            "bms":       "soudure_bms",
            "wrap":      "wrap",
            "fermeture": "fermeture",
            "capa":      "test_capa",        
            "emb":       "emballage",        
            "exp":       "expedition",       
            "recherche": "recherche",
        }
        
        self.ALLOWED_STAGE_KEYS = set(self.STAGE_TO_DBCOL.keys())  # pour sécuriser
    
    def _required_previous_dbcols(self,current_stage: str):
        """
        Retourne la liste des NOMS DE COLONNES MySQL (dans suivi_production)
        à vérifier (=1) avant de valider current_stage.
    
        On s'appuie sur self.stage_order (rangs > 0 = étape active)
        et sur STAGE_TO_DBCOL pour la correspondance.
        """
        if not hasattr(self, "stage_order") or not self.stage_order:
            return []
    
        if current_stage not in self.ALLOWED_STAGE_KEYS:
            return []
    
        cur_rank = self.stage_order.get(current_stage, 0)
        if cur_rank <= 0:
            return []
    
        prev_pairs = []
        for k, v in self.stage_order.items():
            if k in self.ALLOWED_STAGE_KEYS and v > 0 and v < cur_rank:
                dbcol = self.STAGE_TO_DBCOL.get(k)
                if dbcol:
                    prev_pairs.append((k, v, dbcol))
    
        # tri pour lisibilité/diagnostic
        prev_pairs.sort(key=lambda x: x[1])
    
        # on retourne uniquement les colonnes DB
        return [dbcol for (_, _, dbcol) in prev_pairs]
    
    
    def _check_prereqs_and_warn(self, num_batt: str, current_stage: str) -> bool:
        """
        Vérifie dans suivi_production que TOUTES les colonnes prérequis (=1)
        sont validées pour num_batt, d'après le mapping STAGE_TO_DBCOL.
        """
        prev_dbcols = self._required_previous_dbcols(current_stage)
        if not prev_dbcols:
            return True  # pas de prérequis
    
        conn = self.db_manager.connect()
        if not conn:
            return False
        try:
            cursor = conn.cursor(dictionary=True)
    
            # Sécurisation : on n'insère que des noms de colonnes issus de STAGE_TO_DBCOL
            cols_sql = ", ".join(prev_dbcols)
            sql = f"""
                SELECT {cols_sql}
                FROM suivi_production
                WHERE numero_serie_batterie = %s
                LIMIT 1
            """
            cursor.execute(sql, (num_batt,))
            row = cursor.fetchone()
            if not row:
                messagebox.showwarning("Vérification", "Sélectionner une batterie")
                return False
    
            missing = [c for c in prev_dbcols if (row.get(c) or 0) != 1]
            if missing:
                lis = ", ".join(missing)
                messagebox.showwarning(
                    "Pré-requis manquants",
                    f"Impossible de valider '{current_stage}'. Étape(s) non validée(s) : {lis}"
                )
                return False
    
            return True
        except Exception as e:
            messagebox.showerror("SQL", f"Erreur vérification prérequis:\n{e}")
            return False
        finally:
            try:
                cursor.close(); conn.close()
            except:
                pass
                
    def _show_model_selector_and_build(self):
        try:
            df = pd.read_excel(EXCEL_PATH,sheet_name="Flux test modele")
        except Exception as e:
            messagebox.showerror("Excel", f"Impossible de lire l'Excel:\n{e}")
            self.destroy()
            return

        if "nom_modele" not in df.columns:
            messagebox.showerror("Excel", "La colonne 'nom_modele' est absente du fichier.")
            self.destroy()
            return

        self.models = df["nom_modele"].dropna().astype(str).unique().tolist()
        if not self.models:
            messagebox.showerror("Excel", "Aucune valeur dans 'nom_modele'.")
            self.destroy()
            return

        dlg = tk.Toplevel(self)
        dlg.title("Choisir la référence batterie")
        dlg.geometry("420x160")
        dlg.transient(self)
        dlg.grab_set()

        ttk.Label(dlg, text="Référence batterie :").pack(pady=(18, 6))
        cb_var = tk.StringVar()
        cb = ttk.Combobox(dlg, textvariable=cb_var, values=self.models, state="readonly", width=40)
        cb.pack()
        cb.current(0)

        def on_launch():
            
            ref = cb_var.get()
            row = df.loc[df["nom_modele"].astype(str) == ref]
            if row.empty:
                messagebox.showerror("Sélection", "Référence introuvable.")
                return

            column_to_stage = {
                "picking": "picking",
                "soudure_pack": "pack",
                "soudure_nappe": "nappe",
                "soudure_bms": "bms",
                "wrap": "wrap",
                "fermeture": "fermeture_batt",
                "test_capa": "capa",
                "emballage": "emb",
                "expedition": "exp",
                "recherche": "recherche",
            }

            stage_order = {}
            for col, key in column_to_stage.items():
                if col in df.columns:
                    try:
                        val = int(row.iloc[0][col])
                    except Exception:
                        # NaN / texte -> traite comme 0
                        val = 0
                    stage_order[key] = val

            self.selected_model = ref
            self.stage_order = stage_order
            dlg.destroy()

        ttk.Button(dlg, text="Lancer l'application", style="Good.TButton", command=on_launch).pack(pady=16)

        # centre la boîte
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 210
        y = self.winfo_y() + (self.winfo_height() // 2) - 80
        dlg.geometry(f"+{x}+{y}")

        self.wait_window(dlg)

        if self.stage_order is None:
            
            self.destroy()
            return

        
        self._create_widgets_with_order()


    def _create_widgets_with_order(self):
        self.title("Suivi de production - "+str(self.selected_model))
        stage_defs = {
            "picking":   ("Contrôle de picking", self.setup_picking),
            "pack":      ("Contrôle soudure pack", self.setup_pack),
            "nappe":     ("Contrôle soudure nappe", self.setup_nappe),
            "bms":       ("Contrôle soudure BMS", self.setup_bms),
            "wrap":      ("Contrôle wrap", self.setup_wrap),
            "fermeture_batt": ("Contrôle fermeture", self.setup_fermeture),
            "capa":      ("Test de capacité", self.setup_capa),
            "emb":       ("Contrôle emballage", self.setup_emb),
            "exp":       ("Contrôle expédition", self.setup_exp),
            "recherche": ("Recherche de batterie", self.setup_recherche),
        }

        
        ordered_keys = [
            k for k, v in sorted(self.stage_order.items(), key=lambda kv: kv[1])
            if v and v > 0 and k in stage_defs
        ]

        if not ordered_keys:
            messagebox.showwarning("Configuration", "Aucun onglet actif pour cette référence.")
          
            ordered_keys = ["picking"]

        notebook = ttk.Notebook(self)
        notebook.pack(expand=True, fill="both")

        for key in ordered_keys:
            title, setup_fn = stage_defs[key]
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=title)
            setup_fn(frame)

    
    def set_photo(self, label: tk.Label, chemin_image: str, size=(200, 200)):
        """Charge une image et l'affecte AU SEUL label donné."""
        try:
            img = Image.open(chemin_image)
            img = img.resize(size)
            photo = ImageTk.PhotoImage(img)
            label.config(image=photo, text="")
            label.image = photo  # garder une référence pour éviter GC
        except Exception as e:
            messagebox.showerror("Erreur image", f"Impossible de charger l'image : {e}")
            
    def verfier_coherence_ref(self,num_batt):
        modele=self.selected_model
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "Select reference_produit_voltr from produit_voltr where numero_serie_produit = %s "
            param = (num_batt,)
            cursor.execute(query, param)
            modele_act=cursor.fetchone()[0]
            if modele_act != modele:
                reponse=messagebox.askyesno("Modèle de batterie différent",f"Le mdodèle de batterie n'est pas coherent \n Passer du modèle {modele_act} au modèle {modele} pour la batterie {num_batt} ?")
                if not reponse:
                    return 'stop'
                else :
                    cursor.execute("UPDATE produit_voltr SET reference_produit_voltr = %s WHERE numero_serie_produit =%s",(modele,num_batt))
                    messagebox.showinfo("Nouveau modèle", f'la batterie {num_batt} est passé au modèle {modele}.')
                    conn.commit()
                    return 'next'
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()  
    
    def changer_ref_batterie(self,new_ref,num_batt):
        modele=self.selected_model
        if modele == new_ref:
            messagebox.showinfo("Réference identique","Le nouveau modele est identique au modele actuel")
            return 
        else :
            conn = self.db_manager.connect()
            if not conn:
                return
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE produit_voltr SET reference_produit_voltr = %s WHERE numero_serie_produit =%s",(new_ref,num_batt))
                messagebox.showinfo("Nouveau modèle", f'la batterie {num_batt} est passé au modèle {new_ref}.')
                conn.commit()
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
            finally:
                try:
                    cursor.close()
                except:
                    pass
                conn.close()  
                
    #------------------------------ Onglet picking -----------------------------------------------
    
    #Front

    def setup_picking(self, frame):
        
        
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.numero_serie_cell_entry = ttk.Entry(left_frame); self.numero_serie_cell_entry.pack(pady=5)
        self.numero_serie_cell_entry.bind("<KeyRelease>",self.check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.numero_serie_batt_entry = ttk.Entry(left_frame); self.numero_serie_batt_entry.pack(pady=5)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        
        self.listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.listbox_batt.pack(pady=5, fill='both')
        
        self.listbox_batt.bind("<<ListboxSelect>>", self.on_select_batt)
        
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="🔁 Remplacement cellule:").pack(pady=5)
        ttk.Label(right_frame, text="N° série cellule:").pack(pady=5)
        self.numero_cell_r_entry = ttk.Entry(right_frame); self.numero_cell_r_entry.pack(pady=5)
        ttk.Label(right_frame, text="Défaut:").pack(pady=5)
        self.combobox_default = ttk.Combobox(right_frame, state="readonly", values=["Non trouvée", "Tension", "Corrosion", "Déformation"])
        self.combobox_default.pack(pady=5)
        self.combobox_default.bind("<<ComboboxSelected>>",self.tension_defaut)
        
        ttk.Label(right_frame, text="Tension cellule NOK:").pack(pady=5)
        self.tension_cell_entry=ttk.Entry(right_frame,state="disabled");self.tension_cell_entry.pack(pady=5)

        ttk.Button(right_frame, text="🔄 Demande remplacement cellule", command=self.replace_cellule, style="Danger.TButton").pack(pady=10)

        frame_info = tk.Frame(right_frame, width=300, height=100, bg='#D0F5BE'); frame_info.pack(pady=20); frame_info.pack_propagate(False)
        tk.Label(frame_info, text="⚠ Écart maximum de 0.05V", bg='#D0F5BE', fg='black', font=("Segoe UI", 11, 'bold')).pack(expand=True)

        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_picking, style="Good.TButton").pack(pady=10)
        
        self.display_model_list()
            
    #Back
    
    def display_model_list(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv  on sp.numero_serie_batterie = pv.numero_serie_produit where sp.picking_tension is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
    def on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.numero_serie_batt_entry.delete(0, tk.END)
        self.numero_serie_batt_entry.insert(0, selected_value)
                      
    def add_non_conf_batterie(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension_fail = picking_tension_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.numero_serie_batt_entry.delete(0, tk.END)
            
    def valider_picking(self):
        num_batt=str(self.numero_serie_batt_entry.get())
        if not num_batt:
            return
        directive=self.verfier_coherence_ref(num_batt)
        # 1) Pré-requis
        stage_col='picking'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        if directive=='stop':
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension = 1, date_picking_tension = NOW() where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.numero_serie_batt_entry.delete(0, tk.END)
            self.display_model_list()
             
    def tension_defaut(self,event=None):
        defaut=str(self.combobox_default.get())
        if defaut=="Tension":
            self.tension_cell_entry.config(state="normal")
        else :
            self.tension_cell_entry.config(state="disabled")
                       
    def replace_cellule(self):

        self.remplacement_dir = r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc"
        os.makedirs(self.remplacement_dir, exist_ok=True)
        # Template 
        self.template_remplacement_path = r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\Template remplacement cellule (13).xlsx"
        
        num = self.numero_cell_r_entry.get().strip()  
        defaut = self.combobox_default.get().strip()                  
        tension_txt = self.tension_cell_entry.get().strip()    
        
        if not num:
            messagebox.showwarning("Champ manquant", "Renseigne le N° de série cellule.")
            return
        if not defaut:
            messagebox.showwarning("Champ manquant", "Sélectionne un défaut.")
            return
        if defaut.lower().startswith("tension"):
            if not tension_txt:
                messagebox.showwarning("Champ manquant",
                                       "Défaut = Tension ➜ renseigne la tension NOK.")
                return
            # optionnel: vérifier numérique
            try:
                float(tension_txt.replace(",", "."))
            except ValueError:
                messagebox.showwarning("Valeur invalide",
                                       "La tension doit être un nombre (ex: 3.72).")
                return
    
        projet = None
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
    
            cur.execute("""
                SELECT pv.numero_projet,pv.numero_serie_produit FROM produit_voltr as pv join cellule as c on pv.numero_serie_produit=c.affectation_produit WHERE c.numero_serie_cellule = %s
            """, (num,))
            row = cur.fetchone()
            if row and row[0]:
                projet = str(row[0]).strip()
                produit = str(row[1]).strip()
            else:
                messagebox.showerror("Projet introuvable",
                                     f"Aucun projet trouvé pour la cellule {num}.")
                return
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer le projet :\n{e}")
            return
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass
    
        today_str = datetime.now().strftime("%Y-%m-%d")
        safe_projet = re.sub(r"[^\w\- ]+", "_", projet)  # sécurité nom de fichier
        out_path = os.path.join(self.remplacement_dir, f"{safe_projet}-{today_str}.xlsx")
    
        tension_val = ""
        if defaut.lower().startswith("tension"):
            # conserver le format texte tel que saisi, ou caster en float si tu préfères
            tension_val = tension_txt.replace(",", ".")
    
        new_row = [produit, num, defaut, tension_val]
    
        try:
            if not os.path.exists(out_path):
                if not os.path.exists(self.template_remplacement_path):
                    messagebox.showerror("Template manquant",
                                         f"Template introuvable :\n{self.template_remplacement_path}")
                    return
                shutil.copyfile(self.template_remplacement_path, out_path)
    
            wb = load_workbook(out_path)

            SHEET_CANDIDATES = ["Remplacement", "Remplacements", "Feuil1", "Données"]
            ws = None
            for name in SHEET_CANDIDATES:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:

                expected_headers = ["Date", "N° série", "Défaut", "Tension"]
                for sh in wb.worksheets:
                    headers = [ (sh.cell(row=1, column=i).value or "").strip() for i in range(1, 5) ]
                    if [h.lower() for h in headers] == [h.lower() for h in expected_headers]:
                        ws = sh
                        break
            if ws is None:
              
                ws = wb.active
    
            ws.append(new_row)

            max_row = ws.max_row
            max_col = ws.max_column
            last_col_letter = get_column_letter(max_col)
    
            if hasattr(ws, "tables"):
                for tbl in list(ws.tables.values()):
                    ref = tbl.ref
                    start, end = ref.split(":")
                    start_col = "".join([c for c in start if c.isalpha()])
                    start_row = int("".join([c for c in start if c.isdigit()]))

                    new_ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
                    tbl.ref = new_ref
    
            if ws.auto_filter and ws.auto_filter.ref:
                start, _ = ws.auto_filter.ref.split(":")
                start_col = "".join([c for c in start if c.isalpha()])
                start_row = int("".join([c for c in start if c.isdigit()]))
                ws.auto_filter.ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
    
            wb.save(out_path)
    
        except Exception as e:
            messagebox.showerror("Erreur Excel", f"Impossible d'écrire dans le fichier :\n{e}")
            return

        try:
            self.numero_cell_r_entry.delete(0, 'end')
            self.combobox_default.set('')
            self.tension_cell_entry.delete(0, 'end')
        except Exception:
            pass
    
        messagebox.showinfo("Remplacement enregistré", f"Ligne ajoutée dans :\n{out_path}")
        
        
    
    def check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.numero_serie_batt_entry.delete(0, tk.END)
                self.numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
            
    #------------------------------ Onglet soudure pack -----------------------------------------------    
    
    #Front
    
    def setup_pack(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.s_numero_serie_cell_entry = ttk.Entry(left_frame); self.s_numero_serie_cell_entry.pack(pady=5)
        self.s_numero_serie_cell_entry.bind("<KeyRelease>",self.s_check_entry_length)
        
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.s_numero_serie_batt_entry = ttk.Entry(left_frame); self.s_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_pack, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.s_listbox_batt = tk.Listbox(right_frame, font=('Segoe UI', 11), height=10); self.s_listbox_batt.pack(pady=5, fill='both')
        self.s_listbox_batt.bind("<<ListboxSelect>>", self.s_on_select_batt)
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_soudure_pack, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_pack()
            
    #Back
    
    def s_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.s_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.s_numero_serie_batt_entry.delete(0, tk.END)
                self.s_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.s_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_pack(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.soudure_pack is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.s_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.s_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_soudure_pack(self):
        num_batt=str(self.s_numero_serie_batt_entry.get())
        # 1) Pré-requis
        stage_col='pack'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack = 1, date_soudure_pack = NOW() where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.s_numero_serie_batt_entry.delete(0, tk.END)
            self.display_model_list_pack()
            
            
    def add_non_conf_batterie_pack(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.s_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack_fail = soudure_pack_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.s_numero_serie_batt_entry.delete(0, tk.END)
    
    def s_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.s_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.s_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.s_numero_serie_batt_entry.delete(0, tk.END)
        self.s_numero_serie_batt_entry.insert(0, selected_value)
            
    
        
    #------------------------------ Onglet soudure nappe -----------------------------------------------   

    def setup_nappe(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.n_numero_serie_cell_entry = ttk.Entry(left_frame); self.n_numero_serie_cell_entry.pack(pady=5)
        self.n_numero_serie_cell_entry.bind("<KeyRelease>",self.n_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.n_numero_serie_batt_entry = ttk.Entry(left_frame); self.n_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_nappe, style="Danger.TButton").pack(pady=10)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.n_listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.n_listbox_batt.pack(pady=5, fill='both')
        self.n_listbox_batt.bind("<<ListboxSelect>>", self.n_on_select_batt)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Ecart tension modules:").pack(pady=5)
        self.ecart_t_entry = ttk.Entry(right_frame); self.ecart_t_entry.pack(pady=5)
        self.n_label_photo = tk.Label(right_frame, bg="#e0e0e0", width=200, height=200, text="Aperçu photo", anchor='center')
        self.n_label_photo.pack(pady=10)
        self.set_photo(self.n_label_photo, r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\voltr_logo.jpg")
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_soudure_nappe, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_nappe()
        
        #Back
        
    def n_check_entry_length(self, event=None):

        numero_serie_cell = self.n_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.n_numero_serie_batt_entry.delete(0, tk.END)
                self.n_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.n_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_nappe(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.soudure_nappe is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.n_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.n_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_soudure_nappe(self):
        num_batt=str(self.n_numero_serie_batt_entry.get())
        # 1) Pré-requis
        stage_col='nappe'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        if self.ecart_t_entry.get():
            delta_tension_str=self.ecart_t_entry.get()
            delta_t=delta_tension_str.replace(",", ".")
            delta_tension=float(delta_t)
        else: 
            messagebox.showerror("Saisie incompléte !","Renseigner l'ecart de tension entre le module min et le module max")
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_nappe = 1,delta_tension_module = %s, date_soudure_nappe = NOW() where numero_serie_batterie = %s "
            param = (delta_tension,num_batt)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.n_numero_serie_batt_entry.delete(0, tk.END)
            self.display_model_list_nappe()
            
            
    def add_non_conf_batterie_nappe(self):
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.n_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack_fail = soudure_nappe_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.n_numero_serie_batt_entry.delete(0, tk.END)
        
    def n_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.n_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.n_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.n_numero_serie_batt_entry.delete(0, tk.END)
        self.n_numero_serie_batt_entry.insert(0, selected_value)
        
    #------------------------------ Onglet bms -----------------------------------------------   

    def setup_bms(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.b_numero_serie_cell_entry = ttk.Entry(left_frame); self.b_numero_serie_cell_entry.pack(pady=5)
        self.b_numero_serie_cell_entry.bind("<KeyRelease>",self.b_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.b_numero_serie_batt_entry = ttk.Entry(left_frame); self.b_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_bms, style="Danger.TButton").pack(pady=10)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.b_listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.b_listbox_batt.pack(pady=5, fill='both')
        self.b_listbox_batt.bind("<<ListboxSelect>>", self.b_on_select_batt)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        self.b_label_photo = tk.Label(right_frame, bg="#e0e0e0", width=200, height=200, text="Aperçu photo", anchor='center')
        self.b_label_photo.pack(pady=10)
        self.set_photo(self.b_label_photo, r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\voltr_logo.jpg")
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_bms, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_bms()
        
    def b_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.b_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.b_numero_serie_batt_entry.delete(0, tk.END)
                self.b_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.b_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_bms(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.soudure_bms is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.b_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.b_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_bms(self):
        num_batt=str(self.b_numero_serie_batt_entry.get())
        stage_col='bms'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_bms = 1, date_soudure_bms = NOW() where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.b_numero_serie_batt_entry.delete(0, tk.END)
            self.display_model_list_bms()
            
            
    def add_non_conf_batterie_bms(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.b_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_bms_fail = soudure_bms_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.b_numero_serie_batt_entry.delete(0, tk.END)
    
    def b_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.b_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.b_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.b_numero_serie_batt_entry.delete(0, tk.END)
        self.b_numero_serie_batt_entry.insert(0, selected_value)
        
    
        
    #------------------------------ Onglet wrap -----------------------------------------------   

    def setup_wrap(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.w_numero_serie_cell_entry = ttk.Entry(left_frame); self.w_numero_serie_cell_entry.pack(pady=5)
        self.w_numero_serie_cell_entry.bind("<KeyRelease>",self.w_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.w_numero_serie_batt_entry = ttk.Entry(left_frame); self.w_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_wrap, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.w_listbox_batt = tk.Listbox(right_frame, font=('Segoe UI', 11), height=10); self.w_listbox_batt.pack(pady=5, fill='both')
        self.w_listbox_batt.bind("<<ListboxSelect>>", self.w_on_select_batt)
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_wrap, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_wrap()
        
    def w_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.b_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.b_numero_serie_batt_entry.delete(0, tk.END)
                self.b_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.b_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_wrap(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.wrap is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.w_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.w_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_wrap(self):
        num_batt=str(self.w_numero_serie_batt_entry.get())
        stage_col='wrap'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        conn = self.db_manager.connect()
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET wrap = 1, date_wrap = NOW() where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.w_numero_serie_batt_entry.delete(0, tk.END)
            self.display_model_list_wrap()
            
            
    def add_non_conf_batterie_wrap(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.w_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET wrap_fail = wrap_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.w_numero_serie_batt_entry.delete(0, tk.END)
    
    def w_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.w_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.w_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.w_numero_serie_batt_entry.delete(0, tk.END)
        self.w_numero_serie_batt_entry.insert(0, selected_value)
        
        
    #------------------------------ Onglet fermeture -----------------------------------------------   
    def setup_fermeture(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.f_numero_serie_cell_entry = ttk.Entry(left_frame); self.f_numero_serie_cell_entry.pack(pady=5)
        self.f_numero_serie_cell_entry.bind("<KeyRelease>",self.f_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.f_numero_serie_batt_entry = ttk.Entry(left_frame); self.f_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_fermeture, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.f_listbox_batt = tk.Listbox(right_frame, font=('Segoe UI', 11), height=10); self.f_listbox_batt.pack(pady=5, fill='both')
        self.f_listbox_batt.bind("<<ListboxSelect>>", self.f_on_select_batt)
        ttk.Label(left_frame, text="Tension en fin de test:").pack(pady=5)
        self.tension_end_entry = ttk.Entry(left_frame); self.tension_end_entry.pack(pady=5)
        self.f_label_photo = tk.Label(right_frame, bg="#e0e0e0", width=200, height=200, text="Aperçu photo", anchor='center')
        self.f_label_photo.pack(pady=10)
        self.set_photo(self.f_label_photo, r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\voltr_logo.jpg")
        ttk.Button(left_frame, text="✅ Contrôle OK", command=self.valider_fermeture, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_fermeture()
        
    def f_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.f_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.f_numero_serie_batt_entry.delete(0, tk.END)
                self.f_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.f_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_fermeture(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.fermeture_batt is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.f_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.f_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_fermeture(self):
        num_batt=str(self.f_numero_serie_batt_entry.get())
        stage_col='fermeture'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        conn = self.db_manager.connect()
        tension_fin=self.tension_end_entry.get()
        tension_f=tension_fin.replace(",",".")
        tension_end=float(tension_f)
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET fermeture_batt = 1, date_fermeture_batt = NOW(), test_tension= %s where numero_serie_batterie = %s "
            param = (tension_end,num_batt)  
            cursor.execute(query, param)
            
            cursor.execute("Select test_capa from suivi_production where numero_serie_batterie= %s",(num_batt,))
            ress=cursor.fetchall()
            etat_f=[res[0] for res in ress]
            if etat_f:
                cursor.execute("UPDATE produit_voltr SET statut = %s where numero_serie_produit= %s",('stock',num_batt))

        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.f_numero_serie_batt_entry.delete(0, tk.END)
            self.display_model_list_fermeture()
            
            
    def add_non_conf_batterie_fermeture(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.f_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET fermeture_fail = fermeture_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.f_numero_serie_batt_entry.delete(0, tk.END)
    
    def f_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.f_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.f_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.f_numero_serie_batt_entry.delete(0, tk.END)
        self.f_numero_serie_batt_entry.insert(0, selected_value)
        
        

    #------------------------------ Onglet emballage -----------------------------------------------   
    def setup_emb(self, frame):
        
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.emb_numero_serie_batt_entry = ttk.Entry(left_frame); self.emb_numero_serie_batt_entry.pack(pady=5)
        self.emb_numero_serie_batt_entry.bind("<KeyRelease>",self.emb_check_entry_length)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_emb, style="Danger.TButton").pack(pady=10)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.emb_listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.emb_listbox_batt.pack(pady=5, fill='both')
        self.emb_listbox_batt.bind("<<ListboxSelect>>", self.emb_on_select_batt)

        right_frame= ttk.Frame(frame); right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Modèle batterie").pack(pady=5)
        cb_var_m = tk.StringVar()
        self.cb_emb = ttk.Combobox(right_frame, textvariable=cb_var_m, values=self.models, state="disabled", width=40)
        self.cb_emb.pack(pady=5)
        
        self.chk_var_emb = tk.BooleanVar(value=False)
        
        def toggle_combobox():
            self.cb_emb.configure(state="readonly" if self.chk_var_emb.get() else "disabled")

        self.chk_emb = ttk.Checkbutton(right_frame, text="Changer le modèle",
                      variable=self.chk_var_emb, command=toggle_combobox)
        self.chk_emb.pack(pady=5)
        
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_emballage, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_emballage()

    def emb_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.emb_numero_serie_batt_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
                self.emb_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_emballage(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.emballage is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.emb_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.emb_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_emballage(self):
        num_batt=str(self.emb_numero_serie_batt_entry.get())
        stage_col='emb'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        if self.chk_var_emb.get():
            new_ref=str(self.cb_emb.get())
            self.changer_ref_batterie(new_ref,num_batt)
        else:
            directive=self.verfier_coherence_ref(num_batt)
            if directive=='stop':
                return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET emballage = 1, date_emballage= NOW() where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.emb_numero_serie_batt_entry.delete(0, tk.END)
            self.display_model_list_emballage()
            self.display_model_list_exp()
            
            
    def add_non_conf_batterie_emb(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        self.emb_numero_serie_batt_entry.delete(0, tk.END)
    
    def emb_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.emb_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.emb_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.emb_numero_serie_batt_entry.delete(0, tk.END)
        self.emb_numero_serie_batt_entry.insert(0, selected_value)
    
    #------------------------------ Onglet capa -----------------------------------------------   

    def setup_capa(self, frame):
        # Conteneur principal
        wrap = ttk.LabelFrame(frame, text="Test OK", padding=12)
        wrap.pack(fill="both", expand=True, padx=8, pady=8)
        
        # ====== Tableau PRINCIPAL (tests OK) ======
        table_frame = ttk.Frame(wrap)
        table_frame.pack(fill="both", expand=True, padx=8, pady=(6, 12))
        
        cols_ok = ("N° Série", "Modèle", "Capacité", "Tension de fin de test", "Emplacement")
        self.test_tree = ttk.Treeview(table_frame, columns=cols_ok, show="headings", selectmode="browse")
        for c in cols_ok:
            self.test_tree.heading(c, text=c)
            # largeur par défaut + un peu plus pour les champs longs
            if c == "Tension de fin de test":
                width = 170
            elif c in ("Emplacement", "Modèle"):
                width = 160
            else:
                width = 140
            self.test_tree.column(c, anchor="w", width=width, stretch=True)
        
        yscroll_ok = ttk.Scrollbar(table_frame, orient="vertical", command=self.test_tree.yview)
        self.test_tree.configure(yscrollcommand=yscroll_ok.set)
        
        self.test_tree.pack(side="left", fill="both", expand=True)
        yscroll_ok.pack(side="right", fill="y")
        
        # ====== Tableau SECONDAIRE (tests défaillants) ======
        failed_box = ttk.LabelFrame(wrap, text="Tests défaillants", padding=8)
        failed_box.pack(fill="both", expand=True, padx=8, pady=(0, 12))
        
        failed_frame = ttk.Frame(failed_box)
        failed_frame.pack(fill="both", expand=True)
        
        cols_fail = ("N° Série", "Emplacement", "Cause")
        self.fail_tree = ttk.Treeview(failed_frame, columns=cols_fail, show="headings", selectmode="browse")
        for c in cols_fail:
            self.fail_tree.heading(c, text=c)
            width = 150 if c != "Cause" else 260
            self.fail_tree.column(c, anchor="w", width=width, stretch=True)
        
        yscroll_fail = ttk.Scrollbar(failed_frame, orient="vertical", command=self.fail_tree.yview)
        self.fail_tree.configure(yscrollcommand=yscroll_fail.set)
        
        self.fail_tree.pack(side="left", fill="both", expand=True)
        yscroll_fail.pack(side="right", fill="y")
        
        # ====== Gros bouton ovale centré (canvas) ======
        btn_frame = ttk.Frame(wrap)
        btn_frame.pack(fill="x", pady=8)
        
        canvas = tk.Canvas(btn_frame, width=260, height=64, highlightthickness=0, bg=self.cget("background"))
        canvas.pack(pady=6)
        oval = canvas.create_oval(4, 4, 256, 60, fill="#2E62FF", outline="#1b3fb3", width=2)
        txt  = canvas.create_text(130, 32, text="Traiter les fichiers", fill="white", font=("Segoe UI", 11, "bold"))
        
        # Bind sur l’ovale et le texte
        canvas.tag_bind(oval, "<Button-1>", self._on_click)
        canvas.tag_bind(txt,  "<Button-1>", self._on_click)
        
    def _on_click(self,event=None):
        
        ok_pairs = []   # [(numero_serie_batterie, chemin_fichier), ...]
        ko_files = []   # [chemin_fichier, ...]
        
        dossier_path=r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\dossiers\en cours"
        dossier_exploites=r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\dossiers\OK"
        dossier_ko=r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\dossiers\NOK"
        
        """
        dossier_path="C:/Users/User/Desktop/MAJ_TEST/Resultats"
        dossier_exploites="C:/Users/User/Desktop/MAJ_TEST/Fichiers_exploites"
        dossier_ko="C:/Users/User/Desktop/MAJ_TEST/Fichiers_non_ok/cellues K.O"
        """
        """
        dossier_path=r"G:\Drive partagés\VoltR\4_Production\5_Cyclage\1_Résultats de cyclage\Fichier en cours"
        dossier_exploites= "G:/Drive partagés/VoltR/4_Production/5_Cyclage/1_Résultats de cyclage/Fichiers traités/batterie"
        dossier_ko= "G:/Drive partagés/VoltR/4_Production/5_Cyclage/1_Résultats de cyclage/Fichiers NOK/Batteries KO"
        """
        df_cyclage = pd.read_excel(EXCEL_PATH,sheet_name="Cyclage",header=1)
        
        for fichier in os.listdir(dossier_path):#Traite chaque fichier dans le dossier 
            if fichier.endswith((".xlsx", ".xls")):  # Vérifie si le fichier est de type Excel
                chemin_fichier = os.path.join(dossier_path, fichier)
                numero_serie_test = os.path.splitext(os.path.basename(chemin_fichier))[0]
                # Ex: "MC0002031068-4-3-7-INR18650MH1_A.0.1"
                parties = numero_serie_test.split('-', 4)
            
                # Sécurités basiques sur le nom
                if len(parties) < 5:
                    # Nom inattendu -> on ignore/passe
                    continue
            
                numero_serie_batterie = parties[0]
                emplacement = f"{parties[1]}-{parties[2]}-{parties[3]}"
                modele_test = parties[4]
            
                if modele_test == self.selected_model:
                    try:
                        # --- DB: récupérer ref_cell ---
                        conn = self.db_manager.connect()
                        if not conn:
                            # Pas de connexion = échec de traitement
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            
                            continue
            
                        try:
                            cursor = conn.cursor()
                            query = """
                                SELECT reference_cellule
                                FROM cellule
                                WHERE affectation_produit = %s
                                LIMIT 1
                            """
                            cursor.execute(query, (numero_serie_batterie,))
                            row_db = cursor.fetchone()
                        except Exception as e:
                            # Erreur SQL => échec de traitement
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            print(f"[SQL] {numero_serie_batterie} -> {e}")
                           
                            row_db = None
                        finally:
                            try:
                                cursor.close()
                            except:
                                pass
                            conn.close()
            
                        if not row_db or not row_db[0]:
                            # Pas de ref cellule => impossible de lire les seuils
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            
                            continue
            
                        ref_cell = row_db[0]
            
                        # --- Seuils (df_cyclage) ---
                        row = df_cyclage[
                            (df_cyclage["Nom_modele"] == modele_test) &
                            (df_cyclage["Ref cellule"] == ref_cell)
                        ]
            
                        if row.empty:
                            # Seuil introuvable
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            
                            continue
            
                        seuils = row.iloc[0]
                        try:
                            capa_min = float(seuils["Capa mini (Ah)"])
                        except Exception:
                            capa_min = float(str(seuils["Capa mini (Ah)"]).replace(",", "."))
                        try:
                            temps_h = float(seuils["Temps de test (h)"])
                        except Exception:
                            temps_h = float(str(seuils["Temps de test (h)"]).replace(",", "."))
                        try:
                            tension_seuil = float(seuils["Tension seuil (V)"])
                        except Exception:
                            tension_seuil = float(str(seuils["Tension seuil (V)"]).replace(",", "."))
            
                        # --- Lecture mesures ---
                        # Capacité (onglet 'step', ligne 4, col 'Capacity(Ah)')
                        step = pd.read_excel(chemin_fichier, sheet_name="step")
                        try:
                            capa_dch = float(step["Capacity(Ah)"].iloc[3])
                        except Exception as e:
                            # Si la ligne/col manque -> échec de traitement
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            print(f"[STEP capa] {numero_serie_batterie} -> {e}")
                            
                            continue
            
                        # Tension fin test (onglet 'record', Step Index == 4, max sur les 10 derniers points)
                        record = pd.read_excel(chemin_fichier, sheet_name="record")
                        df_dch_last = record[record["Step Index"] == 4].tail(10)
                        if df_dch_last.empty or "Voltage(V)" not in df_dch_last.columns:
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            continue
                        try:
                            max_voltage = float(df_dch_last["Voltage(V)"].max())
                        except Exception as e:
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            print(f"[RECORD volt] {numero_serie_batterie} -> {e}")
                            
                            continue
            
                        # Temps test (dernière valeur 'Total Time')
                        if "Total Time" not in record.columns or record["Total Time"].empty:
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            continue
            
                        last_time_val = record["Total Time"].iloc[-1]
                        tension_finale = record["Voltage(V)"].iloc[-1]
                        # Robustifier la conversion en timedelta
                        try:
                            # Si déjà de type timedelta/chaîne "HH:MM:SS"
                            time_obj = pd.to_timedelta(last_time_val)
                            if pd.isna(time_obj):
                                raise ValueError("NaT")
                        except Exception:
                            # Si Excel time (float en jours)
                            try:
                                time_obj = pd.to_timedelta(float(last_time_val), unit="D")
                            except Exception as e:
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                print(f"[RECORD time] {numero_serie_batterie} -> {e}")
                                
                                continue
            
                        # --- Indicateurs ---
                        indic_capa = "OK" if capa_dch > capa_min else "NOK"
                        indic_v    = "OK" if max_voltage < tension_seuil else "NOK"
                        indic_t    = "OK" if time_obj > pd.to_timedelta(f"{int(temps_h)} hours") else "NOK"
                        
                        
            
                        if indic_capa == "OK" and indic_v == "OK" and indic_t == "OK":
                            # Tous OK -> tree principal (5 colonnes)
                            self.test_tree.insert(
                                "", "end",
                                values=(
                                    numero_serie_batterie,
                                    modele_test,
                                    f"{capa_dch:.3f}",
                                    f"{tension_finale:.3f}",
                                    emplacement
                                )
                            )
                            ok_pairs.append((numero_serie_batterie, chemin_fichier))
                            conn = self.db_manager.connect()
                            if not conn:
                                messagebox.showerror("Erreur DB", "Connexion DB impossible pour l'update.")
                                return
                        
                            try:
                                cur = conn.cursor()
                        
                                sql = """
                                    UPDATE suivi_production
                                    SET valeur_test_capa= %s
                                    WHERE numero_serie_batterie = %s
                                """
                        
                                params = (capa_dch,numero_serie_batterie)
                                cur.execute(sql, params)
                            except Exception as e:
                                messagebox.showerror("Erreur DB", f"Update échoué : {e}")
                            finally:
                                try:
                                    cur.close()
                                except:
                                    pass
                                conn.close()
                        else:
                            # Au moins un NOK -> tree défaillants (3 colonnes) avec cause
                            causes = []
                            if indic_capa == "NOK":
                                causes.append("Capacité")
                            if indic_v == "NOK":
                                causes.append("Tension")
                            if indic_t == "NOK":
                                causes.append("Temps")
                            cause_txt = " & ".join(causes) if causes else "traitement"
                            self.fail_tree.insert(
                                "", "end",
                                values=(numero_serie_batterie, emplacement, cause_txt)
                            )
                            ko_files.append(chemin_fichier)
            
                    except Exception as e:
                        # Toute autre erreur après le if modele_test == self.selected_model
                        self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                        ko_files.append(chemin_fichier)
                        print(f"[TRAITEMENT] {numero_serie_batterie} -> {e}")
                else:
                    # Modèle différent => on ignore ce fichier
                    pass                       
            
        # === Après la boucle: mise à jour BDD pour les OK ===
        ok_serials = list({s for (s, _p) in ok_pairs})  # dédoublonné
        if ok_serials:
            self._update_ok_in_db(ok_serials)
    
        # === Déplacement des fichiers ===
        self._move_processed_files(ok_pairs, ko_files, dossier_exploites, dossier_ko)    
        
    def _update_ok_in_db(self, ok_serials):
        """
        Met à jour la base pour tous les numéros série OK.
        Adapte la requête SQL selon ton schéma.
        """
        conn = self.db_manager.connect()
        if not conn:
            messagebox.showerror("Erreur DB", "Connexion DB impossible pour l'update.")
            return
    
        try:
            cur = conn.cursor()
    
            sql = """
                UPDATE suivi_production
                   SET test_capa = 1,
                       date_test_capa = NOW()
                 WHERE numero_serie_batterie = %s
            """
    
            params = [(s,) for s in ok_serials]
            cur.executemany(sql, params)
            
            for s in ok_serials:
                cur.execute("Select fermeture_batt from suivi_production where numero_serie_batterie= %s",(s,))
                ress=cur.fetchall()
                etat_f=[res[0] for res in ress]
                if etat_f:
                    cur.execute("UPDATE produit_voltr SET statut = %s where numero_serie_produit= %s",('stock',s))
            conn.commit()
            print(f"Update OK sur {cur.rowcount} lignes.")
        except Exception as e:
            messagebox.showerror("Erreur DB", f"Update échoué : {e}")
        finally:
            try:
                cur.close()
            except:
                pass
            conn.close()
    
    def _move_processed_files(self, ok_pairs, ko_files, dossier_exploites, dossier_ko):
        """
        Déplace les fichiers:
          - OK -> dossier_exploites
          - KO/erreurs -> dossier_ko
        """
        os.makedirs(dossier_exploites, exist_ok=True)
        os.makedirs(dossier_ko, exist_ok=True)
    
        def _safe_move(src_path, dst_dir):
            try:
                base = os.path.basename(src_path)
                dst = os.path.join(dst_dir, base)
                if os.path.exists(dst):
                    name, ext = os.path.splitext(base)
                    i = 1
                    candidate = os.path.join(dst_dir, f"{name} ({i}){ext}")
                    while os.path.exists(candidate):
                        i += 1
                        candidate = os.path.join(dst_dir, f"{name} ({i}){ext}")
                    dst = candidate
                shutil.move(src_path, dst)
                return True
            except Exception as e:
                print(f"[MOVE] {src_path} -> {dst_dir} : {e}")
                return False
    
        # OK -> exploités
        for _, src in ok_pairs:
            _safe_move(src, dossier_exploites)
    
        # KO -> ko
        for src in ko_files:
            _safe_move(src, dossier_ko)
    
        
    #------------------------------ Onglet recherche -----------------------------------------------   
        
    def setup_recherche(self, frame):
        # ----- Layout principal : gauche | boutons | droite -----
        container = ttk.Frame(frame); container.pack(fill="both", expand=True, padx=12, pady=12)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=0)
        container.columnconfigure(2, weight=2)
        container.rowconfigure(0, weight=1)
    
        # ------- Colonne gauche : entrées + combobox + listbox -------
        left = ttk.LabelFrame(container, text="Recherche", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
       
    
        ttk.Label(left, text="N° série cellule").grid(row=0, column=0, sticky="w")
        self.rech_entry_cell = ttk.Entry(left, width=28)
        self.rech_entry_cell.grid(row=1, column=0, sticky="we", pady=(0,8))
        # Remplissage auto du n° batterie quand l'entry cellule atteint 12 chars
        self.rech_entry_cell.bind("<KeyRelease>", self._rech_on_cell_entry)
    
        ttk.Label(left, text="N° série batterie").grid(row=2, column=0, sticky="w")
        self.rech_entry_batt = ttk.Entry(left, width=28)
        self.rech_entry_batt.grid(row=3, column=0, sticky="we", pady=(0,8))
    
        ttk.Label(left, text="Référence batterie").grid(row=4, column=0, sticky="w")
        # valeur par défaut nulle (vide)
        self.rech_model_var = tk.StringVar(value="")
        self.rech_combo = ttk.Combobox(left, textvariable=self.rech_model_var,
                                       values=(self.models or []), state="readonly", width=30)
        self.rech_combo.grid(row=5, column=0, sticky="we", pady=(0,8))
        self.rech_combo.bind("<<ComboboxSelected>>", lambda e: self._rech_on_model_change())
    
        ttk.Label(left, text="Liste batterie").grid(row=6, column=0, sticky="w")
        lb_frame = ttk.Frame(left); lb_frame.grid(row=7, column=0, sticky="nsew")
        left.rowconfigure(7, weight=1)
    
        # multi-sélection
        self.rech_listbox = tk.Listbox(lb_frame, height=10, activestyle="dotbox", selectmode="extended")
        yscroll = ttk.Scrollbar(lb_frame, orient="vertical", command=self.rech_listbox.yview)
        self.rech_listbox.configure(yscrollcommand=yscroll.set)
        self.rech_listbox.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
    
        # Double-clic => déplacer à droite
        self.rech_listbox.bind("<Double-1>", lambda e: self._rech_move_right())
    
        # --------- Colonne boutons centraux ----------
        mid = ttk.Frame(container); mid.grid(row=0, column=1, sticky="ns")
        for i in range(3): mid.rowconfigure(i, weight=1)
        ttk.Button(mid, text="→", width=3, command=self._rech_move_right).grid(row=0, column=0, pady=4)
        ttk.Button(mid, text="←", width=3, command=self._rech_remove_selected_right).grid(row=1, column=0, pady=4)
    
        # --------- Colonne droite : table dynamique ----------
        right = ttk.LabelFrame(container, text="Sélection / Détails", padding=10)
        right.grid(row=0, column=2, sticky="nsew", padx=(8,0))
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        
        
        # largeur fixe (par ex. 500 px, ajuste comme tu veux)
        right.configure(width=800)
        right.grid_propagate(False)   # bloque l’expansion auto
        
        
        self.rech_right_title = ttk.Label(right, text="", font=("", 10, "bold"))
        self.rech_right_title.grid(row=0, column=0, sticky="w", pady=(0,6))
    
        tv_frame = ttk.Frame(right); tv_frame.grid(row=1, column=0, sticky="nsew")
        # + scroll vertical & horizontal
        self.rech_tree = ttk.Treeview(tv_frame, columns=(), show="headings", selectmode="extended")
        y2 = ttk.Scrollbar(tv_frame, orient="vertical", command=self.rech_tree.yview)
        x2 = ttk.Scrollbar(tv_frame, orient="horizontal", command=self.rech_tree.xview)
        self.rech_tree.configure(yscrollcommand=y2.set, xscrollcommand=x2.set)
        self.rech_tree.grid(row=0, column=0, sticky="nsew")
        y2.grid(row=0, column=1, sticky="ns")
        x2.grid(row=1, column=0, sticky="ew")
        tv_frame.rowconfigure(0, weight=1); tv_frame.columnconfigure(0, weight=1)
    
        # set utilisé pour éviter les doublons à droite (clé = numero_serie_batterie)
        self._rech_right_keys = set()
        
    def _rech_on_cell_entry(self, event=None):
        """Quand l'entry cellule atteint 12 chars, on cherche la batterie associée et on remplit l'entry batterie."""
        numero_serie_cell = self.rech_entry_cell.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # affectation_produit = numero_serie_batterie (selon ta logique existante)
            cur.execute("SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s", (numero_serie_cell,))
            row = cur.fetchone()
            self.rech_entry_batt.delete(0, tk.END)
            if row and row[0]:
                self.rech_entry_batt.insert(0, str(row[0]))
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Lookup cellule→batterie impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_on_model_change(self):
        """Quand on choisit un modèle, on alimente la liste des n° batteries via la jointure demandée."""
        ref = self.rech_model_var.get().strip()
        self.rech_listbox.delete(0, tk.END)
        if not ref:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # Liste des NUMÉROS DE SÉRIE BATTERIE pour la référence choisie
            # sp = suivi_production / p = produit
            sql = ("""
                SELECT DISTINCT sp.numero_serie_batterie
                FROM suivi_production sp
                JOIN produit_voltr p
                  ON sp.numero_serie_batterie = p.numero_serie_produit
                WHERE p.reference_produit_voltr = %s
                ORDER BY sp.numero_serie_batterie
            """)
            cur.execute(sql, (ref,))
            for (num_batt,) in cur.fetchall():
                if num_batt:
                    self.rech_listbox.insert(tk.END, str(num_batt))
            self.rech_right_title.config(text=f"Modèle sélectionné : {ref}")
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement liste batteries impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_move_right(self):
        """Ajoute à droite : 1) le n° saisi dans l’entry batterie (si présent),
        2) tous les n° sélectionnés dans la liste. Charge les LIGNES de suivi_production correspondantes."""
        # rassembler les cibles
        targets = set()
        batt_from_entry = self.rech_entry_batt.get().strip()
        if batt_from_entry:
            targets.add(batt_from_entry)
        for idx in self.rech_listbox.curselection():
            targets.add(self.rech_listbox.get(idx))
    
        if not targets:
            return
    
        # Charger les lignes depuis suivi_production
        rows, colnames = self._rech_load_suivi_rows(list(targets))
        if not rows:
            return
    
        # Configurer le tree si c'est la 1ère fois ou si colonnes différentes
        self._rech_configure_tree_for_columns(colnames)
    
        # Insérer sans doublon (clé = numero_serie_batterie)
        try:
            k_idx = colnames.index("numero_serie_batterie")  # l’utilisateur veut dédupliquer sur ce champ
        except ValueError:
            # si absent (peu probable), on déduplique sur la 1ère colonne
            k_idx = 0
    
        added = 0
        for r in rows:
            key = str(r[k_idx]) if r[k_idx] is not None else ""
            if key and key not in self._rech_right_keys:
                self.rech_tree.insert("", tk.END, values=[("" if v is None else v) for v in r])
                self._rech_right_keys.add(key)
                added += 1
    
        if added == 0:
            self.rech_right_title.config(text="Aucun nouvel élément (déduplication active)")
    
    def _rech_remove_selected_right(self):
        """Supprimer les lignes sélectionnées dans la table de droite."""
        sel = self.rech_tree.selection()
        if not sel:
            return
        # identifier l'index de numero_serie_batterie pour nettoyer le set
        cols = self.rech_tree.cget("columns")
        try:
            k_idx = cols.index("numero_serie_batterie")
        except ValueError:
            k_idx = 0
        for iid in sel:
            vals = self.rech_tree.item(iid, "values")
            # protéger si vide
            if vals:
                key = str(vals[k_idx])
                if key in self._rech_right_keys:
                    self._rech_right_keys.remove(key)
            self.rech_tree.delete(iid)
    
    def _rech_load_suivi_rows(self, numero_serie_batteries):
        """Retourne (rows, colnames) pour les lignes de suivi_production filtrées par n° série batterie."""
        if not numero_serie_batteries:
            return [], []
        conn = self.db_manager.connect()
        if not conn:
            return [], []
        try:
            cur = conn.cursor()
            # On charge TOUTE la ligne de suivi_production (c’est ce que tu as demandé)
            # IN sécurisé : on fabrique la liste de %s
            placeholders = ", ".join(["%s"] * len(numero_serie_batteries))
            sql = f"SELECT * FROM suivi_production WHERE numero_serie_batterie IN ({placeholders})"
            cur.execute(sql, tuple(numero_serie_batteries))
            rows = cur.fetchall()
            colnames = [d[0] for d in cur.description]
            return rows, colnames
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement suivi_production impossible :\n{e}")
            return [], []
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_configure_tree_for_columns(self, colnames):
        """Configure le Treeview (colonnes, largeur, ancrage, headings) + stretch + scroll horiz/vert déjà mis côté UI."""
        # si colonnes déjà identiques, ne rien faire
        current = self.rech_tree.cget("columns")
        if tuple(current) == tuple(colnames):
            return
    
        # reset
        for c in current:
            try: self.rech_tree.heading(c, text="")
            except: pass
        self.rech_tree.delete(*self.rech_tree.get_children())
        self._rech_right_keys.clear()
    
        self.rech_tree["columns"] = colnames
        # headings & options
        for c in colnames:
            self.rech_tree.heading(c, text=c)
            # largeur heuristique : min 100, sinon ~len*9 px
            width = max(100, min(380, int(len(c) * 9)))
            self.rech_tree.column(c, width=width, minwidth=80, stretch=True, anchor="w")

    
    #------------------------------ Onglet expedition -----------------------------------------------   

    def setup_exp(self, frame):
        # ---------- Layout principal : gauche | droite ----------
        container = ttk.Frame(frame); container.pack(fill="both", expand=True, padx=12, pady=12)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(0, weight=1)
    
        # ========== Colonne gauche : entrées & combos ==========
        left = ttk.LabelFrame(container, text="Expédition", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
    
        # N° série produit + ajout auto si 9 chiffres
        ttk.Label(left, text="N° série produit:").pack(pady=(0,4), anchor="w")
        self.exp_numero_serie_batt_entry = ttk.Entry(left, width=30)
        self.exp_numero_serie_batt_entry.pack(pady=(0,8), fill="x")
        self.exp_numero_serie_batt_entry.bind("<KeyRelease>", self._exp_on_entry_change)
    
        ttk.Button(left, text="❌ Non conforme", command=self.add_non_conf_batterie, style="Danger.TButton").pack(pady=6, anchor="w")
    
        # Client
        ttk.Label(left, text="Client").pack(pady=(8,4), anchor="w")
        cb_var_client = tk.StringVar()
        self.cb_cl = ttk.Combobox(left, textvariable=cb_var_client, state="readonly", width=40)
        self.cb_cl.pack(fill="x")
    
        # Alimentation des clients
        conn = self.db_manager.connect()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT nom_client FROM client ORDER BY nom_client")
                mots = [row[0].replace(" ", "-") for row in cursor.fetchall()]
                self.cb_cl["values"] = mots
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Chargement clients :\n{e}")
            finally:
                try: cursor.close()
                except: pass
                conn.close()
    
        # Projet
        ttk.Label(left, text="Projet").pack(pady=(8,4), anchor="w")
        self.cb_pr = ttk.Combobox(left, state="readonly", width=40)
        self.cb_pr.pack(fill="x")
        
        # Alimentation des clients
        conn = self.db_manager.connect()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT nom_projet FROM projet ORDER BY nom_projet")
                mots = [row[0] for row in cursor.fetchall()]
                self.cb_pr["values"] = mots
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Chargement projet :\n{e}")
            finally:
                try: cursor.close()
                except: pass
                conn.close()
    
        # Modèle + checkbox d’activation
        ttk.Label(left, text="Modèle batterie").pack(pady=(8,4), anchor="w")
        cb_var_m_exp = tk.StringVar()
        self.cb_exp = ttk.Combobox(left, textvariable=cb_var_m_exp, values=getattr(self, "models", []),
                                   state="disabled", width=40)
        self.cb_exp.pack(fill="x")
    
        self.chk_exp_var = tk.BooleanVar(value=False)
        def toggle_combobox_exp():
            self.cb_exp.configure(state="readonly" if self.chk_exp_var.get() else "disabled")
        ttk.Checkbutton(left, text="Changer le modèle", variable=self.chk_exp_var,
                        command=toggle_combobox_exp).pack(pady=(6,0), anchor="w")
    
        # Commentaire
        ttk.Label(left, text="Commentaire").pack(pady=(8,4), anchor="w")
        self.exp_comm_entry = ttk.Entry(left)
        self.exp_comm_entry.pack(fill="x")
    
        # ========= Colonne droite : 2 Listbox (disponibles | sélectionnées) =========
        right = ttk.Frame(container)
        right.grid(row=0, column=1, sticky="nsew", padx=(8,0))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)
        right.rowconfigure(4, weight=1)
    
        # Liste des batteries disponibles (emballées)
        ttk.Label(right, text="Batteries emballées (disponibles)").grid(row=0, column=0, sticky="w", pady=(0,4))
        self.exp_listbox_batt = tk.Listbox(right, font=('Segoe UI', 11), height=10)
        self.exp_listbox_batt.grid(row=1, column=0, sticky="nsew")
        self.exp_listbox_batt.bind("<<ListboxSelect>>", lambda e: None)  # neutre
        self.exp_listbox_batt.bind("<Double-Button-1>", self._exp_on_available_double_click)
    
        ttk.Separator(right, orient="horizontal").grid(row=2, column=0, sticky="ew", pady=10)
    
        # Sélection (avec compteur)
        header_sel = ttk.Frame(right); header_sel.grid(row=3, column=0, sticky="ew", pady=(0,4))
        ttk.Label(header_sel, text="Batteries sélectionnées").pack(side="left")
        ttk.Label(header_sel, text="Quantité:").pack(side="right")
        self._exp_count_var = tk.IntVar(value=0)
        self._exp_count_lbl = ttk.Label(header_sel, textvariable=self._exp_count_var)
        self._exp_count_lbl.pack(side="right", padx=(0,8))
    
        self.send_listbox_batt = tk.Listbox(right, font=('Segoe UI', 11), height=10)
        self.send_listbox_batt.grid(row=4, column=0, sticky="nsew")
        self.send_listbox_batt.bind("<Double-Button-1>", self._exp_on_selected_double_click)
    
        ttk.Button(right, text="✅ Contrôle OK", command=self.valider_exp, style="Good.TButton").grid(
            row=5, column=0, pady=10, sticky="e"
        )
    
        # Charge la liste des disponibles
        self.display_model_list_exp()
    
    
    def display_model_list_exp(self):
        """Alimente la listbox des batteries emballées (disponibles)."""
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT numero_serie_batterie FROM suivi_production WHERE emballage = 1"
            cursor.execute(query)
            rows = cursor.fetchall()
            liste_batteries = [str(r[0]) for r in rows]
            self.exp_listbox_batt.delete(0, tk.END)
            for batt in liste_batteries:
                self.exp_listbox_batt.insert(tk.END, batt)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try: cursor.close()
            except: pass
            conn.close()
    
    
    # ===================== Handlers & utilitaires =====================
    
    def _exp_on_entry_change(self, event):
        """Ajoute la batterie dans 'sélectionnées' quand l'entrée atteint 9 chiffres."""
        txt = self.exp_numero_serie_batt_entry.get().strip()
        if len(txt) == 9:
            self._exp_add_to_selection(txt)
            self.exp_numero_serie_batt_entry.delete(0, tk.END)
    
    def _exp_on_available_double_click(self, event):
        """Double-clic sur une batterie disponible -> ajoute à la sélection."""
        sel = self.exp_listbox_batt.curselection()
        if not sel:
            return
        value = self.exp_listbox_batt.get(sel[0])
        self._exp_add_to_selection(value)
    
    def _exp_on_selected_double_click(self, event):
        """Double-clic sur une batterie sélectionnée -> la retire de la sélection."""
        sel = self.send_listbox_batt.curselection()
        if not sel:
            return
        value = self.send_listbox_batt.get(sel[0])
        # Retire la première occurrence (il n'y a pas de doublon, donc OK)
        self.send_listbox_batt.delete(sel[0])
        self._exp_update_counter()
    
    def _exp_add_to_selection(self, numero):
        """Ajoute sans doublon à la listbox sélectionnée puis met à jour le compteur."""
        # Anti-doublon
        current = set(self.send_listbox_batt.get(0, tk.END))
        if numero in current:
            return
        self.send_listbox_batt.insert(tk.END, numero)
        self._exp_update_counter()
    
    def _exp_update_counter(self):
        """Mise à jour du compteur de batteries sélectionnées."""
        self._exp_count_var.set(self.send_listbox_batt.size())
        
    def _exp_get_selected_batteries(self):
        """Retourne toutes les batteries présentes dans la listbox 'sélectionnées'."""
        return list(self.send_listbox_batt.get(0, tk.END))
    
    def exp_add_non_conf_batterie(self):
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        self.exp_numero_serie_batt_entry.delete(0, tk.END)
        
        
    def valider_exp(self):
        
        # 0) Récup sélection
        numeros = list(self.send_listbox_batt.get(0, tk.END))
        if not numeros:
            messagebox.showwarning("Avertissement", "Aucune batterie sélectionnée.")
            return
        
        # 1) Référence cible (combobox si case cochée, sinon self.modele)
        target_ref = ""
        if getattr(self, "chk_exp_var", None) and self.chk_exp_var.get():
            target_ref = self.cb_exp.get().strip()
        if not target_ref:
            target_ref = getattr(self, "selected_model", "").strip()
        
        # 2) Lecture des références actuelles via IN (...)
        placeholders = ", ".join(["%s"] * len(numeros))
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
        
            # -- Ajuste ce SELECT selon ton schéma (JOIN si id_produit, sinon colonne directe)
            sql_sel = f"""
                SELECT sp.numero_serie_batterie, p.reference_produit_voltr
                FROM suivi_production AS sp
                JOIN produit_voltr AS p ON p.numero_serie_produit = sp.numero_serie_batterie
                WHERE sp.numero_serie_batterie IN ({placeholders})
            """
            cur.execute(sql_sel, tuple(numeros))
            rows = cur.fetchall()
            num2ref = {str(n): str(r) for (n, r) in rows}
        
            # Vérifs de base (numéros non trouvés)
            manquants = [n for n in numeros if n not in num2ref]
            if manquants:
                messagebox.showerror("Erreur",
                                     f"Références introuvables pour : {', '.join(manquants)}")
                return
        
            # 3) Regroupement par référence
            groups = {}
            for n, ref in num2ref.items():
                groups.setdefault(ref, []).append(n)
        
            # 4) Cohérence vs target_ref
            #    - si target_ref définie : tout ce qui n'a pas cette ref est "à corriger"
            #    - si target_ref vide : on ne corrige pas, mais on exige l'homogénéité
            to_fix = []
            if target_ref:
                for ref, nums in groups.items():
                    if ref != target_ref:
                        to_fix.extend(nums)
        
                if to_fix:
                    # 5) Proposition de correction
                    if not messagebox.askyesno(
                        "Référence différente",
                        f"{len(to_fix)} batterie(s) ne sont pas en '{target_ref}'.\n"
                        "Voulez-vous les mettre toutes à cette référence ?"
                    ):
                        # L’utilisateur refuse -> on arrête le process
                        return
        
                    # 6) Mise à jour des références
                    placeholders_fix = ", ".join(["%s"] * len(to_fix))
                    # -- VERSION avec id_produit (JOIN sur reference_produit)
                    sql_upd = f"""
                        UPDATE produit_voltr 
                        SET reference_produit_voltr = %s
                        WHERE produit_voltr.numero_serie_produit IN ({placeholders_fix})
                    """
                    cur.execute(sql_upd, (target_ref, *to_fix))
                    conn.commit()
        
                    # 7) Re-lecture pour contrôle
                    cur.execute(sql_sel, tuple(numeros))
                    rows = cur.fetchall()
                    num2ref = {str(n): str(r) for (n, r) in rows}
                    groups = {}
                    for n, ref in num2ref.items():
                        groups.setdefault(ref, []).append(n)
        
            # 8) Garde-fou : exiger une seule référence finale
            if len(groups) > 1:
                refs = ", ".join(groups.keys())
                messagebox.showerror("Références multiples",
                                     f"Plusieurs références restent présentes : {refs}\n"
                                     f"Process interrompu.")
                return
        
            # 9) Suite du process OK (une seule ref, corrigée si besoin)
            ref_unique = next(iter(groups.keys()))
            numeros_final = groups[ref_unique]
            
            # 9) FLUX MÉTIER : marquer expédiées
            comment = self.exp_comm_entry.get().strip()
            placeholders_final = ", ".join(["%s"] * len(numeros))
            projet=self.cb_pr.get().strip()
            nom_client=self.cb_cl.get().strip()
            
            conn2 = self.db_manager.connect()
            if conn2:
                try:
                    cursor = conn2.cursor()
                    cursor.execute("SELECT nom_client FROM client ORDER BY nom_client")
                    mots = [row[0].replace(" ", "-") for row in cursor.fetchall()]
                    self.cb_cl["values"] = mots
                except Exception as e:
                    messagebox.showerror("Erreur SQL", f"Chargement clients :\n{e}")
                finally:
                    try: cursor.close()
                    except: pass
                    conn2.close()
            
            #Obtenir id_client
            if nom_client:
                nom_c_bdd=nom_client.replace("-"," ")
                conn2 = self.db_manager.connect()
                if conn2:
                    try:
                        cursor = conn2.cursor()
                        cursor.execute("SELECT id_client FROM client where nom_client = %s",(nom_c_bdd,))
                        id_client = [row[0] for row in cursor.fetchall()]
                    except Exception as e:
                        messagebox.showerror("Erreur SQL", f"Chargement id_client :\n{e}")
                    finally:
                        try: cursor.close()
                        except: pass
                        conn2.close()
    
            if comment:
                
                sql_sp = f"""
                    UPDATE suivi_production
                    SET expedition = 1,
                        date_expedition = NOW(),
                        commentaire = %s
                    WHERE numero_serie_batterie IN ({placeholders_final})
                """
                params_sp = (comment, *numeros)
                
            else:
                
                sql_sp = f"""
                    UPDATE suivi_production
                    SET expedition = 1,
                        date_expedition = NOW()
                    WHERE numero_serie_batterie IN ({placeholders_final})
                """
                params_sp = tuple(numeros)
    
            cur.execute(sql_sp, params_sp)
            
            set_parts = ["statut = 'expediee'"]
            params_mark = []
            
            if nom_client:
                if id_client:
                    set_parts.append("id_client = %s")
                    params_mark.append(id_client)
            
            if projet:
                set_parts.append("numero_projet = %s")
                params_mark.append(projet)
            
            sql_mark = f"""
                UPDATE produit_voltr
                SET {', '.join(set_parts)}
                WHERE numero_serie_produit IN ({placeholders_final})
            """
            
            params_mark = tuple(params_mark) + tuple(numeros)
            
            cur.execute(sql_mark, params_mark)
            
            conn.commit()
    
            # 10) UI : succès + refresh
            messagebox.showinfo("Succès", f"{len(numeros)} batterie(s) marquées expédiées.")
            self.send_listbox_batt.delete(0, tk.END)       # vide la sélection
            self._exp_update_counter()                      # remet le compteur à jour
            self.display_model_list_exp()                   # recharge les "disponibles"
            self.exp_comm_entry.delete(0, tk.END)           # vide le commentaire
            
        except Exception as e:
            try:
                conn.rollback()
            except:
                pass
            messagebox.showerror("Erreur SQL", f"valider_exp :\n{e}")
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass
        
if __name__ == "__main__":
    app = StockApp()
    app.mainloop()
