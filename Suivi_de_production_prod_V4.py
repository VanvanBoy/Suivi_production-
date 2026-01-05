# -*- coding: utf-8 -*-
"""
Created on Thu Aug 21 15:18:57 2025

@author: User
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os, re, shutil
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from ttkthemes import ThemedTk
from PIL import Image, ImageTk
import pandas as pd
import mysql.connector
import webbrowser
from datetime import datetime

EXCEL_PATH = r"G:\Drive partagés\VoltR\11_Data\IHM\Instructions IHM\Suivi_prod_par_modele.xlsx" 

class DBManager:
    def __init__(self):
        
        def get_db_credentials():
            # Fonction pour obtenir les informations d'identification de l'utilisateur
            user = simpledialog.askstring("Login", "Enter your MySQL username:")
            password = simpledialog.askstring("Login", "Enter your MySQL password:", show='*')
            return user, password
        
        self.user,self.password=get_db_credentials()
        """
        self.config = {
            'user': self.user,
            'password': self.password,
            'host': '34.77.226.40',
            'database': 'bdd_23102025',
            'auth_plugin': 'mysql_native_password'
        }
        
        """
        self.config = {
            'user': self.user,
            'password': self.password,
            'host': '34.77.226.40',
            'database': 'cellules_batteries_cloud',
            'auth_plugin': 'mysql_native_password'
        }
        
        """
        self.config = {
            'user': 'root',
            'password': 'VoltR99!',
            'host': '127.0.0.1',
            'database': 'bdd_29072025',
            'auth_plugin': 'mysql_native_password'
        }
        
        """
        
    def connect(self):
        try:
            return mysql.connector.connect(**self.config)
        except mysql.connector.Error as err:
            messagebox.showerror("Erreur DB", f"Erreur lors de la connexion : {err}")
            return None


class StockApp(ThemedTk):

    
    def __init__(self):
        super().__init__(theme="xpnative")
        self.title("Suivi de production")
        self.geometry("1000x650")
        
        self.refresh_ms = 10_000  # 10 secondes
        self._tab_refresh_job = None
        self._refreshing = False
        
        # annuler proprement à la fermeture
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
        self.focus_targets = {}  # stage -> widget qui doit recevoir le focu
        
        self.db_manager = DBManager()
        self.picking_file_path = None

        # styles
        self.style = ttk.Style()
        self.style.configure("TLabel", font=('Segoe UI', 11), padding=4)
        self.style.configure("TButton", font=('Segoe UI', 11), padding=6)
        self.style.configure("TEntry", font=('Segoe UI', 11))
        self.style.configure("TCombobox", font=('Segoe UI', 11))
        self.style.configure("Danger.TButton", foreground="black", background="#F06F65")
        self.style.map("Danger.TButton", background=[('active', '#e55b50')], foreground=[('disabled', 'black')])
        self.style.configure("Good.TButton", foreground="black", background="#0ED329")
        self.style.map("Good.TButton", background=[('active', '#0ED329')], foreground=[('disabled', 'black')])

        
        self.selected_model = None
        self.stage_order = None  
    
        self.STAGE_TO_DBCOL = {
            "picking":   "picking_tension",          
            "pack":      "soudure_pack",     
            "nappe":     "soudure_nappe",
            "bms":       "soudure_bms",
            "wrap":      "wrap",
            "fermeture_batt": "fermeture_batt",
            "capa":      "test_capa",        
            "emb":       "emballage",        
            "exp":       "expedition",       
            "recherche": "recherche",
            "recyclage": "recyclage"   
        }
        
        self.ALLOWED_STAGE_KEYS = set(self.STAGE_TO_DBCOL.keys())  # pour sécuriser
        
        self._show_model_selector_and_build()
    
    def on_close(self):
        self._cancel_tab_refresh()
        self.destroy()
        
    def get_db_credentials(self):
        # Fonction pour obtenir les informations d'identification de l'utilisateur
        user = simpledialog.askstring("Login", "Enter your MySQL username:")
        password = simpledialog.askstring("Login", "Enter your MySQL password:", show='*')
        return user, password
    
    def _required_previous_dbcols(self,current_stage: str):
        """
        Retourne la liste des NOMS DE COLONNES MySQL (dans suivi_production)
        à vérifier (=1) avant de valider current_stage.
    
        On s'appuie sur self.stage_order (rangs > 0 = étape active)
        et sur STAGE_TO_DBCOL pour la correspondance.
        """
        if not hasattr(self, "stage_order") or not self.stage_order:
            return []
    
        if current_stage not in self.ALLOWED_STAGE_KEYS:
            return []
    
        cur_rank = self.stage_order.get(current_stage, 0)
        if cur_rank <= 0:
            return []
    
        prev_pairs = []
        for k, v in self.stage_order.items():
            if k in self.ALLOWED_STAGE_KEYS and v > 0 and v < cur_rank:
                dbcol = self.STAGE_TO_DBCOL.get(k)
                if dbcol:
                    prev_pairs.append((k, v, dbcol))
    
        # tri pour lisibilité/diagnostic
        prev_pairs.sort(key=lambda x: x[1])
    
        # on retourne uniquement les colonnes DB
        return [dbcol for (_, _, dbcol) in prev_pairs]
    
    
    def _check_prereqs_and_warn(self, num_batt: str, current_stage: str) -> bool:
        """
        Vérifie dans suivi_production que TOUTES les colonnes prérequis (=1)
        sont validées pour num_batt, d'après le mapping STAGE_TO_DBCOL.
        """
        prev_dbcols = self._required_previous_dbcols(current_stage)
        if not prev_dbcols:
            return True  # pas de prérequis
    
        conn = self.db_manager.connect()
        if not conn:
            return False
        try:
            cursor = conn.cursor(dictionary=True)
    
            # Sécurisation : on n'insère que des noms de colonnes issus de STAGE_TO_DBCOL
            cols_sql = ", ".join(prev_dbcols)
            sql = f"""
                SELECT {cols_sql}
                FROM suivi_production
                WHERE numero_serie_batterie = %s
                LIMIT 1
            """
            cursor.execute(sql, (num_batt,))
            row = cursor.fetchone()
            if not row:
                messagebox.showwarning("Vérification", "Sélectionner une batterie")
                return False
    
            missing = [c for c in prev_dbcols if (row.get(c) or 0) != 1]
            if missing:
                lis = ", ".join(missing)
                messagebox.showwarning(
                    "Pré-requis manquants",
                    f"Impossible de valider '{current_stage}'. Étape(s) non validée(s) : {lis}"
                )
                return False
    
            return True
        except Exception as e:
            messagebox.showerror("SQL", f"Erreur vérification prérequis:\n{e}")
            return False
        finally:
            try:
                cursor.close(); conn.close()
            except:
                pass
                
    def _show_model_selector_and_build(self):
        try:
            df = pd.read_excel(EXCEL_PATH,sheet_name="Flux test modele")
        except Exception as e:
            messagebox.showerror("Excel", f"Impossible de lire l'Excel:\n{e}")
            self.destroy()
            return

        if "nom_modele" not in df.columns:
            messagebox.showerror("Excel", "La colonne 'nom_modele' est absente du fichier.")
            self.destroy()
            return

        self.models = df["nom_modele"].dropna().astype(str).unique().tolist()
        if not self.models:
            messagebox.showerror("Excel", "Aucune valeur dans 'nom_modele'.")
            self.destroy()
            return

        dlg = tk.Toplevel(self)
        dlg.title("Choisir la référence batterie")
        dlg.geometry("420x160")
        dlg.transient(self)
        dlg.grab_set()

        ttk.Label(dlg, text="Référence batterie :").pack(pady=(18, 6))
        cb_var = tk.StringVar()
        cb = ttk.Combobox(dlg, textvariable=cb_var, values=self.models, state="readonly", width=40)
        cb.pack()
        cb.current(0)

        def on_launch():
            
            ref = cb_var.get()
            row = df.loc[df["nom_modele"].astype(str) == ref]
            if row.empty:
                messagebox.showerror("Sélection", "Référence introuvable.")
                return

            column_to_stage = {
                "picking": "picking",
                "soudure_pack": "pack",
                "soudure_nappe": "nappe",
                "soudure_bms": "bms",
                "wrap": "wrap",
                "fermeture": "fermeture_batt",
                "test_capa": "capa",
                "emballage": "emb",
                "expedition": "exp",
                "recherche": "recherche",
                "recyclage": "recyclage",
            }

            stage_order = {}
            for col, key in column_to_stage.items():
                if col in df.columns:
                    try:
                        val = int(row.iloc[0][col])
                    except Exception:
                        # NaN / texte -> traite comme 0
                        val = 0
                    stage_order[key] = val

            self.selected_model = ref
            self.stage_order = stage_order
            
            dlg.destroy()

        ttk.Button(dlg, text="Lancer l'application", style="Good.TButton", command=on_launch).pack(pady=16)

        # centre la boîte
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 210
        y = self.winfo_y() + (self.winfo_height() // 2) - 80
        dlg.geometry(f"+{x}+{y}")

        self.wait_window(dlg)

        if self.stage_order is None:
            
            self.destroy()
            return

        
        self._create_widgets_with_order()
        
        stage_to_func = {
            'picking': self.display_model_list,
            'pack':  self.display_model_list_pack,
            'bms':  self.display_model_list_bms,
            'fermeture_batt':  self.display_model_list_fermeture,
            'emb':  self.display_model_list_emballage,
            'exp':  self.display_model_list_exp,
            'nappe':  self.display_model_list_nappe,
            'wrap':  self.display_model_list_wrap,
        }
        
        self.funcs_to_run= [
            stage_to_func[k]
            for k, v in sorted(self.stage_order.items(), key=lambda kv: kv[1])
            if v > 0 and k in stage_to_func
        ]
        
        for f in self.funcs_to_run:
            print(f.__name__)
            
    
    def verif_etape_act_non_ok(self,current_stage,num_prod):
        
        if current_stage not in self.ALLOWED_STAGE_KEYS:
            return False
        
        dbcol = self.STAGE_TO_DBCOL[current_stage]
        
        query=f"select {dbcol} from suivi_production where numero_serie_batterie =%s"
        param=(num_prod,)
        
        conn=self.db_manager.connect()
        cursor=conn.cursor()
        cursor.execute(query,param)
        
        bool_stage=cursor.fetchone()[0]
        if bool_stage:
            bool_stage=int(bool_stage)
            
        if bool_stage==1:
            messagebox.showerror('Etape deja validée',"Etape deja validée !")
            return False
        else :
            return True
        

    def _create_widgets_with_order(self):
        
        self.title("Suivi de production - "+str(self.selected_model))
        stage_defs = {
            "picking":   ("Contrôle de picking", self.setup_picking),
            "pack":      ("Contrôle soudure pack", self.setup_pack),
            "nappe":     ("Contrôle soudure nappe", self.setup_nappe),
            "bms":       ("Contrôle soudure BMS", self.setup_bms),
            "wrap":      ("Contrôle wrap", self.setup_wrap),
            "fermeture_batt": ("Contrôle fermeture", self.setup_fermeture),
            "capa":      ("Test de capacité", self.setup_capa),
            "emb":       ("Contrôle emballage", self.setup_emb),
            "exp":       ("Contrôle expédition", self.setup_exp),
            "recherche": ("Recherche de batterie", self.setup_recherche),
            "recyclage": ("Gestion recyclage", self.setup_recyclage)
        }

        
        ordered_keys = [
            k for k, v in sorted(self.stage_order.items(), key=lambda kv: kv[1])
            if v and v > 0 and k in stage_defs
        ]

        if not ordered_keys:
            messagebox.showwarning("Configuration", "Aucun onglet actif pour cette référence.")
          
            ordered_keys = ["picking"]
        
        self.ordered_keys = ordered_keys

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill="both")
        
        self.tab_to_stage = {}   # tab_id (str) -> 'picking' | 'pack' | ...
        self.stage_refreshers = {}  # 'picking' -> fonction reload

        for key in ordered_keys:
            title, setup_fn = stage_defs[key]
            print(title)
            frame = ttk.Frame(self.notebook)
            self.notebook.add(frame, text=title)
            focus_widget=setup_fn(frame)
            if focus_widget is not None:
                self.register_focus_target(key, focus_widget)
            # mémorise le mapping tab -> stage
            self.tab_to_stage[str(frame)] = key
            
        # mappe les fonctions de refresh (remplace par tes vraies fonctions)
        self.stage_refreshers = {
            'picking':         self.display_model_list,
            'pack':            self.display_model_list_pack,
            'nappe':           self.display_model_list_nappe,
            'bms':             self.display_model_list_bms,
            'wrap':            self.display_model_list_wrap,
            'fermeture_batt':  self.display_model_list_fermeture,
            'capa':            self.afficher_numero_en_attente,
        }
    
        # Bind: lorsque l’onglet change → reload immédiat + restart timer
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
    
        # Premier reload immédiat, puis démarrage du cycle 10 s
        self._refresh_active_tab_now()
        self._schedule_next_tab_tick()
        self.after(0, self._focus_active_tab)
    
    def register_focus_target(self, stage: str, widget):
        """Enregistre le widget qui doit recevoir le focus au changement d'onglet."""
        if not hasattr(self, "focus_targets"):
            self.focus_targets = {}
        if widget is not None:
            self.focus_targets[stage] = widget
            
    def make_tab_chain(self, widgets, submit_button=None, ring=True, enter_from_fields=False):
        """
        Définit l'ordre Tab explicite et fait en sorte que seul le bouton
        déclenche l'action quand il est focalisé.
        - widgets : liste ordonnée des widgets (Entry/Combobox/Button/…).
                    Si tu veux que Tab atteigne le bouton, mets-le en dernier.
        - submit_button : ttk.Button (optionnel). Seul le bouton recevra <Return>.
        - ring : True => Tab boucle sur le 1er widget
        - enter_from_fields : si True, <Return> sur les champs déclenchera submit_button
                              (NE PAS utiliser si dangereux). Par défaut False.
        """
        if not widgets:
            return
    
        # assure takefocus sur les widgets inclus
        for w in widgets:
            try:
                w['takefocus'] = True
            except Exception:
                pass
    
        def next_idx(i):
            return (i + 1) % len(widgets) if ring else min(i + 1, len(widgets) - 1)
        def prev_idx(i):
            return (i - 1) % len(widgets) if ring else max(i - 1, 0)
    
        for i, w in enumerate(widgets):
            # handlers capturant i (closure sûre grâce à i=i)
            def go_next(e, i=i):
                widgets[next_idx(i)].focus_set()
                return "break"
            def go_prev(e, i=i):
                widgets[prev_idx(i)].focus_set()
                return "break"
    
            # Bind Tab / Shift-Tab (généralement suffisants)
            w.bind("<Tab>", go_next)
            w.bind("<Shift-Tab>", go_prev)
    
            # Certains environnements linux/old-tk envoient ISO_Left_Tab pour Shift-Tab.
            # On essaye de binder mais on ignore proprement l'erreur si le keysym n'existe pas.
            try:
                w.bind("<ISO_Left_Tab>", go_prev)
            except tk.TclError:
                # keysym non supporté sur cette plateforme : on ignore silencieusement
                pass
    
            # IMPORTANT: on NE bind pas <Return> sur les champs par défaut (sécurité)
            if enter_from_fields and submit_button is not None:
                w.bind("<Return>", lambda e, b=submit_button: b.invoke())
    
        # Sur le bouton : Enter et Espace déclenchent l'action (mais seulement si le bouton a le focus)
        if submit_button is not None:
            try:
                submit_button['takefocus'] = True
            except Exception:
                pass
            # binding sûr sur le bouton lui-même
            submit_button.bind("<Return>", lambda e, b=submit_button: b.invoke())
            submit_button.bind("<space>", lambda e, b=submit_button: b.invoke())
        
    
    def _get_active_stage(self) -> str | None:
        if not hasattr(self, "notebook"):
            return None
        tab_id = self.notebook.select()
        return self.tab_to_stage.get(tab_id)
    
    def _refresh_active_tab_now(self):
        if self._refreshing:
            return
        stage = self._get_active_stage()
        if not stage:
            return
        ref_fn = self.stage_refreshers.get(stage)
        if not callable(ref_fn):
            return
        self._refreshing = True
        try:
            ref_fn()  # ta fonction qui va en BDD et recharge le Treeview de l’onglet
        except Exception as e:
            print(f"[refresh:{stage}] {e}")
        finally:
            self._refreshing = False
    
    def _on_tab_changed(self, event=None):
        self._cancel_tab_refresh()
        self._refresh_active_tab_now()   # reload immédiat en arrivant sur le nouvel onglet
        self._focus_active_tab()         # <<< donne le focus dans le nouvel ongle
        self._schedule_next_tab_tick()   # redémarre le timer
    
    def _tab_tick(self):
        self._refresh_active_tab_now()
        self._schedule_next_tab_tick()
    
    def _schedule_next_tab_tick(self):
        # petit jitter pour éviter que 10 postes tapent la DB exactement en même temps
        import random
        jitter = random.randint(0, 1500)
        self._tab_refresh_job = self.after(self.refresh_ms + jitter, self._tab_tick)
    
    def _cancel_tab_refresh(self):
        if self._tab_refresh_job is not None:
            try:
                self.after_cancel(self._tab_refresh_job)
            except Exception:
                pass
            self._tab_refresh_job = None
            
    def register_focus_target(self, stage: str, widget):
        """À appeler depuis tes setup_* pour déclarer le champ à focus par défaut."""
        if widget is not None:
            self.focus_targets[stage] = widget
    
    def _find_first_input(self, container):
        """Fallback: cherche récursivement le premier Entry/Combobox/Text dans un frame."""
        for w in container.winfo_children():
            # si c'est un champ saisissable
            if isinstance(w, (ttk.Entry, tk.Entry, ttk.Combobox, tk.Text)):
                return w
            # si c'est un conteneur, on descend
            if isinstance(w, (ttk.Frame, tk.Frame, ttk.LabelFrame, ttk.Labelframe, ttk.Panedwindow)):
                found = self._find_first_input(w)
                if found:
                    return found
        return None
    
    def _focus_active_tab(self):
        """Donne le focus au widget cible de l'onglet actif (ou fallback)."""
        stage = self._get_active_stage()
        if not stage or not hasattr(self, "notebook"):
            return
    
        tab_id = self.notebook.select()
        tab = self.nametowidget(tab_id)
    
        target = self.focus_targets.get(stage)
        if target is None or not target.winfo_exists():
            target = self._find_first_input(tab)
    
        if target and target.winfo_exists():
            # after(0) pour que l'onglet ait fini de s'afficher
            self.after(0, lambda: (
                target.focus_set(),
                # si Entry/Combobox: on met le curseur à la fin + on sélectionne tout (optionnel)
                (hasattr(target, "icursor") and target.icursor('end')),
                (hasattr(target, "selection_range") and target.selection_range(0, 'end'))
            ))

    
    def set_photo(self, label: tk.Label, chemin_image: str, size=(200, 200)):
        """Charge une image et l'affecte AU SEUL label donné."""
        try:
            img = Image.open(chemin_image)
            img = img.resize(size)
            photo = ImageTk.PhotoImage(img)
            label.config(image=photo, text="")
            label.image = photo  # garder une référence pour éviter GC
        except Exception as e:
            messagebox.showerror("Erreur image", f"Impossible de charger l'image : {e}")
    
    def convert_comma_to_dot(self,event):
        # Fonction pour convertir les virgules en points dans la zone de saisie
        widget = event.widget
        text = widget.get()
        text = text.replace(',', '.')
        widget.delete(0, tk.END)
        widget.insert(0, text)
            
    def verfier_coherence_ref(self,num_batt):
        modele=self.selected_model
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "Select reference_produit_voltr from produit_voltr where numero_serie_produit = %s "
            param = (num_batt,)
            cursor.execute(query, param)
            modele_act=cursor.fetchone()[0]
            if modele_act != modele:
                reponse=messagebox.askyesno("Modèle de batterie différent",f"Le mdodèle de batterie n'est pas coherent \n Passer du modèle {modele_act} au modèle {modele} pour la batterie {num_batt} ?")
                if not reponse:
                    return 'stop'
                else :
                    cursor.execute("UPDATE produit_voltr SET reference_produit_voltr = %s WHERE numero_serie_produit =%s",(modele,num_batt))
                    messagebox.showinfo("Nouveau modèle", f'la batterie {num_batt} est passé au modèle {modele}.')
                    conn.commit()
                    return 'next'
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()  
    
    def changer_ref_batterie(self,new_ref,num_batt):
        modele=self.selected_model
        if modele == new_ref:
            messagebox.showinfo("Réference identique","Le nouveau modele est identique au modele actuel")
            return 
        else :
            conn = self.db_manager.connect()
            if not conn:
                return
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE produit_voltr SET reference_produit_voltr = %s WHERE numero_serie_produit =%s",(new_ref,num_batt))
                messagebox.showinfo("Nouveau modèle", f'la batterie {num_batt} est passé au modèle {new_ref}.')
                conn.commit()
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
            finally:
                try:
                    cursor.close()
                except:
                    pass
                conn.close()  
                
    def build_stage_query(self, current_stage):
        # Trouver l’index de l’étape courante
        idx = self.ordered_keys.index(current_stage)
    
        # Étapes précédentes
        previous_stages = self.ordered_keys[:idx]
    
        # Colonnes SQL associées
        prev_cols = [self.STAGE_TO_DBCOL[s] for s in previous_stages]
        current_col = self.STAGE_TO_DBCOL[current_stage]
    
        # Condition : toutes les étapes précédentes doivent être validées (=1)
        prev_conditions = " AND ".join([f"sp.{col} = 1" for col in prev_cols]) if prev_cols else "1=1"
    
        # Condition : l’étape courante doit être non validée (=0 ou NULL)
        current_condition = f"(sp.{current_col} = 0 OR sp.{current_col} IS NULL)"
    
        if current_stage != 'exp':
            # Construire la requête
            query = f"""
                SELECT sp.numero_serie_batterie
                FROM suivi_production AS sp
                JOIN produit_voltr AS pv
                  ON sp.numero_serie_batterie = pv.numero_serie_produit
                WHERE {prev_conditions}
                  AND {current_condition}
                  AND pv.reference_produit_voltr = %s
            """
        else: 
            # Construire la requête (pas de filtre ref produit)
            query = f"""
                SELECT sp.numero_serie_batterie
                FROM suivi_production AS sp
                JOIN produit_voltr AS pv
                  ON sp.numero_serie_batterie = pv.numero_serie_produit
                WHERE {prev_conditions}
                  AND {current_condition}
                  AND (recyclage=0 or recyclage is null)
            """
            
        return query

                
    #------------------------------ Onglet picking -----------------------------------------------
    
    #Front

    def setup_picking(self, frame):
        
        
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.numero_serie_cell_entry = ttk.Entry(left_frame)
        self.numero_serie_cell_entry.pack(pady=5)
        self.numero_serie_cell_entry.bind("<KeyRelease>", self.check_entry_length)
        
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.numero_serie_batt_entry = ttk.Entry(left_frame)
        self.numero_serie_batt_entry.pack(pady=5)
        
        
        
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        
        # --- Bloc dédié Listbox + Scrollbar ---
        listbox_frame = tk.Frame(left_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
        
        self.listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.listbox_batt.pack(side="left", fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
        
        self.listbox_batt.config(yscrollcommand=scrollbar.set)
        self.listbox_batt.bind("<<ListboxSelect>>", self.on_select_batt)
        # --- fin bloc listbox ---
        
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie,
            style="Danger.TButton"
        ).pack(pady=10)
        
        # --- Cadre droit inchangé ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        
        ttk.Label(right_frame, text="🔁 Remplacement cellule:").pack(pady=5)
        ttk.Label(right_frame, text="N° série cellule:").pack(pady=5)
        self.numero_cell_r_entry = ttk.Entry(right_frame)
        self.numero_cell_r_entry.pack(pady=5)
        
        ttk.Label(right_frame, text="Défaut:").pack(pady=5)
        self.combobox_default = ttk.Combobox(
            right_frame, state="readonly",
            values=["Non trouvée", "Tension", "Corrosion", "Déformation"]
        )
        self.combobox_default.pack(pady=5)
        self.combobox_default.bind("<<ComboboxSelected>>", self.tension_defaut)
        
        ttk.Label(right_frame, text="Tension cellule NOK:").pack(pady=5)
        self.tension_cell_entry = ttk.Entry(right_frame, state="disabled")
        self.tension_cell_entry.pack(pady=5)
        self.tension_cell_entry.bind("<KeyRelease>", self.convert_comma_to_dot)
        
        ttk.Button(
            right_frame, text="🔄 Demande remplacement cellule",
            command=self.replace_cellule, style="Danger.TButton"
        ).pack(pady=10)
        
        frame_info = tk.Frame(right_frame, width=300, height=100, bg='#D0F5BE')
        frame_info.pack(pady=20)
        frame_info.pack_propagate(False)
        tk.Label(
            frame_info, text="⚠ Écart maximum de 0.05V",
            bg='#D0F5BE', fg='black',
            font=("Segoe UI", 11, 'bold')
        ).pack(expand=True)
        
        self.btn_valider_picking = ttk.Button(
        right_frame, text="✅ Contrôle OK",
        command=self.valider_picking, style="Good.TButton"
        )
        self.btn_valider_picking.pack(pady=10)
        
        self.display_model_list()
        
        self.make_tab_chain(
            [
                self.numero_serie_batt_entry,
                self.btn_valider_picking  # placer le bouton en dernier si tu veux que Tab atteigne le bouton
            ],
            submit_button=self.btn_valider_picking,
            ring=True,
            enter_from_fields=False  # sécurité : Enter depuis les champs ne déclenche pas le bouton
        )
        
        return self.numero_serie_batt_entry

            
    #Back
    
    def display_model_list(self):
        stage_act='picking'
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            if modele[:8]=="PPTR018A":
                query="""
                SELECT sp.numero_serie_batterie
                FROM suivi_production AS sp
                JOIN produit_voltr AS pv
                  ON sp.numero_serie_batterie = pv.numero_serie_produit
                WHERE pv.reference_produit_voltr LIKE %s
                AND (sp.picking_tension = 0 or sp.picking_tension is null)
                AND (sp.recyclage = 0 or sp.recyclage is null)
                  """             
                param=(modele[:8]+"%",)
                
                
            else :
                query=self.build_stage_query(stage_act)
                param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
            
            try:
                first_frac, last_frac = self.listbox_batt.yview()
            except Exception:
                first_frac = 0.0
    
            # Vide la Listbox
            self.listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.listbox_batt.insert(tk.END, batt)
            
            try:
                self.listbox_batt.yview_moveto(first_frac)
            except Exception:
                pass
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
    def on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.numero_serie_batt_entry.delete(0, tk.END)
        self.numero_serie_batt_entry.insert(0, selected_value)
                      
    def add_non_conf_batterie(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if reponse:
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 

        num_batt=str(self.numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension_fail = picking_tension_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.numero_serie_batt_entry.delete(0, tk.END)
            
    def valider_picking(self):
        num_batt=str(self.numero_serie_batt_entry.get())
        if not num_batt:
            return
        if str(self.selected_model)[:8]!='PPTR018A':
            directive=self.verfier_coherence_ref(num_batt)
            if directive=='stop':
                return
        # 1) Pré-requis
        stage_col='picking'
        if not self.verif_etape_act_non_ok(stage_col, num_batt):
            return
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        conn = self.db_manager.connect()
        visa=self.db_manager.user
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension = 1, date_picking_tension = NOW(), visa_picking_tension = %s where numero_serie_batterie = %s "
            param = (visa,num_batt)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.numero_serie_batt_entry.delete(0, tk.END)
            """
            self.display_model_list()
            self.display_model_list_pack()
            self.display_model_list_nappe()
            self.display_model_list_bms()
            self.display_model_list_wrap()
            self.display_model_list_fermeture()
            self.display_model_list_emballage()
            self.display_model_list_exp()
            """
            
            self._focus_active_tab()
            for f in self.funcs_to_run:
                f()
             
    def tension_defaut(self,event=None):
        defaut=str(self.combobox_default.get())
        if defaut=="Tension":
            self.tension_cell_entry.config(state="normal")
        else :
            self.tension_cell_entry.config(state="disabled")
                       
    def replace_cellule(self):

        self.remplacement_dir = r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc"
        os.makedirs(self.remplacement_dir, exist_ok=True)
        # Template 
        self.template_remplacement_path = r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\Template remplacement cellule (13).xlsx"
        
        num = self.numero_cell_r_entry.get().strip()  
        defaut = self.combobox_default.get().strip()                  
        tension_txt = self.tension_cell_entry.get().strip()    
        
        if not num:
            messagebox.showwarning("Champ manquant", "Renseigne le N° de série cellule.")
            return
        if not defaut:
            messagebox.showwarning("Champ manquant", "Sélectionne un défaut.")
            return
        if defaut.lower().startswith("tension"):
            if not tension_txt:
                messagebox.showwarning("Champ manquant",
                                       "Défaut = Tension ➜ renseigne la tension NOK.")
                return
            # optionnel: vérifier numérique
            try:
                float(tension_txt.replace(",", "."))
            except ValueError:
                messagebox.showwarning("Valeur invalide",
                                       "La tension doit être un nombre (ex: 3.72).")
                return
    
        projet = None
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
    
            cur.execute("""
                SELECT pv.numero_projet,pv.numero_serie_produit FROM produit_voltr as pv join cellule as c on pv.numero_serie_produit=c.affectation_produit WHERE c.numero_serie_cellule = %s
            """, (num,))
            row = cur.fetchone()
            if row and row[0]:
                projet = str(row[0]).strip()
                produit = str(row[1]).strip()
            else:
                messagebox.showerror("Projet introuvable",
                                     f"Aucun projet trouvé pour la cellule {num}.")
                return
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer le projet :\n{e}")
            return
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass
    
        today_str = datetime.now().strftime("%Y-%m-%d")
        safe_projet = re.sub(r"[^\w\- ]+", "_", projet)  # sécurité nom de fichier
        out_path = os.path.join(self.remplacement_dir, f"{safe_projet}-{today_str}.xlsx")
    
        tension_val = ""
        if defaut.lower().startswith("tension"):
            # conserver le format texte tel que saisi, ou caster en float si tu préfères
            tension_val = tension_txt.replace(",", ".")
    
        new_row = [produit, num, defaut, tension_val]
    
        try:
            if not os.path.exists(out_path):
                if not os.path.exists(self.template_remplacement_path):
                    messagebox.showerror("Template manquant",
                                         f"Template introuvable :\n{self.template_remplacement_path}")
                    return
                shutil.copyfile(self.template_remplacement_path, out_path)
    
            wb = load_workbook(out_path)

            SHEET_CANDIDATES = ["Remplacement", "Remplacements", "Feuil1", "Données"]
            ws = None
            for name in SHEET_CANDIDATES:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:

                expected_headers = ["Date", "N° série", "Défaut", "Tension"]
                for sh in wb.worksheets:
                    headers = [ (sh.cell(row=1, column=i).value or "").strip() for i in range(1, 5) ]
                    if [h.lower() for h in headers] == [h.lower() for h in expected_headers]:
                        ws = sh
                        break
            if ws is None:
              
                ws = wb.active
    
            ws.append(new_row)

            max_row = ws.max_row
            max_col = ws.max_column
            last_col_letter = get_column_letter(max_col)
    
            if hasattr(ws, "tables"):
                for tbl in list(ws.tables.values()):
                    ref = tbl.ref
                    start, end = ref.split(":")
                    start_col = "".join([c for c in start if c.isalpha()])
                    start_row = int("".join([c for c in start if c.isdigit()]))

                    new_ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
                    tbl.ref = new_ref
    
            if ws.auto_filter and ws.auto_filter.ref:
                start, _ = ws.auto_filter.ref.split(":")
                start_col = "".join([c for c in start if c.isalpha()])
                start_row = int("".join([c for c in start if c.isdigit()]))
                ws.auto_filter.ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
    
            wb.save(out_path)
    
        except Exception as e:
            messagebox.showerror("Erreur Excel", f"Impossible d'écrire dans le fichier :\n{e}")
            return

        try:
            self.numero_cell_r_entry.delete(0, 'end')
            self.combobox_default.set('')
            self.tension_cell_entry.delete(0, 'end')
        except Exception:
            pass
    
        messagebox.showinfo("Remplacement enregistré", f"Ligne ajoutée dans :\n{out_path}")
        
        
    
    def check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.numero_serie_batt_entry.delete(0, tk.END)
                self.numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
            
    #------------------------------ Onglet soudure pack -----------------------------------------------    
    
    #Front
    
    def setup_pack(self, frame):
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.s_numero_serie_cell_entry = ttk.Entry(left_frame)
        self.s_numero_serie_cell_entry.pack(pady=5)
        self.s_numero_serie_cell_entry.bind("<KeyRelease>", self.s_check_entry_length)
    
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.s_numero_serie_batt_entry = ttk.Entry(left_frame)
        self.s_numero_serie_batt_entry.pack(pady=5)
        
        if str(self.selected_model)[:8]=='PPTR018A':
        
            ttk.Label(left_frame, text="Choisir une reference EOP").pack(pady=5)
            
            VALS = ["PPTR018AA", "PPTR018AB", "PPTR018AC","PPTR018AD"]

            # Variable liée
            self.choice = tk.StringVar(value=VALS[0])
                
            self.s_mod_combobox=ttk.Combobox(left_frame, textvariable=self.choice, values=VALS, state="readonly", width=20)
            self.s_mod_combobox.insert(0,str(self.selected_model))  # sélectionne l'élément par défaut
            self.s_mod_combobox.pack(pady=5)
            
            ttk.Label(left_frame, text="Mesure d'impedance (Ohms)").pack(pady=5)
            self.impedance_entry= ttk.Entry(left_frame)
            self.impedance_entry.pack(pady=5)
            self.impedance_entry.bind("<KeyRelease>", self.convert_comma_to_dot)
            
            
            ttk.Label(left_frame, text="Mesure tension (V)").pack(pady=5)
            self.tension_eop_entry= ttk.Entry(left_frame)
            self.tension_eop_entry.pack(pady=5)
            self.tension_eop_entry.bind("<KeyRelease>", self.convert_comma_to_dot)
            
            
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie_pack,
            style="Danger.TButton"
        ).pack(pady=10)
    
        # --- Frame droite ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
    
        # --- Bloc Listbox + Scrollbar ---
        listbox_frame = tk.Frame(right_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
    
        self.s_listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.s_listbox_batt.pack(side="left", fill="both", expand=True)
    
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.s_listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
    
        self.s_listbox_batt.config(yscrollcommand=scrollbar.set)
        self.s_listbox_batt.bind("<<ListboxSelect>>", self.s_on_select_batt)
        # --- fin bloc listbox ---
    
        self.btn_valider_pack = ttk.Button(
            left_frame, text="✅ Contrôle OK",
            command=self.valider_soudure_pack,
            style="Good.TButton"
        )
        self.btn_valider_pack.pack(pady=10)
        
        self.display_model_list_pack()
        
        self.make_tab_chain(
            [
                self.s_numero_serie_batt_entry,
                self.btn_valider_pack  # placer le bouton en dernier si tu veux que Tab atteigne le bouton
            ],
            submit_button=self.btn_valider_pack,
            ring=True,
            enter_from_fields=False  # sécurité : Enter depuis les champs ne déclenche pas le bouton
        )   
        
        return self.s_numero_serie_batt_entry

    #Back
    
    def s_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.s_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.s_numero_serie_batt_entry.delete(0, tk.END)
                self.s_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.s_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_pack(self):
        stage_act='pack'
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            if modele[:8]=="PPTR018A":
                query ="""
                    SELECT sp.numero_serie_batterie
                    FROM suivi_production AS sp
                    JOIN produit_voltr AS pv
                      ON sp.numero_serie_batterie = pv.numero_serie_produit
                    WHERE sp.picking_tension = 1
                      AND (sp.soudure_pack=0 or sp.soudure_pack is null)
                      AND (sp.recyclage = 0 or sp.recyclage is null)
                      AND pv.reference_produit_voltr like %s
                     """
                param=(modele[:8]+"%",)
                
                
            else :
                
                query=self.build_stage_query(stage_act)
                param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.s_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.s_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_soudure_pack(self):
        
        modele=str(self.selected_model)
        if modele[:8]=='PPTR018A':
            new_modele=self.s_mod_combobox.get()
            if new_modele=="— choisir —":
                messagebox.showerror("Modele EOP","Choisir un modele d'EOP")
                return
        num_batt=str(self.s_numero_serie_batt_entry.get())
        # 1) Pré-requis
        stage_col='pack'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        conn = self.db_manager.connect()
        visa=self.db_manager.user
        if not conn:
            return
        if modele[:8]!='PPTR018A':
            directive=self.verfier_coherence_ref(num_batt)
            if directive=='stop':
                return
            
        try:
            cursor = conn.cursor()
            if modele[:8]=='PPTR018A':
                impedance=self.impedance_entry.get()
                controle_tension=self.tension_eop_entry.get()
                if impedance:
                    impedance=float(impedance)
                else :
                    messagebox.showerror("Impedance non renseignée", "Veuillez renseigner l'impedance")
                    
                if controle_tension:
                    controle_tension=float(controle_tension)
                else :
                    messagebox.showerror("Tension non renseignée", "Veuillez renseigner la tension")
                df_cyclage = pd.read_excel(EXCEL_PATH,sheet_name="Cyclage",header=1)
                try:
                    query = """
                        SELECT reference_cellule
                        FROM cellule
                        WHERE affectation_produit = %s
                        LIMIT 1
                    """
                    cursor.execute(query, (num_batt,))
                    row_db = cursor.fetchone()
                except Exception as e:
                    # Erreur SQL => échec de traitement
                    messagebox.showerror("Reference cellule introuvable !", f"Pas de reference cellule pour la batterie {num_batt}")
                    row_db = None
                    return
                
                if not row_db or not row_db[0]:
                    messagebox.showerror("Reference cellule introuvable !", f"Pas de reference cellule pour la batterie {num_batt}")   
                    return
                    
                ref_cell = row_db[0]
                # --- Seuils (df_cyclage) ---
                row = df_cyclage[
                    (df_cyclage["Nom_modele"] == new_modele) &
                    (df_cyclage["Ref cellule"] == ref_cell)
                ]

                if row.empty:
                    messagebox.showerror("Duo batterie/cellule introuvable !", f"Pas de batterie {num_batt} avec la cellule {ref_cell} dans le suivi de production")   
                    return
                    
                seuils = row.iloc[0]
                try:
                    imp_min = float(seuils["Impedance borne min"])
                except Exception:
                    imp_min = float(str(seuils["Impedance borne min"]).replace(",", "."))
                try:
                    imp_max = float(seuils["Impedance borne max"])
                except Exception:
                    imp_max = float(str(seuils["Impedance borne max"]).replace(",", "."))
                
                if imp_min <= impedance <= imp_max :
                    query = "UPDATE suivi_production SET soudure_pack = 1, date_soudure_pack = NOW(),test_impedance= %s, date_impedance=NOW(), test_tension=%s, date_test_tension= NOW(), visa_soudure_pack= %s where numero_serie_batterie = %s "
                    param = (impedance,controle_tension,visa,num_batt)  
                    cursor.execute(query, param)
                    messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
                else :
                    messagebox.showerror("Erreur impedance", "La valeure d'impedance est NOK")
                    return
                    
            else:    
                query = "UPDATE suivi_production SET soudure_pack = 1, date_soudure_pack = NOW(), visa_soudure_pack= %s where numero_serie_batterie = %s "
                param = (visa,num_batt)  
                cursor.execute(query, param)
                messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
                
            if modele[:8]=='PPTR018A':
            
                query_modele='UPDATE produit_voltr set reference_produit_voltr= %s where numero_serie_produit = %s'
                param_modele=(new_modele,num_batt)
                cursor.execute(query_modele, param_modele)

        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.s_numero_serie_batt_entry.delete(0, tk.END)
            self._focus_active_tab()
            for f in self.funcs_to_run:
                f()
            
            
    def add_non_conf_batterie_pack(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if reponse:
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
           
        
        num_batt=str(self.s_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack_fail = soudure_pack_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.s_numero_serie_batt_entry.delete(0, tk.END)
    
    def s_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.s_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.s_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.s_numero_serie_batt_entry.delete(0, tk.END)
        self.s_numero_serie_batt_entry.insert(0, selected_value)
            
    
        
    #------------------------------ Onglet soudure nappe -----------------------------------------------   

    def setup_nappe(self, frame):
        
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.n_numero_serie_cell_entry = ttk.Entry(left_frame)
        self.n_numero_serie_cell_entry.pack(pady=5)
        self.n_numero_serie_cell_entry.bind("<KeyRelease>", self.n_check_entry_length)
    
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.n_numero_serie_batt_entry = ttk.Entry(left_frame)
        self.n_numero_serie_batt_entry.pack(pady=5)
    
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie_nappe,
            style="Danger.TButton"
        ).pack(pady=10)
    
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
    
        # --- Bloc Listbox + Scrollbar ---
        listbox_frame = tk.Frame(left_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
    
        self.n_listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.n_listbox_batt.pack(side="left", fill="both", expand=True)
    
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.n_listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
    
        self.n_listbox_batt.config(yscrollcommand=scrollbar.set)
        self.n_listbox_batt.bind("<<ListboxSelect>>", self.n_on_select_batt)
        # --- fin bloc listbox ---
    
        # --- Frame droite ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(right_frame, text="Ecart tension modules:").pack(pady=5)
        self.ecart_t_entry = ttk.Entry(right_frame)
        self.ecart_t_entry.pack(pady=5)
        self.ecart_t_entry.bind("<KeyRelease>", self.convert_comma_to_dot)
    
        self.n_label_photo = tk.Label(
            right_frame,
            bg="#e0e0e0",
            width=200,
            height=200,
            text="Aperçu photo",
            anchor='center'
        )
        self.n_label_photo.pack(pady=10)
        self.set_photo(
            self.n_label_photo,
            r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\voltr_logo.jpg"
        )
    
        self.btn_valider_nappe = ttk.Button(
            right_frame, text="✅ Contrôle OK",
            command=self.valider_soudure_nappe,
            style="Good.TButton"
        )
        self.btn_valider_nappe.pack(pady=10)
        
        self.make_tab_chain(
            [
                self.n_numero_serie_batt_entry,
                self.btn_valider_nappe  # placer le bouton en dernier si tu veux que Tab atteigne le bouton
            ],
            submit_button=self.btn_valider_nappe,
            ring=True,
            enter_from_fields=False  # sécurité : Enter depuis les champs ne déclenche pas le bouton
        )
    
        self.display_model_list_nappe()
        
        return self.n_numero_serie_batt_entry
            
        #Back
        
    def n_check_entry_length(self, event=None):

        numero_serie_cell = self.n_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.n_numero_serie_batt_entry.delete(0, tk.END)
                self.n_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.n_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_nappe(self):
        stage_act='nappe'
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query=self.build_stage_query(stage_act)
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
            
            try:
                first_frac, last_frac = self.n_listbox_batt.yview()
            except Exception:
                first_frac = 0.0

    
            # Vide la Listbox
            self.n_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.n_listbox_batt.insert(tk.END, batt)
                
            # Restaure la position de scroll (clamped automatiquement par Tk)
            try:
                self.n_listbox_batt.yview_moveto(first_frac)
            except Exception:
                pass
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_soudure_nappe(self):
        num_batt=str(self.n_numero_serie_batt_entry.get())
        # 1) Pré-requis
        stage_col='nappe'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        if self.ecart_t_entry.get():
            delta_tension_str=self.ecart_t_entry.get()
            delta_t=delta_tension_str.replace(",", ".")
            delta_tension=float(delta_t)
        else: 
            messagebox.showerror("Saisie incompléte !","Renseigner l'ecart de tension entre le module min et le module max")
            return
        conn = self.db_manager.connect()
        visa=self.db_manager.user
        if not conn:
            return
        
        if delta_tension > 0.05:
            reponse = messagebox.askyesno("Non conformité", "Ecart de tension trop élevé, ouvrir une non-conformité ?")
            if not reponse:
                return
            
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
            
            num_batt=str(self.n_numero_serie_batt_entry.get())
            
            self.verfier_coherence_ref(num_batt)
            
            try:
                cursor = conn.cursor()
                query = "UPDATE suivi_production SET soudure_pack_fail = soudure_nappe_fail + 1 where numero_serie_batterie = %s "
                param = (num_batt,)
                cursor.execute(query, param)
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
            finally:
                try:
                    conn.commit()
                    cursor.close()
                except:
                    pass
                conn.close()
                self.n_numero_serie_batt_entry.delete(0, tk.END)
                self.ecart_t_entry.delete(0, tk.END)
                    
        else :
            try:
                cursor = conn.cursor()
                query = "UPDATE suivi_production SET soudure_nappe = 1,delta_tension_module = %s, date_soudure_nappe = NOW(), visa_soudure_nappe= %s where numero_serie_batterie = %s "
                param = (delta_tension,visa,num_batt)  
                cursor.execute(query, param)
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
            finally:
                try:
                    conn.commit()
                    cursor.close()
                except:
                    pass
                conn.close()
                messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
                self.n_numero_serie_batt_entry.delete(0, tk.END)
                self.ecart_t_entry.delete(0, tk.END)
                self._focus_active_tab()
                for f in self.funcs_to_run:
                    f()
            
    def add_non_conf_batterie_nappe(self):
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.n_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack_fail = soudure_nappe_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.n_numero_serie_batt_entry.delete(0, tk.END)
            self.ecart_t_entry.delete(0, tk.END)
        
    def n_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.n_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.n_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.n_numero_serie_batt_entry.delete(0, tk.END)
        self.n_numero_serie_batt_entry.insert(0, selected_value)
        
    #------------------------------ Onglet bms -----------------------------------------------   

    def setup_bms(self, frame):
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.b_numero_serie_cell_entry = ttk.Entry(left_frame)
        self.b_numero_serie_cell_entry.pack(pady=5)
        self.b_numero_serie_cell_entry.bind("<KeyRelease>", self.b_check_entry_length)
    
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.b_numero_serie_batt_entry = ttk.Entry(left_frame)
        self.b_numero_serie_batt_entry.pack(pady=5)
    
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie_bms,
            style="Danger.TButton"
        ).pack(pady=10)
    
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
    
        # --- Bloc Listbox + Scrollbar ---
        listbox_frame = tk.Frame(left_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
    
        self.b_listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.b_listbox_batt.pack(side="left", fill="both", expand=True)
    
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.b_listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
    
        self.b_listbox_batt.config(yscrollcommand=scrollbar.set)
        self.b_listbox_batt.bind("<<ListboxSelect>>", self.b_on_select_batt)
        # --- fin bloc listbox ---
    
        # --- Frame droite ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
    
        self.b_label_photo = tk.Label(
            right_frame,
            bg="#e0e0e0",
            width=200,
            height=200,
            text="Aperçu photo",
            anchor='center'
        )
        self.b_label_photo.pack(pady=10)
    
        self.set_photo(
            self.b_label_photo,
            r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\voltr_logo.jpg"
        )
    
        self.btn_valider_bms=ttk.Button(
            right_frame, text="✅ Contrôle OK",
            command=self.valider_bms,
            style="Good.TButton"
        )
        self.btn_valider_bms.pack(pady=10)
        
        
        self.make_tab_chain(
            [
                self.b_numero_serie_batt_entry,
                self.btn_valider_bms  # placer le bouton en dernier si tu veux que Tab atteigne le bouton
            ],
            submit_button=self.btn_valider_bms,
            ring=True,
            enter_from_fields=False  # sécurité : Enter depuis les champs ne déclenche pas le bouton
        )
    
        self.display_model_list_bms()
        
        return self.b_numero_serie_batt_entry

    def b_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.b_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.b_numero_serie_batt_entry.delete(0, tk.END)
                self.b_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.b_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_bms(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        stage_act='bms'
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query=self.build_stage_query(stage_act)
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.b_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.b_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_bms(self):
        num_batt=str(self.b_numero_serie_batt_entry.get())
        stage_col='bms'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        conn = self.db_manager.connect()
        visa=self.db_manager.user
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_bms = 1, date_soudure_bms = NOW(), visa_soudure_bms= %s where numero_serie_batterie = %s"
            param = (visa,num_batt)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.b_numero_serie_batt_entry.delete(0, tk.END)
            self._focus_active_tab()
            for f in self.funcs_to_run:
                f()
            
            
    def add_non_conf_batterie_bms(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
                
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.b_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_bms_fail = soudure_bms_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.b_numero_serie_batt_entry.delete(0, tk.END)
    
    def b_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.b_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.b_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.b_numero_serie_batt_entry.delete(0, tk.END)
        self.b_numero_serie_batt_entry.insert(0, selected_value)
        
    
        
    #------------------------------ Onglet wrap -----------------------------------------------   

    def setup_wrap(self, frame):
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.w_numero_serie_cell_entry = ttk.Entry(left_frame)
        self.w_numero_serie_cell_entry.pack(pady=5)
        self.w_numero_serie_cell_entry.bind("<KeyRelease>", self.w_check_entry_length)
    
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.w_numero_serie_batt_entry = ttk.Entry(left_frame)
        self.w_numero_serie_batt_entry.pack(pady=5)
    
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie_wrap,
            style="Danger.TButton"
        ).pack(pady=10)
    
        # --- Frame droite ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
    
        # --- Bloc Listbox + Scrollbar ---
        listbox_frame = tk.Frame(right_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
    
        self.w_listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.w_listbox_batt.pack(side="left", fill="both", expand=True)
    
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.w_listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
    
        self.w_listbox_batt.config(yscrollcommand=scrollbar.set)
        self.w_listbox_batt.bind("<<ListboxSelect>>", self.w_on_select_batt)
        # --- fin bloc listbox ---
    
        self.btn_valider_wrap=ttk.Button(
            left_frame, text="✅ Contrôle OK",
            command=self.valider_wrap,
            style="Good.TButton"
        )
        self.btn_valider_wrap.pack(pady=10)
        
        self.make_tab_chain(
            [
                self.w_numero_serie_batt_entry,
                self.btn_valider_wrap  # placer le bouton en dernier si tu veux que Tab atteigne le bouton
            ],
            submit_button=self.btn_valider_wrap,
            ring=True,
            enter_from_fields=False  # sécurité : Enter depuis les champs ne déclenche pas le bouton
        )
    
    
        self.display_model_list_wrap()
        
        return self.w_numero_serie_batt_entry

        
    def w_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.b_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.b_numero_serie_batt_entry.delete(0, tk.END)
                self.b_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.b_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_wrap(self):
        stage_act='wrap'
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query=self.build_stage_query(stage_act)
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.w_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.w_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_wrap(self):
        num_batt=str(self.w_numero_serie_batt_entry.get())
        stage_col='wrap'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        conn = self.db_manager.connect()
        visa=self.db_manager.user
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET wrap = 1, date_wrap = NOW(), visa_wrap= %s where numero_serie_batterie = %s "
            param = (visa,num_batt)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.w_numero_serie_batt_entry.delete(0, tk.END)
            self._focus_active_tab()
            for f in self.funcs_to_run:
                f()
            
            
    def add_non_conf_batterie_wrap(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
                
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.w_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET wrap_fail = wrap_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.w_numero_serie_batt_entry.delete(0, tk.END)
    
    def w_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.w_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.w_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.w_numero_serie_batt_entry.delete(0, tk.END)
        self.w_numero_serie_batt_entry.insert(0, selected_value)
        
        
    #------------------------------ Onglet fermeture -----------------------------------------------   
    def setup_fermeture(self, frame):
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.f_numero_serie_cell_entry = ttk.Entry(left_frame)
        self.f_numero_serie_cell_entry.pack(pady=5)
        self.f_numero_serie_cell_entry.bind("<KeyRelease>", self.f_check_entry_length)
    
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.f_numero_serie_batt_entry = ttk.Entry(left_frame)
        self.f_numero_serie_batt_entry.pack(pady=5)
    
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie_fermeture,
            style="Danger.TButton"
        ).pack(pady=10)
    
        ttk.Label(left_frame, text="Tension en fin de test:").pack(pady=5)
        self.tension_end_entry = ttk.Entry(left_frame)
        self.tension_end_entry.pack(pady=5)
        self.tension_end_entry.bind("<KeyRelease>", self.convert_comma_to_dot)
    
        self.btn_valider_fermeture=ttk.Button(
            left_frame, text="✅ Contrôle OK",
            command=self.valider_fermeture,
            style="Good.TButton"
        )
        self.btn_valider_fermeture.pack(pady=10)
        
        self.make_tab_chain(
            [
                self.f_numero_serie_batt_entry,
                self.btn_valider_fermeture  # placer le bouton en dernier si tu veux que Tab atteigne le bouton
            ],
            submit_button=self.btn_valider_fermeture,
            ring=True,
            enter_from_fields=False  # sécurité : Enter depuis les champs ne déclenche pas le bouton
        )
        
        
    
        # --- Frame droite ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
    
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
    
        # --- Bloc Listbox + Scrollbar ---
        listbox_frame = tk.Frame(right_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
    
        self.f_listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.f_listbox_batt.pack(side="left", fill="both", expand=True)
    
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.f_listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
    
        self.f_listbox_batt.config(yscrollcommand=scrollbar.set)
        self.f_listbox_batt.bind("<<ListboxSelect>>", self.f_on_select_batt)
        # --- fin bloc listbox ---
    
        self.f_label_photo = tk.Label(
            left_frame,
            bg="#e0e0e0",
            width=200,
            height=200,
            text="Aperçu photo",
            anchor='center'
        )
        self.f_label_photo.pack(pady=10)
    
        self.set_photo(
            self.f_label_photo,
            r"G:\Drive partagés\VoltR\11_Data\IHM\Executable\IHM_suivi_prod_beta\Suivi_prod_rsc\voltr_logo.jpg"
        )
    
        self.display_model_list_fermeture()
        
        return self.f_numero_serie_batt_entry

        
    def f_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.f_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.f_numero_serie_batt_entry.delete(0, tk.END)
                self.f_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.f_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_fermeture(self):
        stage_act='fermeture_batt'
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query=self.build_stage_query(stage_act)
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.f_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.f_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_fermeture(self):
        num_batt=str(self.f_numero_serie_batt_entry.get())
        stage_col='fermeture'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        directive=self.verfier_coherence_ref(num_batt)
        if directive=='stop':
            return
        conn = self.db_manager.connect()
        visa=self.db_manager.user
        tension_fin=self.tension_end_entry.get()
        tension_f=tension_fin.replace(",",".")
        tension_end=float(tension_f)
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET fermeture_batt = 1, date_fermeture_batt = NOW(), test_tension= %s, visa_fermeture= %s where numero_serie_batterie = %s "
            param = (tension_end,visa,num_batt)  
            cursor.execute(query, param)
            
            cursor.execute("UPDATE produit_voltr SET statut = %s where numero_serie_produit= %s",('stock',num_batt))

        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.f_numero_serie_batt_entry.delete(0, tk.END)
            self._focus_active_tab()
            for f in self.funcs_to_run:
                f()
            
            
    def add_non_conf_batterie_fermeture(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
                
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.f_numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET fermeture_fail = fermeture_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.f_numero_serie_batt_entry.delete(0, tk.END)
    
    def f_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.f_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.f_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.f_numero_serie_batt_entry.delete(0, tk.END)
        self.f_numero_serie_batt_entry.insert(0, selected_value)
        
        

    #------------------------------ Onglet emballage -----------------------------------------------   
    
    def setup_emb(self, frame):
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        
        # --- Frame droite ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.emb_numero_serie_batt_entry = ttk.Entry(left_frame)
        self.emb_numero_serie_batt_entry.pack(pady=5)
        self.emb_numero_serie_batt_entry.bind("<KeyRelease>", self.emb_check_entry_length)
        
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
        
        # --- Bloc Listbox + Scrollbar ---
        listbox_frame = tk.Frame(right_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
        
        self.emb_listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.emb_listbox_batt.pack(side="left", fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.emb_listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
        
        self.emb_listbox_batt.config(yscrollcommand=scrollbar.set)
        self.emb_listbox_batt.bind("<<ListboxSelect>>", self.emb_on_select_batt)
        # --- fin bloc listbox ---
        
        ttk.Label(left_frame, text="Modèle batterie").pack(pady=5)
        
        cb_var_m = tk.StringVar()
        self.cb_emb = ttk.Combobox(
            left_frame,
            textvariable=cb_var_m,
            values=self.models,
            state="disabled",
            width=40
        )
        self.cb_emb.pack(pady=5)
        
        self.chk_var_emb = tk.BooleanVar(value=False)
        
        def toggle_combobox():
            self.cb_emb.configure(state="readonly" if self.chk_var_emb.get() else "disabled")
        
        self.chk_emb = ttk.Checkbutton(
            left_frame,
            text="Changer le modèle",
            variable=self.chk_var_emb,
            command=toggle_combobox
        )
        self.chk_emb.pack(pady=5)
        
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie_emb,
            style="Danger.TButton"
        ).pack(pady=10)
        
        ttk.Button(
            left_frame, text="✅ Contrôle OK",
            command=self.valider_emballage,
            style="Good.TButton"
        ).pack(pady=10)
        
        self.display_model_list_emballage()

    def emb_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.emb_numero_serie_batt_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
                self.emb_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_emballage(self):
        stage_act='emb'
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query=self.build_stage_query(stage_act)
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.emb_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.emb_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_emballage(self):
        num_batt=str(self.emb_numero_serie_batt_entry.get())
        stage_col='emb'
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        if self.chk_var_emb.get():
            new_ref=str(self.cb_emb.get())
            self.changer_ref_batterie(new_ref,num_batt)
        else:
            directive=self.verfier_coherence_ref(num_batt)
            if directive=='stop':
                return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET emballage = 1, date_emballage= NOW() where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.emb_numero_serie_batt_entry.delete(0, tk.END)
            for f in self.funcs_to_run:
                f()
            
    def add_non_conf_batterie_emb(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
                
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from) 
        
        self.emb_numero_serie_batt_entry.delete(0, tk.END)
    
    def emb_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.emb_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.emb_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.emb_numero_serie_batt_entry.delete(0, tk.END)
        self.emb_numero_serie_batt_entry.insert(0, selected_value)
    
    #------------------------------ Onglet capa -----------------------------------------------   

    def setup_capa(self, frame):
        
        # Conteneur principal
        wrap = ttk.LabelFrame(frame, text="Test OK & En attente de cyclage", padding=12)
        wrap.pack(fill="both", expand=True, padx=8, pady=8)
        
        # ====== Tableau PRINCIPAL (tests OK) + mini treeview à droite ======
        table_frame = ttk.Frame(wrap)
        table_frame.pack(fill="both", expand=True, padx=8, pady=(6, 12))
    
        # --- Colonne de droite : mini treeview "N° Série" ---
        right_frame = ttk.Frame(table_frame)
        right_frame.pack(side="right", fill="y", padx=(6, 0))  # se place tout à droite
    
        self.ok_series_tree = ttk.Treeview(
            right_frame,
            columns=("N° Série",),
            show="headings",
            selectmode="browse",
            height=12  # petit format
        )
        self.ok_series_tree.heading("N° Série", text="N° Série")
        self.ok_series_tree.column("N° Série", anchor="w", width=180, stretch=True)
    
        yscroll_ok_series = ttk.Scrollbar(right_frame, orient="vertical", command=self.ok_series_tree.yview)
        self.ok_series_tree.configure(yscrollcommand=yscroll_ok_series.set)
    
        self.ok_series_tree.pack(side="left", fill="y")
        yscroll_ok_series.pack(side="right", fill="y")
    
        # --- Tableau principal (tests OK) ---
        cols_ok = ("N° Série", "Modèle", "Capacité", "Tension de fin de test", "Emplacement")
        self.test_tree = ttk.Treeview(table_frame, columns=cols_ok, show="headings", selectmode="browse")
        for c in cols_ok:
            self.test_tree.heading(c, text=c)
            if c == "Tension de fin de test":
                width = 170
            elif c in ("Emplacement", "Modèle"):
                width = 160
            else:
                width = 140
            self.test_tree.column(c, anchor="w", width=width, stretch=True)
    
        yscroll_ok = ttk.Scrollbar(table_frame, orient="vertical", command=self.test_tree.yview)
        self.test_tree.configure(yscrollcommand=yscroll_ok.set)
    
        # ordre des pack: scrollbar (droite), table (gauche)
        yscroll_ok.pack(side="right", fill="y")
        self.test_tree.pack(side="left", fill="both", expand=True)
    
        # ====== Tableau SECONDAIRE (tests défaillants) ======
        failed_box = ttk.LabelFrame(wrap, text="Tests défaillants & Non prêtes", padding=8)
        failed_box.pack(fill="both", expand=True, padx=8, pady=(0, 12))
    
        failed_box.columnconfigure(0, weight=1)
        failed_box.columnconfigure(1, weight=1)
    
        # --- Partie gauche : Tests défaillants ---
        failed_frame = ttk.Frame(failed_box)
        failed_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
    
        cols_fail = ("N° Série", "Emplacement", "Cause")
        self.fail_tree = ttk.Treeview(failed_frame, columns=cols_fail, show="headings", selectmode="browse")
        for c in cols_fail:
            self.fail_tree.heading(c, text=c)
            width = 150 if c != "Cause" else 260
            self.fail_tree.column(c, anchor="w", width=width, stretch=True)
    
        yscroll_fail = ttk.Scrollbar(failed_frame, orient="vertical", command=self.fail_tree.yview)
        self.fail_tree.configure(yscrollcommand=yscroll_fail.set)
    
        self.fail_tree.pack(side="left", fill="both", expand=True)
        yscroll_fail.pack(side="right", fill="y")
    
        # --- Partie droite : Batteries non prêtes ---
        notready_frame = ttk.Frame(failed_box)
        notready_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
    
        cols_notready = ("N° Série",)
        self.notready_tree = ttk.Treeview(
            notready_frame, columns=cols_notready, show="headings", selectmode="browse"
        )
        self.notready_tree.heading("N° Série", text="N° Série")
        self.notready_tree.column("N° Série", anchor="w", width=180, stretch=True)
    
        yscroll_notready = ttk.Scrollbar(notready_frame, orient="vertical", command=self.notready_tree.yview)
        self.notready_tree.configure(yscrollcommand=yscroll_notready.set)
    
        self.notready_tree.pack(side="left", fill="both", expand=True)
        yscroll_notready.pack(side="right", fill="y")
    
        # ====== Gros bouton ovale centré (canvas) ======
        btn_frame = ttk.Frame(wrap)
        btn_frame.pack(fill="x", pady=8)
    
        canvas = tk.Canvas(btn_frame, width=260, height=64, highlightthickness=0, bg=self.cget("background"))
        canvas.pack(pady=6)
        oval = canvas.create_oval(4, 4, 256, 60, fill="#2E62FF", outline="#1b3fb3", width=2)
        txt  = canvas.create_text(130, 32, text="Traiter les fichiers", fill="white", font=("Segoe UI", 11, "bold"))
    
        canvas.tag_bind(oval, "<Button-1>", self._on_click)
        canvas.tag_bind(txt,  "<Button-1>", self._on_click)
        
        self.afficher_numero_en_attente()
        
    def afficher_numero_en_attente(self):
        conn = self.db_manager.connect()

        try:
            modele=self.selected_model
            stage_act='capa'
            cursor = conn.cursor()
            query=self.build_stage_query(stage_act)
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  

            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
            
        except Exception as e:
            # Erreur SQL => échec de traitement
            messagebox.shoxinfo('error!', f"recuperation de la liste des batteries a tester impossible : {e}")
            
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            self.ok_series_tree.delete(*self.ok_series_tree.get_children())
            for batt in liste_batteries:
                self.ok_series_tree.insert("", "end", values=(batt,))
            
    def _on_click(self,event=None):
        
        ok_pairs = []   # [(numero_serie_batterie, chemin_fichier), ...]
        ko_files = []   # [chemin_fichier, ...]
        non_test_files=[]
        
        conn = self.db_manager.connect()

        try:
            modele=self.selected_model
            stage_act='capa'
            cursor = conn.cursor()
            query=self.build_stage_query(stage_act)
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  

            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
            
        except Exception as e:
            # Erreur SQL => échec de traitement
            messagebox.shoxinfo('error!', f"recuperation de la liste des batteries a tester impossible : {e}")
            
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
        
        
        dossier_path=r"G:\Drive partagés\VoltR\4_Production\5_Cyclage\1_Résultats de cyclage\Fichier en cours"
        dossier_exploites=r"G:\Drive partagés\VoltR\4_Production\5_Cyclage\1_Résultats de cyclage\Fichiers traités\Batteries"
        dossier_ko=r"G:\Drive partagés\VoltR\4_Production\5_Cyclage\1_Résultats de cyclage\Fichiers NOK\Batteries KO"
        """
        dossier_path="C:/Users/User/Desktop/MAJ_TEST/Resultats"
        dossier_exploites="C:/Users/User/Desktop/MAJ_TEST/Fichiers_exploites"
        dossier_ko="C:/Users/User/Desktop/MAJ_TEST/Fichiers_non_ok/cellues K.O"
        """
        """
        dossier_path=r"G:\Drive partagés\VoltR\4_Production\5_Cyclage\1_Résultats de cyclage\Fichier en cours"
        dossier_exploites= "G:/Drive partagés/VoltR/4_Production/5_Cyclage/1_Résultats de cyclage/Fichiers traités/batterie"
        dossier_ko= "G:/Drive partagés/VoltR/4_Production/5_Cyclage/1_Résultats de cyclage/Fichiers NOK/Batteries KO"
        """
        df_cyclage = pd.read_excel(EXCEL_PATH,sheet_name="Cyclage",header=1)
        
        for fichier in os.listdir(dossier_path):#Traite chaque fichier dans le dossier 
            if fichier.endswith((".xlsx", ".xls")):  # Vérifie si le fichier est de type Excel
                chemin_fichier = os.path.join(dossier_path, fichier)
                numero_serie_test = os.path.splitext(os.path.basename(chemin_fichier))[0]
                # Ex: "MC0002031068-4-3-7-INR18650MH1_A.0.1"
                parties = numero_serie_test.split('-', 4)
            
                # Sécurités basiques sur le nom
                if len(parties) < 5:
                    # Nom inattendu -> on ignore/passe
                    continue
            
                numero_serie_batterie = parties[0]
                emplacement = f"{parties[1]}-{parties[2]}-{parties[3]}"
                modele_test = parties[4]
                
                if numero_serie_batterie not in liste_batteries:
                    
                    self.notready_tree.insert("", "end", values=(numero_serie_batterie))
                    continue
                    
            
                if modele_test == self.selected_model:
                    try:
                        
                        # --- DB: récupérer ref_cell ---
                        conn = self.db_manager.connect()
                        if not conn:
                            # Pas de connexion = échec de traitement
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "conn BDD"))
                            
                            continue
            
                        try:
                            cursor = conn.cursor()
                            query = """
                                SELECT reference_cellule
                                FROM cellule
                                WHERE affectation_produit = %s
                                LIMIT 1
                            """
                            cursor.execute(query, (numero_serie_batterie,))
                            row_db = cursor.fetchone()
                        except Exception as e:
                            # Erreur SQL => échec de traitement
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            print(f"[SQL] {numero_serie_batterie} -> {e}")
                           
                            row_db = None
                        finally:
                            try:
                                cursor.close()
                            except:
                                pass
                            conn.close()
            
                        if not row_db or not row_db[0]:
                            # Pas de ref cellule => impossible de lire les seuils
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            
                            continue
            
                        ref_cell = row_db[0]
            
                        # --- Seuils (df_cyclage) ---
                        row = df_cyclage[
                            (df_cyclage["Nom_modele"] == modele_test) &
                            (df_cyclage["Ref cellule"] == ref_cell)
                        ]
            
                        if row.empty:
                            # Seuil introuvable
                            self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                            
                            continue
            
                        seuils = row.iloc[0]
                        try:
                            capa_min = float(seuils["Capa mini (Ah)"])
                        except Exception:
                            capa_min = float(str(seuils["Capa mini (Ah)"]).replace(",", "."))
                        try:
                            temps_h = float(seuils["Temps de test (h)"])
                        except Exception:
                            temps_h = float(str(seuils["Temps de test (h)"]).replace(",", "."))
                        try:
                            tension_seuil = float(seuils["Tension seuil (V)"])
                        except Exception:
                            tension_seuil = float(str(seuils["Tension seuil (V)"]).replace(",", "."))
            
                        # --- Lecture mesures ---
                        
                        if modele_test[:7]=="EMBR036":
                            
                            # Capacité (onglet 'step', ligne 4, col 'Capacity(Ah)')
                            step = pd.read_excel(chemin_fichier, sheet_name="step")
                            try:
                                capa_dch = float(step["Capacity(Ah)"].iloc[4])
                            except Exception as e:
                                # Si la ligne/col manque -> échec de traitement
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                print(f"[STEP capa] {numero_serie_batterie} -> {e}")
                                
                                continue
                
                            # Tension fin test (onglet 'record', Step Index == 4, max sur les 10 derniers points)
                            record = pd.read_excel(chemin_fichier, sheet_name="record")
                            df_dch_last = record[record["Step Index"] == 6].head(10)
                            if df_dch_last.empty or "Voltage(V)" not in df_dch_last.columns:
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                continue
                            try:
                                max_voltage = float(df_dch_last["Voltage(V)"].max())
                            except Exception as e:
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                print(f"[RECORD volt] {numero_serie_batterie} -> {e}")
                                
                                continue
                
                            # Temps test (dernière valeur 'Total Time')
                            if "Total Time" not in record.columns or record["Total Time"].empty:
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                continue
                
                            last_time_val = record["Total Time"].iloc[-1]
                            tension_finale = record["Voltage(V)"].iloc[-1]
                            # Robustifier la conversion en timedelta
                            try:
                                # Si déjà de type timedelta/chaîne "HH:MM:SS"
                                time_obj = pd.to_timedelta(last_time_val)
                                if pd.isna(time_obj):
                                    raise ValueError("NaT")
                            except Exception:
                                # Si Excel time (float en jours)
                                try:
                                    time_obj = pd.to_timedelta(float(last_time_val), unit="D")
                                except Exception as e:
                                    self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                    print(f"[RECORD time] {numero_serie_batterie} -> {e}")
                                    
                                    continue
                            
                        elif modele_test[:8]=="PPTR018A" or modele_test[:8]=="LNBR008A":
                            
                            # Capacité (onglet 'step', ligne 4, col 'Capacity(Ah)')
                            step = pd.read_excel(chemin_fichier, sheet_name="step")
                            try:
                                capa_dch = float(step["Capacity(Ah)"].iloc[3])
                            except Exception as e:
                                # Si la ligne/col manque -> échec de traitement
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                print(f"[STEP capa] {numero_serie_batterie} -> {e}")
                                
                                continue
                
                            # Tension fin test (onglet 'record', Step Index == 4, max sur les 10 derniers points)
                            record = pd.read_excel(chemin_fichier, sheet_name="record")
                            df_dch_last = record[record["Step Index"] == 5].head(10)
                            if df_dch_last.empty or "Voltage(V)" not in df_dch_last.columns:
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                continue
                            try:
                                max_voltage = float(df_dch_last["Voltage(V)"].max())
                            except Exception as e:
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                print(f"[RECORD volt] {numero_serie_batterie} -> {e}")
                                
                                continue
                
                            # Temps test (dernière valeur 'Total Time')
                            if "Total Time" not in record.columns or record["Total Time"].empty:
                                self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                continue
                
                            last_time_val = record["Total Time"].iloc[-1]
                            tension_finale = record["Voltage(V)"].iloc[-1]
                            # Robustifier la conversion en timedelta
                            try:
                                # Si déjà de type timedelta/chaîne "HH:MM:SS"
                                time_obj = pd.to_timedelta(last_time_val)
                                if pd.isna(time_obj):
                                    raise ValueError("NaT")
                            except Exception:
                                # Si Excel time (float en jours)
                                try:
                                    time_obj = pd.to_timedelta(float(last_time_val), unit="D")
                                except Exception as e:
                                    self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                                    print(f"[RECORD time] {numero_serie_batterie} -> {e}")
                                    
                                    continue
                     
                        # --- Indicateurs ---
                        indic_capa = "OK" if capa_dch > capa_min else "NOK"
                        indic_v    = "OK" if max_voltage < tension_seuil else "NOK"
                        indic_t    = "OK" if time_obj > pd.to_timedelta(f"{int(temps_h)} hours") else "NOK"
                        
                        if indic_capa == "OK" and indic_v == "OK" and indic_t == "OK":
                            # Tous OK -> tree principal (5 colonnes)
                            self.test_tree.insert(
                                "", "end",
                                values=(
                                    numero_serie_batterie,
                                    modele_test,
                                    f"{capa_dch:.3f}",
                                    f"{tension_finale:.3f}",
                                    emplacement
                                )
                            )
                            ok_pairs.append((numero_serie_batterie, chemin_fichier))
                            conn = self.db_manager.connect()
                            if not conn:
                                messagebox.showerror("Erreur DB", "Connexion DB impossible pour l'update.")
                                return
                        
                            try:
                                cur = conn.cursor()
                        
                                sql = """
                                    UPDATE suivi_production
                                    SET valeur_test_capa = %s
                                    WHERE numero_serie_batterie = %s
                                """
                        
                                params = (capa_dch,numero_serie_batterie)
                                cur.execute(sql, params)
                            except Exception as e:
                                messagebox.showerror("Erreur DB", f"Update échoué : {e}")
                            finally:
                                try:
                                    cur.close()
                                except:
                                    pass
                                conn.commit()
                                conn.close()
                        else:
                            # Au moins un NOK -> tree défaillants (3 colonnes) avec cause
                            causes = []
                            if indic_capa == "NOK":
                                causes.append("Capacité")
                            if indic_v == "NOK":
                                causes.append("Tension")
                            if indic_t == "NOK":
                                causes.append("Temps")
                            cause_txt = " & ".join(causes) if causes else "traitement"
                            self.fail_tree.insert(
                                "", "end",
                                values=(numero_serie_batterie, emplacement, cause_txt)
                            )
                            ko_files.append(chemin_fichier)
            
                    except Exception as e:
                        # Toute autre erreur après le if modele_test == self.selected_model
                        self.fail_tree.insert("", "end", values=(numero_serie_batterie, emplacement, "traitement"))
                        ko_files.append(chemin_fichier)
                        print(f"[TRAITEMENT] {numero_serie_batterie} -> {e}")
                else:
                    # Modèle différent => on ignore ce fichier
                    pass                       
            
        # === Après la boucle: mise à jour BDD pour les OK ===
        ok_serials = list({s for (s, _p) in ok_pairs})  # dédoublonné
        if ok_serials:
            self._update_ok_in_db(ok_serials)
    
        # === Déplacement des fichiers ===
        self._move_processed_files(ok_pairs, ko_files, dossier_exploites, dossier_ko)    
        
    def _update_ok_in_db(self, ok_serials):
        """
        Met à jour la base pour tous les numéros série OK.
        Adapte la requête SQL selon ton schéma.
        """
        conn = self.db_manager.connect()
        if not conn:
            messagebox.showerror("Erreur DB", "Connexion DB impossible pour l'update.")
            return
    
        try:
            cur = conn.cursor()
    
            sql = """
                UPDATE suivi_production
                   SET test_capa = 1,
                       date_test_capa = NOW()
                 WHERE numero_serie_batterie = %s
            """
    
            params = [(s,) for s in ok_serials]
            cur.executemany(sql, params)
            
            for s in ok_serials:
                cur.execute("Select fermeture_batt from suivi_production where numero_serie_batterie= %s",(s,))
                ress=cur.fetchall()
                etat_f=[res[0] for res in ress]
                if etat_f:
                    cur.execute("UPDATE produit_voltr SET statut = %s where numero_serie_produit= %s",('stock',s))
            conn.commit()
            print(f"Update OK sur {cur.rowcount} lignes.")
        except Exception as e:
            messagebox.showerror("Erreur DB", f"Update échoué : {e}")
        finally:
            try:
                cur.close()
            except:
                pass
            conn.close()
            self.afficher_numero_en_attente()
    
    def _move_processed_files(self, ok_pairs, ko_files, dossier_exploites, dossier_ko):
        """
        Déplace les fichiers:
          - OK -> dossier_exploites
          - KO/erreurs -> dossier_ko
        """
        os.makedirs(dossier_exploites, exist_ok=True)
        os.makedirs(dossier_ko, exist_ok=True)
    
        def _safe_move(src_path, dst_dir):
            try:
                base = os.path.basename(src_path)
                dst = os.path.join(dst_dir, base)
                if os.path.exists(dst):
                    name, ext = os.path.splitext(base)
                    i = 1
                    candidate = os.path.join(dst_dir, f"{name} ({i}){ext}")
                    while os.path.exists(candidate):
                        i += 1
                        candidate = os.path.join(dst_dir, f"{name} ({i}){ext}")
                    dst = candidate
                shutil.move(src_path, dst)
                return True
            except Exception as e:
                print(f"[MOVE] {src_path} -> {dst_dir} : {e}")
                return False
    
        # OK -> exploités
        for _, src in ok_pairs:
            _safe_move(src, dossier_exploites)
    
        # KO -> ko
        for src in ko_files:
            _safe_move(src, dossier_ko)
    
        
    #------------------------------ Onglet recherche -----------------------------------------------   
        
    def setup_recherche(self, frame):
        # ----- Layout principal : gauche | boutons | droite -----
        container = ttk.Frame(frame); container.pack(fill="both", expand=True, padx=12, pady=12)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=0)
        container.columnconfigure(2, weight=2)
        container.rowconfigure(0, weight=1)
    
        # ------- Colonne gauche : entrées + combobox + listbox -------
        left = ttk.LabelFrame(container, text="Recherche", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
       
    
        ttk.Label(left, text="N° série cellule").grid(row=0, column=0, sticky="w")
        self.rech_entry_cell = ttk.Entry(left, width=28)
        self.rech_entry_cell.grid(row=1, column=0, sticky="we", pady=(0,8))
        # Remplissage auto du n° batterie quand l'entry cellule atteint 12 chars
        self.rech_entry_cell.bind("<KeyRelease>", self._rech_on_cell_entry)
    
        ttk.Label(left, text="N° série batterie").grid(row=2, column=0, sticky="w")
        self.rech_entry_batt = ttk.Entry(left, width=28)
        self.rech_entry_batt.grid(row=3, column=0, sticky="we", pady=(0,8))
    
        ttk.Label(left, text="Référence batterie").grid(row=4, column=0, sticky="w")
        # valeur par défaut nulle (vide)
        self.rech_model_var = tk.StringVar(value="")
        self.rech_combo = ttk.Combobox(left, textvariable=self.rech_model_var,
                                       values=(self.models or []), state="readonly", width=30)
        self.rech_combo.grid(row=5, column=0, sticky="we", pady=(0,8))
        self.rech_combo.bind("<<ComboboxSelected>>", lambda e: self._rech_on_model_change())
    
        ttk.Label(left, text="Liste batterie").grid(row=6, column=0, sticky="w")
        lb_frame = ttk.Frame(left); lb_frame.grid(row=7, column=0, sticky="nsew")
        left.rowconfigure(7, weight=1)
    
        # multi-sélection
        self.rech_listbox = tk.Listbox(lb_frame, height=10, activestyle="dotbox", selectmode="extended")
        yscroll = ttk.Scrollbar(lb_frame, orient="vertical", command=self.rech_listbox.yview)
        self.rech_listbox.configure(yscrollcommand=yscroll.set)
        self.rech_listbox.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
    
        # Double-clic => déplacer à droite
        self.rech_listbox.bind("<Double-1>", lambda e: self._rech_move_right())
    
        # --------- Colonne boutons centraux ----------
        mid = ttk.Frame(container); mid.grid(row=0, column=1, sticky="ns")
        for i in range(3): mid.rowconfigure(i, weight=1)
        ttk.Button(mid, text="→", width=3, command=self._rech_move_right).grid(row=0, column=0, pady=4)
        ttk.Button(mid, text="←", width=3, command=self._rech_remove_selected_right).grid(row=1, column=0, pady=4)
    
        # --------- Colonne droite : table dynamique ----------
        right = ttk.LabelFrame(container, text="Sélection / Détails", padding=10)
        right.grid(row=0, column=2, sticky="nsew", padx=(8,0))
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        
        # largeur fixe (par ex. 500 px, ajuste comme tu veux)
        right.configure(width=800)
        right.grid_propagate(False)   # bloque l’expansion auto
        
        self.rech_right_title = ttk.Label(right, text="", font=("", 10, "bold"))
        self.rech_right_title.grid(row=0, column=0, sticky="w", pady=(0,6))
    
        tv_frame = ttk.Frame(right); tv_frame.grid(row=1, column=0, sticky="nsew")
        # + scroll vertical & horizontal
        self.rech_tree = ttk.Treeview(tv_frame, columns=(), show="headings", selectmode="extended")
        y2 = ttk.Scrollbar(tv_frame, orient="vertical", command=self.rech_tree.yview)
        x2 = ttk.Scrollbar(tv_frame, orient="horizontal", command=self.rech_tree.xview)
        self.rech_tree.configure(yscrollcommand=y2.set, xscrollcommand=x2.set)
        self.rech_tree.grid(row=0, column=0, sticky="nsew")
        y2.grid(row=0, column=1, sticky="ns")
        x2.grid(row=1, column=0, sticky="ew")
        tv_frame.rowconfigure(0, weight=1); tv_frame.columnconfigure(0, weight=1)
    
        # set utilisé pour éviter les doublons à droite (clé = numero_serie_batterie)
        self._rech_right_keys = set()
        
    def _rech_on_cell_entry(self, event=None):
        """Quand l'entry cellule atteint 12 chars, on cherche la batterie associée et on remplit l'entry batterie."""
        numero_serie_cell = self.rech_entry_cell.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # affectation_produit = numero_serie_batterie (selon ta logique existante)
            cur.execute("SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s", (numero_serie_cell,))
            row = cur.fetchone()
            self.rech_entry_batt.delete(0, tk.END)
            if row and row[0]:
                self.rech_entry_batt.insert(0, str(row[0]))
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Lookup cellule→batterie impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_on_model_change(self):
        """Quand on choisit un modèle, on alimente la liste des n° batteries via la jointure demandée."""
        ref = self.rech_model_var.get().strip()
        self.rech_listbox.delete(0, tk.END)
        if not ref:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # Liste des NUMÉROS DE SÉRIE BATTERIE pour la référence choisie
            # sp = suivi_production / p = produit
            sql = ("""
                SELECT DISTINCT sp.numero_serie_batterie
                FROM suivi_production sp
                JOIN produit_voltr p
                  ON sp.numero_serie_batterie = p.numero_serie_produit
                WHERE p.reference_produit_voltr = %s
                ORDER BY sp.numero_serie_batterie
            """)
            cur.execute(sql, (ref,))
            for (num_batt,) in cur.fetchall():
                if num_batt:
                    self.rech_listbox.insert(tk.END, str(num_batt))
            self.rech_right_title.config(text=f"Modèle sélectionné : {ref}")
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement liste batteries impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_move_right(self):
        """Ajoute à droite : 1) le n° saisi dans l’entry batterie (si présent),
        2) tous les n° sélectionnés dans la liste. Charge les LIGNES de suivi_production correspondantes."""
        # rassembler les cibles
        targets = set()
        batt_from_entry = self.rech_entry_batt.get().strip()
        if batt_from_entry:
            targets.add(batt_from_entry)
        for idx in self.rech_listbox.curselection():
            targets.add(self.rech_listbox.get(idx))
    
        if not targets:
            return
    
        # Charger les lignes depuis suivi_production
        rows, colnames = self._rech_load_suivi_rows(list(targets))
        if not rows:
            return
    
        # Configurer le tree si c'est la 1ère fois ou si colonnes différentes
        self._rech_configure_tree_for_columns(colnames)
    
        # Insérer sans doublon (clé = numero_serie_batterie)
        try:
            k_idx = colnames.index("numero_serie_batterie")  # l’utilisateur veut dédupliquer sur ce champ
        except ValueError:
            # si absent (peu probable), on déduplique sur la 1ère colonne
            k_idx = 0
    
        added = 0
        for r in rows:
            key = str(r[k_idx]) if r[k_idx] is not None else ""
            if key and key not in self._rech_right_keys:
                self.rech_tree.insert("", tk.END, values=[("" if v is None else v) for v in r])
                self._rech_right_keys.add(key)
                added += 1
    
        if added == 0:
            self.rech_right_title.config(text="Aucun nouvel élément (déduplication active)")
    
    def _rech_remove_selected_right(self):
        """Supprimer les lignes sélectionnées dans la table de droite."""
        sel = self.rech_tree.selection()
        if not sel:
            return
        # identifier l'index de numero_serie_batterie pour nettoyer le set
        cols = self.rech_tree.cget("columns")
        try:
            k_idx = cols.index("numero_serie_batterie")
        except ValueError:
            k_idx = 0
        for iid in sel:
            vals = self.rech_tree.item(iid, "values")
            # protéger si vide
            if vals:
                key = str(vals[k_idx])
                if key in self._rech_right_keys:
                    self._rech_right_keys.remove(key)
            self.rech_tree.delete(iid)
    
    def _rech_load_suivi_rows(self, numero_serie_batteries):
        """Retourne (rows, colnames) pour les lignes de suivi_production filtrées par n° série batterie."""
        if not numero_serie_batteries:
            return [], []
        conn = self.db_manager.connect()
        if not conn:
            return [], []
        try:
            cur = conn.cursor()
            # On charge TOUTE la ligne de suivi_production (c’est ce que tu as demandé)
            # IN sécurisé : on fabrique la liste de %s
            placeholders = ", ".join(["%s"] * len(numero_serie_batteries))
            sql = f"SELECT * FROM suivi_production WHERE numero_serie_batterie IN ({placeholders})"
            cur.execute(sql, tuple(numero_serie_batteries))
            rows = cur.fetchall()
            colnames = [d[0] for d in cur.description]
            return rows, colnames
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement suivi_production impossible :\n{e}")
            return [], []
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def _rech_configure_tree_for_columns(self, colnames):
        """Configure le Treeview (colonnes, largeur, ancrage, headings) + stretch + scroll horiz/vert déjà mis côté UI."""
        # si colonnes déjà identiques, ne rien faire
        current = self.rech_tree.cget("columns")
        if tuple(current) == tuple(colnames):
            return
    
        # reset
        for c in current:
            try: self.rech_tree.heading(c, text="")
            except: pass
        self.rech_tree.delete(*self.rech_tree.get_children())
        self._rech_right_keys.clear()
    
        self.rech_tree["columns"] = colnames
        # headings & options
        for c in colnames:
            self.rech_tree.heading(c, text=c)
            # largeur heuristique : min 100, sinon ~len*9 px
            width = max(100, min(380, int(len(c) * 9)))
            self.rech_tree.column(c, width=width, minwidth=80, stretch=True, anchor="w")

    
    #------------------------------ Onglet expedition -----------------------------------------------   

    def setup_exp(self, frame):
        # ---------- Layout principal : gauche | droite ----------
        container = ttk.Frame(frame); container.pack(fill="both", expand=True, padx=12, pady=12)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(0, weight=1)
    
        # ========== Colonne gauche : entrées & combos ==========
        left = ttk.LabelFrame(container, text="Expédition", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
    
        # N° série produit + ajout auto si 9 chiffres
        ttk.Label(left, text="N° série produit:").pack(pady=(0,4), anchor="w")
        self.exp_numero_serie_batt_entry = ttk.Entry(left, width=30)
        self.exp_numero_serie_batt_entry.pack(pady=(0,8), fill="x")
        self.exp_numero_serie_batt_entry.bind("<KeyRelease>", self._exp_on_entry_change)
    
        ttk.Button(left, text="❌ Non conforme", command=self.add_non_conf_batterie, style="Danger.TButton").pack(pady=6, anchor="w")
    
        # Client
        ttk.Label(left, text="Client").pack(pady=(8,4), anchor="w")
        cb_var_client = tk.StringVar()
        self.cb_cl = ttk.Combobox(left, textvariable=cb_var_client, state="readonly", width=40)
        self.cb_cl.pack(fill="x")
    
        # Alimentation des clients
        conn = self.db_manager.connect()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT nom_client FROM client ORDER BY nom_client")
                mots = [row[0].replace(" ", "-") for row in cursor.fetchall()]
                self.cb_cl["values"] = mots
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Chargement clients :\n{e}")
            finally:
                try: cursor.close()
                except: pass
                conn.close()
    
        # Projet
        ttk.Label(left, text="Projet").pack(pady=(8,4), anchor="w")
        self.cb_pr = ttk.Combobox(left, state="readonly", width=40)
        self.cb_pr.pack(fill="x")
        
        # Alimentation des clients
        conn = self.db_manager.connect()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT nom_projet FROM projet ORDER BY nom_projet")
                mots = [row[0] for row in cursor.fetchall()]
                self.cb_pr["values"] = mots
            except Exception as e:
                messagebox.showerror("Erreur SQL", f"Chargement projet :\n{e}")
            finally:
                try: cursor.close()
                except: pass
                conn.close()
    
        # Modèle + checkbox d’activation
        ttk.Label(left, text="Modèle batterie").pack(pady=(8,4), anchor="w")
        cb_var_m_exp = tk.StringVar()
        self.cb_exp = ttk.Combobox(left, textvariable=cb_var_m_exp, values=getattr(self, "models", []),
                                   state="disabled", width=40)
        self.cb_exp.pack(fill="x")
    
        self.chk_exp_var = tk.BooleanVar(value=False)
        def toggle_combobox_exp():
            self.cb_exp.configure(state="readonly" if self.chk_exp_var.get() else "disabled")
            
        ttk.Checkbutton(left, text="Changer le modèle", variable=self.chk_exp_var,
                        command=toggle_combobox_exp).pack(pady=(6,0), anchor="w")
    
        # Commentaire
        ttk.Label(left, text="Commentaire").pack(pady=(8,4), anchor="w")
        self.exp_comm_entry = ttk.Entry(left)
        self.exp_comm_entry.pack(fill="x")
    
        # ========= Colonne droite : 2 Listbox (disponibles | sélectionnées) =========
        right = ttk.Frame(container)
        right.grid(row=0, column=1, sticky="nsew", padx=(8,0))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)
        right.rowconfigure(4, weight=1)
    
        # Liste des batteries disponibles (emballées)
        ttk.Label(right, text="Batteries emballées (disponibles)").grid(row=0, column=0, sticky="w", pady=(0,4))
        self.exp_listbox_batt = tk.Listbox(right, font=('Segoe UI', 11), height=10)
        self.exp_listbox_batt.grid(row=1, column=0, sticky="nsew")
        self.exp_listbox_batt.bind("<<ListboxSelect>>", lambda e: None)  # neutre
        self.exp_listbox_batt.bind("<Double-Button-1>", self._exp_on_available_double_click)
    
        ttk.Separator(right, orient="horizontal").grid(row=2, column=0, sticky="ew", pady=10)
    
        # Sélection (avec compteur)
        header_sel = ttk.Frame(right); header_sel.grid(row=3, column=0, sticky="ew", pady=(0,4))
        ttk.Label(header_sel, text="Batteries sélectionnées").pack(side="left")
        ttk.Label(header_sel, text="Quantité:").pack(side="right")
        self._exp_count_var = tk.IntVar(value=0)
        self._exp_count_lbl = ttk.Label(header_sel, textvariable=self._exp_count_var)
        self._exp_count_lbl.pack(side="right", padx=(0,8))
    
        self.send_listbox_batt = tk.Listbox(right, font=('Segoe UI', 11), height=10)
        self.send_listbox_batt.grid(row=4, column=0, sticky="nsew")
        self.send_listbox_batt.bind("<Double-Button-1>", self._exp_on_selected_double_click)
    
        ttk.Button(right, text="✅ Contrôle OK", command=self.valider_exp, style="Good.TButton").grid(
            row=5, column=0, pady=10, sticky="e"
        )
    
        # Charge la liste des disponibles
        self.display_model_list_exp()
    
    
    def display_model_list_exp(self):
        """Alimente la listbox des batteries emballées (disponibles)."""
        conn = self.db_manager.connect()
        stage_act='exp'
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query=self.build_stage_query(stage_act)
            cursor.execute(query)
            rows = cursor.fetchall()
            liste_batteries = [str(r[0]) for r in rows]
            self.exp_listbox_batt.delete(0, tk.END)
            for batt in liste_batteries:
                self.exp_listbox_batt.insert(tk.END, batt)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try: cursor.close()
            except: pass
            conn.close()
        
    #------------------------------ Onglet recyclage -----------------------------------------------   
    def setup_recyclage(self,frame):

        left = ttk.Frame(frame)
        left.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        
        right = ttk.Frame(frame)
        right.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        
        ttk.Label(left, text="N° série cellule").pack(pady=5)
        self.r_entry_cell = ttk.Entry(left, width=28)
        self.r_entry_cell.pack(pady=5)
        # Remplissage auto du n° batterie quand l'entry cellule atteint 12 chars
        self.r_entry_cell.bind("<KeyRelease>", self.r_on_cell_entry)
    
        ttk.Label(left, text="N° série batterie").pack(pady=5)
        self.r_entry_batt = ttk.Entry(left, width=28)
        self.r_entry_batt.pack(pady=5)
    
        ttk.Label(left, text="Référence batterie").pack(pady=5)
        # valeur par défaut nulle (vide)
        self.r_model_var = tk.StringVar(value="")
        self.r_combo = ttk.Combobox(left, textvariable=self.r_model_var,
                                       values=(self.models or []), state="readonly", width=30)
        self.r_combo.pack(pady=5)
        self.r_combo.bind("<<ComboboxSelected>>", lambda e: self.r_on_model_change())
        
        ttk.Label(left, text="Cause recyclage:").pack(pady=5)
        
        self.cause_var = tk.StringVar(value="")
        self.cause_combo = ttk.Combobox(left, textvariable=self.cause_var,
                                       values=("erreur_soudure","choc","autres"), state="readonly", width=30)
        self.cause_combo.pack(pady=5)
        
        ttk.Button(
            left, text="🔄 Recyler la batterie",
            command=self.recycle_batterie_sp, style="Danger.TButton"
        ).pack(pady=10)
    
        ttk.Label(right, text="Liste batterie").pack(pady=5)
        
        self.r_listbox = tk.Listbox(right, height=10, activestyle="dotbox", selectmode="extended")
        yscroll = ttk.Scrollbar(right, orient="vertical", command=self.r_listbox.yview)
        self.r_listbox.configure(yscrollcommand=yscroll.set)
        self.r_listbox.pack(side="right", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        self.r_listbox.bind("<<ListboxSelect>>", self.on_r_listbox_select)
        
    def on_r_listbox_select(self, event):
        # obtenir indices sélectionnés (peut être plusieurs si selectmode="extended")
        sel = self.r_listbox.curselection()
        if not sel:
            return
        # on prend le premier sélectionné
        idx = sel[0]
        text = self.r_listbox.get(idx)
    
        # --- si text est déjà le numero de série simple ---
        # numero = text.strip()
    
        # --- OU : si text contient d'autres champs, extraire le numéro ---
        # Exemples d'extraction (choisis celle qui correspond à ton format)
        # 1) format "123456789012" => direct
        # 2) format "1;123456789012;moduleA" => split par ';' et prendre le 2ème
        # 3) format "1 | 123456789012 | module A" => split par '|' et strip
        numero = None
        if ";" in text:
            parts = [p.strip() for p in text.split(";")]
            # si le numéro est en 2e position
            if len(parts) >= 2:
                numero = parts[1]
        elif "|" in text:
            parts = [p.strip() for p in text.split("|")]
            # chercher la première partie qui ressemble à un n° (ex: longueur 12, chiffres)
            for p in parts:
                if p and any(ch.isdigit() for ch in p):
                    numero = p
                    break
        else:
            # par défaut on prend toute la chaîne
            numero = text.strip()
    
        # si tu utilises une StringVar pour l'entry, mets-la ; sinon delete/insert
        try:
            self.r_entry_batt.delete(0, "end")
            if numero:
                self.r_entry_batt.insert(0, numero)
        except Exception as e:
            # fallback si self.r_entry_batt est une StringVar
            if hasattr(self, "r_model_var") and isinstance(self.r_model_var, tk.StringVar):
                self.r_model_var.set(numero or "")
            else:
                raise
    
    def r_on_cell_entry(self, event=None):
        """Quand l'entry cellule atteint 12 chars, on cherche la batterie associée et on remplit l'entry batterie."""
        numero_serie_cell = self.r_entry_cell.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # affectation_produit = numero_serie_batterie (selon ta logique existante)
            cur.execute("SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s", (numero_serie_cell,))
            row = cur.fetchone()
            self.r_entry_batt.delete(0, tk.END)
            if row and row[0]:
                self.rech_entry_batt.insert(0, str(row[0]))
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Lookup cellule→batterie impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
        
    def r_on_model_change(self):
        """Quand on choisit un modèle, on alimente la liste des n° batteries via la jointure demandée."""
        ref = self.r_model_var.get().strip()
        self.r_listbox.delete(0, tk.END)
        if not ref:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
            # Liste des NUMÉROS DE SÉRIE BATTERIE pour la référence choisie
            # sp = suivi_production / p = produit
            sql = ("""
                SELECT DISTINCT sp.numero_serie_batterie
                FROM suivi_production sp
                JOIN produit_voltr p
                  ON sp.numero_serie_batterie = p.numero_serie_produit
                WHERE p.reference_produit_voltr = %s
                AND recyclage is null 
                ORDER BY sp.numero_serie_batterie
            """)
            cur.execute(sql, (ref,))
            for (num_batt,) in cur.fetchall():
                if num_batt:
                    self.r_listbox.insert(tk.END, str(num_batt))
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Chargement liste batteries impossible :\n{e}")
        finally:
            try: cur.close()
            except: pass
            conn.close()
    
    def recycle_batterie_sp(self):
        type_obj = "batterie"
        cause = self.cause_combo.get()
        if not cause:
            messagebox.showerror("Pas de cause", "Veuillez sélectionner une cause !")
            return
    
        numero_serie_batt = self.r_entry_batt.get().strip()
        if not numero_serie_batt:
            messagebox.showerror("Erreur", "Veuillez renseigner le numéro de série de la batterie.")
            return
    
        conn = None
        cursor = None
        try:
            conn = self.db_manager.connect()
            cursor = conn.cursor()
    
            # 1) Récupérer référence et poids de la batterie
            query = "SELECT pv.reference_produit_voltr, rv.poids FROM produit_voltr as pv join ref_batterie_voltr as rv on pv.reference_produit_voltr=rv.reference_batterie_voltr WHERE numero_serie_produit = %s"
            cursor.execute(query, (numero_serie_batt,))
            row_prod = cursor.fetchone()
            if row_prod is None:
                messagebox.showerror("Non trouvé", f"Aucune batterie trouvée pour le n° {numero_serie_batt}")
                return
    
            reference_batt = row_prod[0]
            poids_batt = row_prod[1] or 0
    
            # 2) Lire la feuille Excel et trouver le dest_recyclage
            df_cyclage = pd.read_excel(EXCEL_PATH, sheet_name="Cyclage", header=1)
            sel = df_cyclage[df_cyclage["Nom_modele"] == reference_batt]
            if sel.empty:
                messagebox.showerror("Erreur Excel", f"Modèle {reference_batt} introuvable dans {EXCEL_PATH} sheet Cyclage.")
                return
            seuils = sel.iloc[0]
            dest_recyclage = str(seuils.get("Recyclage", "")).strip()
            if not dest_recyclage:
                messagebox.showerror("Erreur", f"Pas de destination de recyclage définie pour {reference_batt}.")
                return
            type_fut = dest_recyclage
    
            # 3) Chercher un fut ouvert pour ce type
            cursor.execute(
                "SELECT id_fut, poids FROM fut_recyclage WHERE exutoire = %s AND etat_fut = %s LIMIT 1",
                (type_fut, "en cours")
            )
            fut_row = cursor.fetchone()
            if fut_row is None:
                messagebox.showerror("Erreur !", "Aucun fut d'exutoire eo_org_mtl n'est ouvert")
                return
    
            id_fut, poids_fut = fut_row[0], fut_row[1] or 0
    
            # 4) Mettre à jour le poids du fut
            poids_tot = poids_fut + poids_batt
            cursor.execute("UPDATE fut_recyclage SET poids = %s WHERE id_fut = %s", (poids_tot, id_fut))
    
            # 5) Insérer la ligne de recyclage (remarquer VALUES (...) et les colonnes explicitement)
            query_recy = """
                INSERT INTO recyclage
                    (numero_serie, type_objet, id_fut, sur_site, date_rebut, cause)
                VALUES (%s, %s, %s, %s, NOW(), %s)
            """
            param_recy = (numero_serie_batt, type_obj, id_fut, "oui", cause)
            cursor.execute(query_recy, param_recy)
    
            # 6) Mettre à jour le suivi_production
            query_sp = "UPDATE suivi_production SET recyclage = 1, date_recyclage = NOW() WHERE numero_serie_batterie = %s"
            cursor.execute(query_sp, (numero_serie_batt,))
            cursor.execute("Update produit_voltr set statut=%s where numero_serie_produit=%s",("recyclee",numero_serie_batt))
    
            # 7) Commit et message utilisateur
            conn.commit()
            emplacement = f"fut {id_fut}"
            messagebox.showinfo("Recyclage réussi",
                                f"La batterie {numero_serie_batt} recyclée dans un fut {type_fut} : {emplacement}")
    
        except Exception as e:
            # rollback en cas d'erreur et message
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
            messagebox.showerror("Erreur BDD", f"Erreur lors du recyclage : {e}")
        finally:
            if cursor:
                try:
                    cursor.close()
                except Exception:
                    pass
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
                
    # ===================== Handlers & utilitaires =====================
    
    def _exp_on_entry_change(self, event):
        """Ajoute la batterie dans 'sélectionnées' quand l'entrée atteint 9 chiffres."""
        txt = self.exp_numero_serie_batt_entry.get().strip()
        if len(txt) == 9:
            self._exp_add_to_selection(txt)
            self.exp_numero_serie_batt_entry.delete(0, tk.END)
    
    def _exp_on_available_double_click(self, event):
        """Double-clic sur une batterie disponible -> ajoute à la sélection."""
        sel = self.exp_listbox_batt.curselection()
        if not sel:
            return
        value = self.exp_listbox_batt.get(sel[0])
        self._exp_add_to_selection(value)
    
    def _exp_on_selected_double_click(self, event):
        """Double-clic sur une batterie sélectionnée -> la retire de la sélection."""
        sel = self.send_listbox_batt.curselection()
        if not sel:
            return
        value = self.send_listbox_batt.get(sel[0])
        # Retire la première occurrence (il n'y a pas de doublon, donc OK)
        self.send_listbox_batt.delete(sel[0])
        self._exp_update_counter()
    
    def _exp_add_to_selection(self, numero):
        """Ajoute sans doublon à la listbox sélectionnée puis met à jour le compteur."""
        # Anti-doublon
        current = set(self.send_listbox_batt.get(0, tk.END))
        if numero in current:
            return
        self.send_listbox_batt.insert(tk.END, numero)
        self._exp_update_counter()
    
    def _exp_update_counter(self):
        """Mise à jour du compteur de batteries sélectionnées."""
        self._exp_count_var.set(self.send_listbox_batt.size())
        
    def _exp_get_selected_batteries(self):
        """Retourne toutes les batteries présentes dans la listbox 'sélectionnées'."""
        return list(self.send_listbox_batt.get(0, tk.END))
    
    def exp_add_non_conf_batterie(self):
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
                
        if reponse:
        
            gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
            webbrowser.open_new_tab(gg_from)  
        
        self.exp_numero_serie_batt_entry.delete(0, tk.END)
        
        
    def valider_exp(self):
        
        # 0) Récup sélection
        numeros = list(self.send_listbox_batt.get(0, tk.END))
        if not numeros:
            messagebox.showwarning("Avertissement", "Aucune batterie sélectionnée.")
            return
        
        # 1) Référence cible (combobox si case cochée, sinon self.modele)
        target_ref = ""
        if getattr(self, "chk_exp_var", None) and self.chk_exp_var.get():
            target_ref = self.cb_exp.get().strip()
        if not target_ref:
            target_ref = getattr(self, "selected_model", "").strip()
        
        # 2) Lecture des références actuelles via IN (...)
        placeholders = ", ".join(["%s"] * len(numeros))
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
        
            # -- Ajuste ce SELECT selon ton schéma (JOIN si id_produit, sinon colonne directe)
            sql_sel = f"""
                SELECT sp.numero_serie_batterie, p.reference_produit_voltr
                FROM suivi_production AS sp
                JOIN produit_voltr AS p ON p.numero_serie_produit = sp.numero_serie_batterie
                WHERE sp.numero_serie_batterie IN ({placeholders})
            """
            cur.execute(sql_sel, tuple(numeros))
            rows = cur.fetchall()
            num2ref = {str(n): str(r) for (n, r) in rows}
        
            # Vérifs de base (numéros non trouvés)
            manquants = [n for n in numeros if n not in num2ref]
            if manquants:
                messagebox.showerror("Erreur",
                                     f"Références introuvables pour : {', '.join(manquants)}")
                return
        
            # 3) Regroupement par référence
            groups = {}
            for n, ref in num2ref.items():
                groups.setdefault(ref, []).append(n)
        
            # 4) Cohérence vs target_ref
            #    - si target_ref définie : tout ce qui n'a pas cette ref est "à corriger"
            #    - si target_ref vide : on ne corrige pas, mais on exige l'homogénéité
            to_fix = []
            if target_ref:
                for ref, nums in groups.items():
                    if ref != target_ref:
                        to_fix.extend(nums)
        
                if to_fix:
                    # 5) Proposition de correction
                    if not messagebox.askyesno(
                        "Référence différente",
                        f"{len(to_fix)} batterie(s) ne sont pas en '{target_ref}'.\n"
                        "Voulez-vous les mettre toutes à cette référence ?"
                    ):
                        # L’utilisateur refuse -> on arrête le process
                        return
        
                    # 6) Mise à jour des références
                    placeholders_fix = ", ".join(["%s"] * len(to_fix))
                    # -- VERSION avec id_produit (JOIN sur reference_produit)
                    sql_upd = f"""
                        UPDATE produit_voltr 
                        SET reference_produit_voltr = %s
                        WHERE produit_voltr.numero_serie_produit IN ({placeholders_fix})
                    """
                    cur.execute(sql_upd, (target_ref, *to_fix))
                    conn.commit()
        
                    # 7) Re-lecture pour contrôle
                    cur.execute(sql_sel, tuple(numeros))
                    rows = cur.fetchall()
                    num2ref = {str(n): str(r) for (n, r) in rows}
                    groups = {}
                    for n, ref in num2ref.items():
                        groups.setdefault(ref, []).append(n)
        
            # 8) Garde-fou : exiger une seule référence finale
            if len(groups) > 1:
                refs = ", ".join(groups.keys())
                messagebox.showerror("Références multiples",
                                     f"Plusieurs références restent présentes : {refs}\n"
                                     f"Process interrompu.")
                return
        
            # 9) Suite du process OK (une seule ref, corrigée si besoin)
            ref_unique = next(iter(groups.keys()))
            numeros_final = groups[ref_unique]
            
            # 9) FLUX MÉTIER : marquer expédiées
            comment = self.exp_comm_entry.get().strip()
            placeholders_final = ", ".join(["%s"] * len(numeros))
            projet=self.cb_pr.get().strip()
            nom_client=self.cb_cl.get().strip()
            
            conn2 = self.db_manager.connect()
            if conn2:
                try:
                    cursor = conn2.cursor()
                    cursor.execute("SELECT nom_client FROM client ORDER BY nom_client")
                    mots = [row[0].replace(" ", "-") for row in cursor.fetchall()]
                    self.cb_cl["values"] = mots
                except Exception as e:
                    messagebox.showerror("Erreur SQL", f"Chargement clients :\n{e}")
                finally:
                    try: cursor.close()
                    except: pass
                    conn2.close()
            
            #Obtenir id_client
            if nom_client:
                nom_c_bdd=nom_client.replace("-"," ")
                conn2 = self.db_manager.connect()
                if conn2:
                    try:
                        cursor = conn2.cursor()
                        cursor.execute("SELECT id_client FROM client where nom_client = %s",(nom_c_bdd,))
                        id_client = [row[0] for row in cursor.fetchall()]
                    except Exception as e:
                        messagebox.showerror("Erreur SQL", f"Chargement id_client :\n{e}")
                    finally:
                        try: cursor.close()
                        except: pass
                        conn2.close()
    
            if comment:
                
                sql_sp = f"""
                    UPDATE suivi_production
                    SET expedition = 1,
                        date_expedition = NOW(),
                        commentaire = %s
                    WHERE numero_serie_batterie IN ({placeholders_final})
                """
                params_sp = (comment, *numeros)
                
            else:
                
                sql_sp = f"""
                    UPDATE suivi_production
                    SET expedition = 1,
                        date_expedition = NOW()
                    WHERE numero_serie_batterie IN ({placeholders_final})
                """
                params_sp = tuple(numeros)
    
            cur.execute(sql_sp, params_sp)
            
            set_parts = ["statut = 'expediee'"]
            params_mark = []
            
            if nom_client:
                if id_client:
                    set_parts.append("id_client = %s")
                    params_mark.append(id_client)
            
            if projet:
                set_parts.append("numero_projet = %s")
                params_mark.append(projet)
            
            sql_mark = f"""
                UPDATE produit_voltr
                SET {', '.join(set_parts)}
                WHERE numero_serie_produit IN ({placeholders_final})
            """
            
            params_mark = tuple(params_mark) + tuple(numeros)
            
            cur.execute(sql_mark, params_mark)
            
            conn.commit()
    
            # 10) UI : succès + refresh
            messagebox.showinfo("Succès", f"{len(numeros)} batterie(s) marquées expédiées.")
            self.send_listbox_batt.delete(0, tk.END)       # vide la sélection
            self._exp_update_counter()                      # remet le compteur à jour
            for f in self.funcs_to_run:
                f()              # recharge les "disponibles"
            self.exp_comm_entry.delete(0, tk.END)           # vide le commentaire
            
            
        except Exception as e:
            try:
                conn.rollback()
            except:
                pass
            messagebox.showerror("Erreur SQL", f"valider_exp :\n{e}")
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass
        
if __name__ == "__main__":
    app = StockApp()
    app.mainloop()
