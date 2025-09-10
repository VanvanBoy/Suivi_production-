# -*- coding: utf-8 -*-
"""
Created on Thu Aug 21 15:18:57 2025

@author: User
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os, re, shutil
import tkinter as tk
from tkinter import ttk, messagebox
from ttkthemes import ThemedTk
from PIL import Image, ImageTk
import pandas as pd
import mysql.connector
import webbrowser
from datetime import datetime

EXCEL_PATH = r"G:\Drive partagés\VoltR\11_Data\IHM\Suivi_prod_par_modele.xlsx"  

class DBManager:
    def __init__(self):
        """
        self.config = {
            'user': 'Vanvan',
            'password': 'VoltR99!',
            'host': '34.77.226.40',
            'database': 'cellules_batteries_cloud',
            'auth_plugin': 'mysql_native_password'
        }
        """
        
        self.config = {
            'user': 'root',
            'password': 'VoltR99!',
            'host': '127.0.0.1',
            'database': 'bdd_29072025',
            'auth_plugin': 'mysql_native_password'
        }
        
        
        
    def connect(self):
        try:
            return mysql.connector.connect(**self.config)
        except mysql.connector.Error as err:
            messagebox.showerror("Erreur DB", f"Erreur lors de la connexion : {err}")
            return None

class StockApp(ThemedTk):

    
    def __init__(self):
        super().__init__(theme="xpnative")
        self.title("Suivi de production")
        self.geometry("1000x650")

        self.db_manager = DBManager()
        self.picking_file_path = None

        # styles
        self.style = ttk.Style()
        self.style.configure("TLabel", font=('Segoe UI', 11), padding=4)
        self.style.configure("TButton", font=('Segoe UI', 11), padding=6)
        self.style.configure("TEntry", font=('Segoe UI', 11))
        self.style.configure("TCombobox", font=('Segoe UI', 11))
        self.style.configure("Danger.TButton", foreground="black", background="#F06F65")
        self.style.map("Danger.TButton", background=[('active', '#e55b50')], foreground=[('disabled', 'black')])
        self.style.configure("Good.TButton", foreground="black", background="#0ED329")
        self.style.map("Good.TButton", background=[('active', '#0ED329')], foreground=[('disabled', 'black')])

        
        self.selected_model = None
        self.stage_order = None  

        self._show_model_selector_and_build()

    
    def _show_model_selector_and_build(self):
        try:
            df = pd.read_excel(EXCEL_PATH)
        except Exception as e:
            messagebox.showerror("Excel", f"Impossible de lire l'Excel:\n{e}")
            self.destroy()
            return

        if "nom_modele" not in df.columns:
            messagebox.showerror("Excel", "La colonne 'nom_modele' est absente du fichier.")
            self.destroy()
            return

        self.models = df["nom_modele"].dropna().astype(str).unique().tolist()
        if not self.models:
            messagebox.showerror("Excel", "Aucune valeur dans 'nom_modele'.")
            self.destroy()
            return

        dlg = tk.Toplevel(self)
        dlg.title("Choisir la référence batterie")
        dlg.geometry("420x160")
        dlg.transient(self)
        dlg.grab_set()

        ttk.Label(dlg, text="Référence batterie :").pack(pady=(18, 6))
        cb_var = tk.StringVar()
        cb = ttk.Combobox(dlg, textvariable=cb_var, values=self.models, state="readonly", width=40)
        cb.pack()
        cb.current(0)

        def on_launch():
            ref = cb_var.get()
            row = df.loc[df["nom_modele"].astype(str) == ref]
            if row.empty:
                messagebox.showerror("Sélection", "Référence introuvable.")
                return

            column_to_stage = {
                "picking": "picking",
                "soudure_pack": "pack",
                "soudure_nappe": "nappe",
                "soudure_bms": "bms",
                "wrap": "wrap",
                "fermeture": "fermeture",
                "test_capa": "capa",
                "emballage": "emb",
                "expedition": "exp",
                "recherche": "recherche",
            }

            stage_order = {}
            for col, key in column_to_stage.items():
                if col in df.columns:
                    try:
                        val = int(row.iloc[0][col])
                    except Exception:
                        # NaN / texte -> traite comme 0
                        val = 0
                    stage_order[key] = val

            self.selected_model = ref
            self.stage_order = stage_order
            dlg.destroy()

        ttk.Button(dlg, text="Lancer l'application", style="Good.TButton", command=on_launch).pack(pady=16)

        # centre la boîte
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 210
        y = self.winfo_y() + (self.winfo_height() // 2) - 80
        dlg.geometry(f"+{x}+{y}")

        self.wait_window(dlg)

        if self.stage_order is None:
            
            self.destroy()
            return

        
        self._create_widgets_with_order()


    def _create_widgets_with_order(self):
        stage_defs = {
            "picking":   ("Contrôle de picking", self.setup_picking),
            "pack":      ("Contrôle soudure pack", self.setup_pack),
            "nappe":     ("Contrôle soudure nappe", self.setup_nappe),
            "bms":       ("Contrôle soudure BMS", self.setup_bms),
            "wrap":      ("Contrôle wrap", self.setup_wrap),
            "fermeture": ("Contrôle fermeture", self.setup_fermeture),
            "capa":      ("Test de capacité", self.setup_capa),
            "emb":       ("Contrôle emballage", self.setup_emb),
            "exp":       ("Contrôle expédition", self.setup_exp),
            "recherche": ("Recherche de batterie", self.setup_recherche),
        }

        
        ordered_keys = [
            k for k, v in sorted(self.stage_order.items(), key=lambda kv: kv[1])
            if v and v > 0 and k in stage_defs
        ]

        if not ordered_keys:
            messagebox.showwarning("Configuration", "Aucun onglet actif pour cette référence.")
          
            ordered_keys = ["picking"]

        notebook = ttk.Notebook(self)
        notebook.pack(expand=True, fill="both")

        for key in ordered_keys:
            title, setup_fn = stage_defs[key]
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=title)
            setup_fn(frame)

    
    def set_photo(self, label: tk.Label, chemin_image: str, size=(200, 200)):
        """Charge une image et l'affecte AU SEUL label donné."""
        try:
            img = Image.open(chemin_image)
            img = img.resize(size)
            photo = ImageTk.PhotoImage(img)
            label.config(image=photo, text="")
            label.image = photo  # garder une référence pour éviter GC
        except Exception as e:
            messagebox.showerror("Erreur image", f"Impossible de charger l'image : {e}")
            
    def verfier_coherence_ref(self,num_batt):
        modele=self.selected_model
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "Select reference_produit_voltr from produit_voltr where numero_serie_produit = %s "
            param = (num_batt,)
            cursor.execute(query, param)
            modele_act=cursor.fetchone()[0]
            if modele_act != modele:
                reponse=messagebox.askyesno("Modèle de batterie différent",f"Le mdodèle de batterie n'est pas coherent /n Passer du modèle {modele_act} au modèle {modele} pour la batterie {num_batt} ?")
                if not reponse:
                    return
                else :
                    cursor.execute("UPDATE produit_voltr SET reference_produit_voltr = %s WHERE numero_serie_produit =% s",(modele,num_batt))
                    messagebox.showinfo("Nouveau modèle", f'la batterie {num_batt} est passé au modèle {modele}.')
                    conn.commit()
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()  
    #------------------------------ Onglet picking -----------------------------------------------
    
    #Front

    def setup_picking(self, frame):
        
        
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.numero_serie_cell_entry = ttk.Entry(left_frame); self.numero_serie_cell_entry.pack(pady=5)
        self.numero_serie_cell_entry.bind("<KeyRelease>",self.check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.numero_serie_batt_entry = ttk.Entry(left_frame); self.numero_serie_batt_entry.pack(pady=5)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.listbox_batt.pack(pady=5, fill='both')
        self.listbox_batt.bind("<<ListboxSelect>>", self.on_select_batt)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="🔁 Remplacement cellule:").pack(pady=5)
        ttk.Label(right_frame, text="N° série cellule:").pack(pady=5)
        self.numero_cell_r_entry = ttk.Entry(right_frame); self.numero_cell_r_entry.pack(pady=5)
        ttk.Label(right_frame, text="Défaut:").pack(pady=5)
        self.combobox_default = ttk.Combobox(right_frame, state="readonly", values=["Non trouvée", "Tension", "Corrosion", "Déformation"])
        self.combobox_default.pack(pady=5)
        self.combobox_default.bind("<<ComboboxSelected>>",self.tension_defaut)
        
        ttk.Label(right_frame, text="Tension cellule NOK:").pack(pady=5)
        self.tension_cell_entry=ttk.Entry(right_frame,state="disabled");self.tension_cell_entry.pack(pady=5)

        ttk.Button(right_frame, text="🔄 Remplacer cellule", command=self.replace_cellule, style="Danger.TButton").pack(pady=10)

        frame_info = tk.Frame(right_frame, width=300, height=100, bg='#D0F5BE'); frame_info.pack(pady=20); frame_info.pack_propagate(False)
        tk.Label(frame_info, text="⚠ Écart maximum de 0.05V", bg='#D0F5BE', fg='black', font=("Segoe UI", 11, 'bold')).pack(expand=True)

        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_picking, style="Good.TButton").pack(pady=10)
        
        self.display_model_list()
            
    #Back
    
    def display_model_list(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv  on sp.numero_serie_batterie = pv.numero_serie_produit where sp.picking_tension is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
    def on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.numero_serie_batt_entry.delete(0, tk.END)
        self.numero_serie_batt_entry.insert(0, selected_value)
                      
    def add_non_conf_batterie(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.numero_serie_batt_entry.get())
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension_fail = picking_tension_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.numero_serie_batt_entry.delete(0, tk.END)
            
    def valider_picking(self):
        num_batt=str(self.numero_serie_batt_entry.get())
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension = 1 where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.numero_serie_batt_entry.delete(0, tk.END)
             
    def tension_defaut(self,event=None):
        defaut=str(self.combobox_default.get())
        if defaut=="Tension":
            self.tension_cell_entry.config(state="normal")
        else :
            self.tension_cell_entry.config(state="disabled")
                       
    def replace_cellule(self):

        self.remplacement_dir = r"C:\Users\User\Desktop\Suivi_prod_rsc"
        os.makedirs(self.remplacement_dir, exist_ok=True)
        # Template 
        self.template_remplacement_path = r"C:\Users\User\Desktop\Suivi_prod_rsc\Template remplacement cellule (13).xlsx"
        
        num = self.numero_cell_r_entry.get().strip()  
        defaut = self.combobox_default.get().strip()                  
        tension_txt = self.tension_cell_entry.get().strip()    
        
        if not num:
            messagebox.showwarning("Champ manquant", "Renseigne le N° de série cellule.")
            return
        if not defaut:
            messagebox.showwarning("Champ manquant", "Sélectionne un défaut.")
            return
        if defaut.lower().startswith("tension"):
            if not tension_txt:
                messagebox.showwarning("Champ manquant",
                                       "Défaut = Tension ➜ renseigne la tension NOK.")
                return
            # optionnel: vérifier numérique
            try:
                float(tension_txt.replace(",", "."))
            except ValueError:
                messagebox.showwarning("Valeur invalide",
                                       "La tension doit être un nombre (ex: 3.72).")
                return
    
        projet = None
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
    
            cur.execute("""
                SELECT pv.numero_projet,pv.numero_serie_produit FROM produit_voltr as pv join cellule as c on pv.numero_serie_produit=c.affectation_produit WHERE c.numero_serie_cellule = %s
            """, (num,))
            row = cur.fetchone()
            if row and row[0]:
                projet = str(row[0]).strip()
                produit = str(row[1]).strip()
            else:
                messagebox.showerror("Projet introuvable",
                                     f"Aucun projet trouvé pour la cellule {num}.")
                return
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer le projet :\n{e}")
            return
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass
    
        today_str = datetime.now().strftime("%Y-%m-%d")
        safe_projet = re.sub(r"[^\w\- ]+", "_", projet)  # sécurité nom de fichier
        out_path = os.path.join(self.remplacement_dir, f"{safe_projet}-{today_str}.xlsx")
    
        tension_val = ""
        if defaut.lower().startswith("tension"):
            # conserver le format texte tel que saisi, ou caster en float si tu préfères
            tension_val = tension_txt.replace(",", ".")
    
        new_row = [produit, num, defaut, tension_val]
    
        try:
            if not os.path.exists(out_path):
                if not os.path.exists(self.template_remplacement_path):
                    messagebox.showerror("Template manquant",
                                         f"Template introuvable :\n{self.template_remplacement_path}")
                    return
                shutil.copyfile(self.template_remplacement_path, out_path)
    
            wb = load_workbook(out_path)

            SHEET_CANDIDATES = ["Remplacement", "Remplacements", "Feuil1", "Données"]
            ws = None
            for name in SHEET_CANDIDATES:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:

                expected_headers = ["Date", "N° série", "Défaut", "Tension"]
                for sh in wb.worksheets:
                    headers = [ (sh.cell(row=1, column=i).value or "").strip() for i in range(1, 5) ]
                    if [h.lower() for h in headers] == [h.lower() for h in expected_headers]:
                        ws = sh
                        break
            if ws is None:
              
                ws = wb.active
    
            ws.append(new_row)

            max_row = ws.max_row
            max_col = ws.max_column
            last_col_letter = get_column_letter(max_col)
    
            if hasattr(ws, "tables"):
                for tbl in list(ws.tables.values()):
                    ref = tbl.ref
                    start, end = ref.split(":")
                    start_col = "".join([c for c in start if c.isalpha()])
                    start_row = int("".join([c for c in start if c.isdigit()]))

                    new_ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
                    tbl.ref = new_ref
    
            if ws.auto_filter and ws.auto_filter.ref:
                start, _ = ws.auto_filter.ref.split(":")
                start_col = "".join([c for c in start if c.isalpha()])
                start_row = int("".join([c for c in start if c.isdigit()]))
                ws.auto_filter.ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
    
            wb.save(out_path)
    
        except Exception as e:
            messagebox.showerror("Erreur Excel", f"Impossible d'écrire dans le fichier :\n{e}")
            return

        try:
            self.numero_cell_r_entry.delete(0, 'end')
            self.combobox_default.set('')
            self.tension_cell_entry.delete(0, 'end')
        except Exception:
            pass
    
        messagebox.showinfo("Remplacement enregistré", f"Ligne ajoutée dans :\n{out_path}")
        
        
    
    def check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.numero_serie_batt_entry.delete(0, tk.END)
                self.numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
            
    #------------------------------ Onglet soudure pack -----------------------------------------------    
    
    #Front
    
    def setup_pack(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.s_numero_serie_cell_entry = ttk.Entry(left_frame); self.s_numero_serie_cell_entry.pack(pady=5)
        self.s_numero_serie_cell_entry.bind("<KeyRelease>",self.s_check_entry_length)
        
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.s_numero_serie_batt_entry = ttk.Entry(left_frame); self.s_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_pack, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.s_listbox_batt = tk.Listbox(right_frame, font=('Segoe UI', 11), height=10); self.s_listbox_batt.pack(pady=5, fill='both')
        self.s_listbox_batt.bind("<<ListboxSelect>>", self.s_on_select_batt)
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_soudure_pack, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_pack()
            
    #Back
    
    def s_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.s_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.s_numero_serie_batt_entry.delete(0, tk.END)
                self.s_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.s_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_pack(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.soudure_pack is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.s_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.s_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_soudure_pack(self):
        num_batt=str(self.s_numero_serie_batt_entry.get())
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack = 1 where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.s_numero_serie_batt_entry.delete(0, tk.END)
            
            
    def add_non_conf_batterie_pack(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.s_numero_serie_batt_entry.get())
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack_fail = soudure_pack_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.s_numero_serie_batt_entry.delete(0, tk.END)
    
    def s_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.s_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.s_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.s_numero_serie_batt_entry.delete(0, tk.END)
        self.s_numero_serie_batt_entry.insert(0, selected_value)
            
    
        
    #------------------------------ Onglet soudure nappe -----------------------------------------------   

    def setup_nappe(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.n_numero_serie_cell_entry = ttk.Entry(left_frame); self.n_numero_serie_cell_entry.pack(pady=5)
        self.n_numero_serie_cell_entry.bind("<KeyRelease>",self.n_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.n_numero_serie_batt_entry = ttk.Entry(left_frame); self.n_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_nappe, style="Danger.TButton").pack(pady=10)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.n_listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.n_listbox_batt.pack(pady=5, fill='both')
        self.n_listbox_batt.bind("<<ListboxSelect>>", self.n_on_select_batt)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Ecart tension modules:").pack(pady=5)
        self.ecart_t_entry = ttk.Entry(right_frame); self.ecart_t_entry.pack(pady=5)
        self.n_label_photo = tk.Label(right_frame, bg="#e0e0e0", width=200, height=200, text="Aperçu photo", anchor='center')
        self.n_label_photo.pack(pady=10)
        self.set_photo(self.n_label_photo, r"C:\Users\User\Downloads\voltr_logo.jpg")
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_soudure_nappe, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_nappe()
        
        #Back
        
    def n_check_entry_length(self, event=None):

        numero_serie_cell = self.n_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.n_numero_serie_batt_entry.delete(0, tk.END)
                self.n_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.n_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_nappe(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.soudure_nappe is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.n_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.n_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_soudure_nappe(self):
        num_batt=str(self.n_numero_serie_batt_entry.get())
        if self.ecart_t_entry.get():
            delta_tension=float(self.ecart_t_entry.get())
        else: 
            messagebox.showerror("Saisie incompléte !","Renseigner l'ecart de tension entre le module min et le module max")
            return
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_nappe = 1,delta_tension_module = %s where numero_serie_batterie = %s "
            param = (delta_tension,num_batt)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.n_numero_serie_batt_entry.delete(0, tk.END)
            
            
    def add_non_conf_batterie_nappe(self):
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.n_numero_serie_batt_entry.get())
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_pack_fail = soudure_nappe_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.n_numero_serie_batt_entry.delete(0, tk.END)
        
    def n_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.n_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.n_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.n_numero_serie_batt_entry.delete(0, tk.END)
        self.n_numero_serie_batt_entry.insert(0, selected_value)
        
    #------------------------------ Onglet bms -----------------------------------------------   

    def setup_bms(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.b_numero_serie_cell_entry = ttk.Entry(left_frame); self.b_numero_serie_cell_entry.pack(pady=5)
        self.b_numero_serie_cell_entry.bind("<KeyRelease>",self.b_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.b_numero_serie_batt_entry = ttk.Entry(left_frame); self.b_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_bms, style="Danger.TButton").pack(pady=10)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.b_listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.b_listbox_batt.pack(pady=5, fill='both')
        self.b_listbox_batt.bind("<<ListboxSelect>>", self.b_on_select_batt)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        self.b_label_photo = tk.Label(right_frame, bg="#e0e0e0", width=200, height=200, text="Aperçu photo", anchor='center')
        self.b_label_photo.pack(pady=10)
        self.set_photo(self.b_label_photo, r"C:\Users\User\Downloads\voltr_logo.jpg")
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_bms, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_bms()
        
    def b_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.b_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.b_numero_serie_batt_entry.delete(0, tk.END)
                self.b_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.b_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_bms(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.soudure_bms is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.b_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.b_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_bms(self):
        num_batt=str(self.b_numero_serie_batt_entry.get())
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_bms = 1 where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.b_numero_serie_batt_entry.delete(0, tk.END)
            
            
    def add_non_conf_batterie_bms(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.b_numero_serie_batt_entry.get())
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET soudure_bms_fail = soudure_bms_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.b_numero_serie_batt_entry.delete(0, tk.END)
    
    def b_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.b_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.b_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.b_numero_serie_batt_entry.delete(0, tk.END)
        self.b_numero_serie_batt_entry.insert(0, selected_value)
        
    
        
    #------------------------------ Onglet wrap -----------------------------------------------   

    def setup_wrap(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.w_numero_serie_cell_entry = ttk.Entry(left_frame); self.w_numero_serie_cell_entry.pack(pady=5)
        self.w_numero_serie_cell_entry.bind("<KeyRelease>",self.w_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.w_numero_serie_batt_entry = ttk.Entry(left_frame); self.w_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_wrap, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.w_listbox_batt = tk.Listbox(right_frame, font=('Segoe UI', 11), height=10); self.w_listbox_batt.pack(pady=5, fill='both')
        self.w_listbox_batt.bind("<<ListboxSelect>>", self.w_on_select_batt)
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_wrap, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_wrap()
        
    def w_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.b_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.b_numero_serie_batt_entry.delete(0, tk.END)
                self.b_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.b_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_wrap(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.wrap is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.w_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.w_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_wrap(self):
        num_batt=str(self.w_numero_serie_batt_entry.get())
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET wrap = 1 where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.w_numero_serie_batt_entry.delete(0, tk.END)
            
            
    def add_non_conf_batterie_wrap(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.w_numero_serie_batt_entry.get())
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET wrap_fail = wrap_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.w_numero_serie_batt_entry.delete(0, tk.END)
    
    def w_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.w_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.w_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.w_numero_serie_batt_entry.delete(0, tk.END)
        self.w_numero_serie_batt_entry.insert(0, selected_value)
        
        
    #------------------------------ Onglet fermeture -----------------------------------------------   
    def setup_fermeture(self, frame):
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.f_numero_serie_cell_entry = ttk.Entry(left_frame); self.f_numero_serie_cell_entry.pack(pady=5)
        self.f_numero_serie_cell_entry.bind("<KeyRelease>",self.f_check_entry_length)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.f_numero_serie_batt_entry = ttk.Entry(left_frame); self.f_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_fermeture, style="Danger.TButton").pack(pady=10)

        right_frame = ttk.Frame(frame); right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.f_listbox_batt = tk.Listbox(right_frame, font=('Segoe UI', 11), height=10); self.f_listbox_batt.pack(pady=5, fill='both')
        self.f_listbox_batt.bind("<<ListboxSelect>>", self.f_on_select_batt)
        ttk.Label(right_frame, text="Tension en fin de test:").pack(pady=5)
        self.tension_end_entry = ttk.Entry(right_frame); self.tension_end_entry.pack(pady=5)
        self.f_label_photo = tk.Label(right_frame, bg="#e0e0e0", width=200, height=200, text="Aperçu photo", anchor='center')
        self.f_label_photo.pack(pady=10)
        self.set_photo(self.f_label_photo, r"C:\Users\User\Downloads\voltr_logo.jpg")
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_fermeture, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_fermeture()
        
    def f_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.f_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.f_numero_serie_batt_entry.delete(0, tk.END)
                self.f_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.f_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_fermeture(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.fermeture_batt is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.f_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.f_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_fermeture(self):
        num_batt=str(self.f_numero_serie_batt_entry.get())
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET fermeture_batt = 1 where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.f_numero_serie_batt_entry.delete(0, tk.END)
            
            
    def add_non_conf_batterie_fermeture(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        num_batt=str(self.f_numero_serie_batt_entry.get())
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET fermeture_fail = fermeture_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.f_numero_serie_batt_entry.delete(0, tk.END)
    
    def f_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.f_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.f_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.f_numero_serie_batt_entry.delete(0, tk.END)
        self.f_numero_serie_batt_entry.insert(0, selected_value)
        
        

    #------------------------------ Onglet emballage -----------------------------------------------   
    def setup_emb(self, frame):
        
        left_frame = ttk.Frame(frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.emb_numero_serie_batt_entry = ttk.Entry(left_frame); self.emb_numero_serie_batt_entry.pack(pady=5)
        self.emb_numero_serie_batt_entry.bind("<KeyRelease>",self.emb_check_entry_length)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie_emb, style="Danger.TButton").pack(pady=10)
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        self.emb_listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.emb_listbox_batt.pack(pady=5, fill='both')
        self.emb_listbox_batt.bind("<<ListboxSelect>>", self.emb_on_select_batt)

        right_frame= ttk.Frame(frame); right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Modèle batterie").pack(pady=5)
        cb_var_m = tk.StringVar()
        self.cb_emb = ttk.Combobox(right_frame, textvariable=cb_var_m, values=self.models, state="disabled", width=40)
        self.cb_emb.pack(pady=5)
        
        chk_var = tk.BooleanVar(value=False)
        
        def toggle_combobox():
            self.cb_emb.configure(state="readonly" if chk_var.get() else "disabled")

        chk = ttk.Checkbutton(right_frame, text="Changer le modèle",
                      variable=chk_var, command=toggle_combobox)
        chk.pack(pady=5)
        
        ttk.Button(right_frame, text="✅ Contrôle OK", command=self.valider_emballage, style="Good.TButton").pack(pady=10)
        
        self.display_model_list_emballage()

    def emb_check_entry_length(self, event=None):
        # (Optionnel) on ne déclenche que quand on a 12 caractères
        numero_serie_cell = self.emb_numero_serie_cell_entry.get().strip()
        if len(numero_serie_cell) != 12:
            return
    
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "SELECT affectation_produit FROM cellule WHERE numero_serie_cellule = %s"
            param = (numero_serie_cell,)  # <-- tuple !
            cursor.execute(query, param)
            row = cursor.fetchone()
            if row and row[0]:
                num_batt = str(row[0])
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
                self.emb_numero_serie_batt_entry.insert(0, num_batt)
            else:
                # Efface si pas trouvé (optionnel)
                self.emb_numero_serie_batt_entry.delete(0, tk.END)
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
    
    def display_model_list_emballage(self):
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            query = "SELECT numero_serie_batterie from suivi_production as sp join produit_voltr as pv on sp.numero_serie_batterie = pv.numero_serie_produit where sp.emballage is null and pv.reference_produit_voltr = %s "
            param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
    
            # Vide la Listbox
            self.emb_listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.emb_listbox_batt.insert(tk.END, batt)
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()

    def valider_emballage(self):
        num_batt=str(self.f_numero_serie_batt_entry.get())
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET emballage = 1 where numero_serie_batterie = %s "
            param = (num_batt,)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.f_numero_serie_batt_entry.delete(0, tk.END)
            
            
    def add_non_conf_batterie_emb(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if not reponse:
            return
        
        gg_from="https://docs.google.com/forms/d/e/1FAIpQLSeDivu0XsxeXnRhJrf1AyoVaywsDtKyPdaCJ9_-EfSQ-3-x7A/viewform?usp=sf_link"
        webbrowser.open_new_tab(gg_from) 
        
        self.emb_numero_serie_batt_entry.delete(0, tk.END)
    
    def emb_on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        selection = self.emb_listbox_batt.curselection()
        if not selection:
            return  # rien de sélectionné
        
        # Récupère le texte de la ligne sélectionnée
        selected_value = self.emb_listbox_batt.get(selection[0])
    
        # Mets à jour l'Entry produit
        self.emb_numero_serie_batt_entry.delete(0, tk.END)
        self.emb_numero_serie_batt_entry.insert(0, selected_value)
    
    #------------------------------ Onglet capa -----------------------------------------------   

    def setup_capa(self, frame):
        # Conteneur principal
        wrap = ttk.LabelFrame(frame, text="Test", padding=12)
        wrap.pack(fill="both", expand=True, padx=8, pady=8)
    
        # Tableau
        table_frame = ttk.Frame(wrap)
        table_frame.pack(fill="both", expand=True, padx=8, pady=(6, 12))
    
        cols = ("N° Série", "Modèle", "Capacité", "Tension de fin de test")
        self.test_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        for c in cols:
            self.test_tree.heading(c, text=c)
            self.test_tree.column(c, anchor="w", width=170 if c == "Tension de fin de test" else 140)
    
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.test_tree.yview)
        self.test_tree.configure(yscrollcommand=yscroll.set)
    
        self.test_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
    
        # Gros bouton ovale centré (canvas)
        btn_frame = ttk.Frame(wrap)
        btn_frame.pack(fill="x", pady=8)
        canvas = tk.Canvas(btn_frame, width=260, height=64, highlightthickness=0, bg=self.cget("background"))
        canvas.pack(pady=6)
        # Dessin de l’ovale bleu
        oval = canvas.create_oval(4, 4, 256, 60, fill="#2E62FF", outline="#1b3fb3", width=2)
        txt  = canvas.create_text(130, 32, text="Traiter les fichiers", fill="white", font=("Segoe UI", 11, "bold"))

        def _on_click(_):
            # placeholder : à remplacer par ton traitement
            messagebox.showinfo("Info", "Bouton 'Traiter les fichiers' cliqué (front uniquement).")
    
        # Bind sur l’ovale et le texte
        canvas.tag_bind(oval, "<Button-1>", _on_click)
        canvas.tag_bind(txt,  "<Button-1>", _on_click)
        
    #------------------------------ Onglet recherche -----------------------------------------------   
        
    def setup_recherche(self, frame):
            # ----- Layout principal : gauche | boutons | droite -----
        container = ttk.Frame(frame); container.pack(fill="both", expand=True, padx=12, pady=12)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=0)
        container.columnconfigure(2, weight=2)
        container.rowconfigure(0, weight=1)
    
        # ------- Colonne gauche : entrées + combobox + listbox -------
        left = ttk.LabelFrame(container, text="Recherche", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
    
        ttk.Label(left, text="N° série cellule").grid(row=0, column=0, sticky="w")
        self.rech_entry_cell = ttk.Entry(left, width=28)
        self.rech_entry_cell.grid(row=1, column=0, sticky="we", pady=(0,8))
    
        ttk.Label(left, text="N° série batterie").grid(row=2, column=0, sticky="w")
        self.rech_entry_batt = ttk.Entry(left, width=28)
        self.rech_entry_batt.grid(row=3, column=0, sticky="we", pady=(0,8))
    
        ttk.Label(left, text="Référence batterie").grid(row=4, column=0, sticky="w")
        self.rech_model_var = tk.StringVar(value=(self.selected_model or (self.models[0] if self.models else "")))
        self.rech_combo = ttk.Combobox(left, textvariable=self.rech_model_var,
                                       values=self.models, state="readonly", width=30)
        self.rech_combo.grid(row=5, column=0, sticky="we", pady=(0,8))
        self.rech_combo.bind("<<ComboboxSelected>>", lambda e: self._rech_on_model_change())
    
        ttk.Label(left, text="Liste batterie").grid(row=6, column=0, sticky="w")
        lb_frame = ttk.Frame(left); lb_frame.grid(row=7, column=0, sticky="nsew")
        left.rowconfigure(7, weight=1)
    
        self.rech_listbox = tk.Listbox(lb_frame, height=10, activestyle="dotbox")
        yscroll = ttk.Scrollbar(lb_frame, orient="vertical", command=self.rech_listbox.yview)
        self.rech_listbox.configure(yscrollcommand=yscroll.set)
        self.rech_listbox.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
    
        # Double-clic => déplacer à droite
        self.rech_listbox.bind("<Double-1>", lambda e: self._rech_move_right())
    
        # --------- Colonne boutons centraux ----------
        mid = ttk.Frame(container); mid.grid(row=0, column=1, sticky="ns")
        for i in range(3): mid.rowconfigure(i, weight=1)
        ttk.Button(mid, text="→", width=3).grid(row=0, column=0, pady=4)
        ttk.Button(mid, text="←", width=3).grid(row=1, column=0, pady=4)
    
        # --------- Colonne droite : table dynamique ----------
        right = ttk.LabelFrame(container, text="Sélection / Détails", padding=10)
        right.grid(row=0, column=2, sticky="nsew", padx=(8,0))
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
    
        self.rech_right_title = ttk.Label(right, text="", font=("", 10, "bold"))
        self.rech_right_title.grid(row=0, column=0, sticky="w", pady=(0,6))
    
        tv_frame = ttk.Frame(right); tv_frame.grid(row=1, column=0, sticky="nsew")
        self.rech_tree = ttk.Treeview(tv_frame, columns=(), show="headings", selectmode="browse")
        y2 = ttk.Scrollbar(tv_frame, orient="vertical", command=self.rech_tree.yview)
        self.rech_tree.configure(yscrollcommand=y2.set)
        self.rech_tree.pack(side="left", fill="both", expand=True)
        y2.pack(side="right", fill="y")
    
    #------------------------------ Onglet expedition -----------------------------------------------   

    def setup_exp(self, frame):
        
        top_frame=ttk.Frame(frame); top_frame.pack(side='top',fill='both',expand=True)
        bottom_frame=ttk.Frame(frame); bottom_frame.pack(side='top',fill='both',expand=True)
        
        left_frame = ttk.Frame(top_frame); left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.exp_numero_serie_batt_entry = ttk.Entry(left_frame); self.exp_numero_serie_batt_entry.pack(pady=5)
        ttk.Button(left_frame, text="❌ Non conforme", command=self.add_non_conf_batterie, style="Danger.TButton").pack(pady=10)
        ttk.Label(left_frame, text="Liste des batteries du modèle emballées:").pack(pady=5)
        self.exp_listbox_batt = tk.Listbox(left_frame, font=('Segoe UI', 11), height=10); self.exp_listbox_batt.pack(pady=5, fill='both')

        right_frame= ttk.Frame(top_frame); right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        ttk.Label(right_frame, text="Client").pack(pady=5)
        cb_var_client = tk.StringVar()
        
        conn=self.db_manager.connect()
        cursor=conn.cursor()
        query="Select nom_client from client order by nom_client"
        cursor.execute(query)
        result_client = cursor.fetchall()
        cursor.close()
        conn.close()
        
        mots_possibles_client= list(result_client)
        mots_possibles_client = [item[0].replace(" ", "-") for item in mots_possibles_client]
        
        self.cb_cl = ttk.Combobox(right_frame, textvariable=cb_var_client, values=mots_possibles_client, state="readonly", width=40)
        self.cb_cl.pack(pady=5)
        
        ttk.Label(right_frame, text="Projet").pack(pady=5)
        self.cb_pr = ttk.Combobox(right_frame, state="readonly", width=40)
        self.cb_pr.pack(pady=5)
        
        ttk.Label(right_frame, text="Modèle batterie").pack(pady=5)
        cb_var_m_exp = tk.StringVar()
        self.cb_exp = ttk.Combobox(right_frame, textvariable=cb_var_m_exp, values=self.models, state="disabled", width=40)
        self.cb_exp.pack(pady=5)
        
        chk_exp_var = tk.BooleanVar(value=False)
        
        def toggle_combobox_exp():
            self.cb_exp.configure(state="readonly" if chk_exp_var.get() else "disabled")

        chk_exp = ttk.Checkbutton(right_frame, text="Changer le modèle",
                      variable=chk_exp_var, command=toggle_combobox_exp)
        chk_exp.pack(pady=5)
        
        ttk.Label(right_frame,text='Quantité:').pack(pady=5)
        compteur_qte=ttk.Label(right_frame,text="0").pack(pady=5)
        
        ttk.Label(bottom_frame,text='Batteries selectionnées').pack(pady=5)
        self.send_listbox_batt = tk.Listbox(bottom_frame, font=('Segoe UI', 11), height=10); self.send_listbox_batt.pack(pady=5, fill='both')
        
        ttk.Button(bottom_frame, text="✅ Contrôle OK", command=self.valider_picking, style="Good.TButton").pack(pady=10)
        
if __name__ == "__main__":
    app = StockApp()
    app.mainloop()
