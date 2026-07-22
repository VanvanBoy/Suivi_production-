# -*- coding: utf-8 -*-
"""
Module extrait automatiquement de Suivi_de_production_prod_V5_5_claude.py
Regroupe les méthodes liées a : TabPicking
"""
import os, re, shutil
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from PIL import Image, ImageTk
import pandas as pd
import mysql.connector
import webbrowser
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from config import EXCEL_PATH, REMPLACEMENT_DIR, TEMPLATE_REMPLACEMENT_PATH, NON_CONFORMITE_FORM_URL


class TabPickingMixin:
    def setup_picking(self, frame):
        
        
        left_frame = ttk.Frame(frame)
        left_frame.pack(side="left", fill='both', expand=True, padx=20, pady=20)
        
        ttk.Label(left_frame, text="N° série d'une cellule du produit:").pack(pady=5)
        self.numero_serie_cell_entry = ttk.Entry(left_frame)
        self.numero_serie_cell_entry.pack(pady=5)
        self.numero_serie_cell_entry.bind("<KeyRelease>", self.check_entry_length)
        
        ttk.Label(left_frame, text="N° série produit:").pack(pady=5)
        self.numero_serie_batt_entry = ttk.Entry(left_frame)
        self.numero_serie_batt_entry.pack(pady=5)
        self.numero_serie_batt_entry.bind("<KeyRelease>", self.check_entry_length_batt)
        
        
        
        ttk.Label(left_frame, text="Liste des batteries du modèle:").pack(pady=5)
        
        # --- Bloc dédié Listbox + Scrollbar ---
        listbox_frame = tk.Frame(left_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
        
        self.listbox_batt = tk.Listbox(
            listbox_frame,
            font=('Segoe UI', 11),
            height=10
        )
        self.listbox_batt.pack(side="left", fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.listbox_batt.yview)
        scrollbar.pack(side="right", fill="y")
        
        self.listbox_batt.config(yscrollcommand=scrollbar.set)
        self.listbox_batt.bind("<<ListboxSelect>>", self.on_select_batt)
        # --- fin bloc listbox ---
        
        ttk.Button(
            left_frame, text="❌ Non conforme",
            command=self.add_non_conf_batterie,
            style="Danger.TButton"
        ).pack(pady=10)
        
        # --- Cadre droit inchangé ---
        right_frame = ttk.Frame(frame)
        right_frame.pack(side="right", fill='both', expand=True, padx=20, pady=20)
        
        ttk.Label(right_frame, text="🔁 Remplacement cellule:").pack(pady=5)
        ttk.Label(right_frame, text="N° série cellule:").pack(pady=5)
        self.numero_cell_r_entry = ttk.Entry(right_frame)
        self.numero_cell_r_entry.pack(pady=5)
        
        ttk.Label(right_frame, text="Défaut:").pack(pady=5)
        self.combobox_default = ttk.Combobox(
            right_frame, state="readonly",
            values=["Non trouvée", "Tension", "Corrosion", "Déformation"]
        )
        self.combobox_default.pack(pady=5)
        self.combobox_default.bind("<<ComboboxSelected>>", self.tension_defaut)
        
        ttk.Label(right_frame, text="Tension cellule NOK:").pack(pady=5)
        self.tension_cell_entry = ttk.Entry(right_frame, state="disabled")
        self.tension_cell_entry.pack(pady=5)
        self.tension_cell_entry.bind("<KeyRelease>", self.convert_comma_to_dot)
        
        ttk.Button(
            right_frame, text="🔄 Demande remplacement cellule",
            command=self.replace_cellule, style="Danger.TButton"
        ).pack(pady=10)
        
        frame_info = tk.Frame(right_frame, width=300, height=100, bg='#D0F5BE')
        frame_info.pack(pady=20)
        frame_info.pack_propagate(False)
        tk.Label(
            frame_info, text="⚠ Écart maximum de 0.05V",
            bg='#D0F5BE', fg='black',
            font=("Segoe UI", 11, 'bold')
        ).pack(expand=True)
        
        self.btn_valider_picking = ttk.Button(
        right_frame, text="✅ Contrôle OK",
        command=self.valider_picking, style="Good.TButton"
        )
        self.btn_valider_picking.pack(pady=10)
        
        self.display_model_list()
        
        self.make_tab_chain(
            [
                self.numero_serie_batt_entry,
                self.btn_valider_picking  # placer le bouton en dernier si tu veux que Tab atteigne le bouton
            ],
            submit_button=self.btn_valider_picking,
            ring=True,
            enter_from_fields=False  # sécurité : Enter depuis les champs ne déclenche pas le bouton
        )
        
        return self.numero_serie_batt_entry

            
    #Back
    
    def display_model_list(self):
        stage_act='picking'
        modele=str(self.selected_model)
        conn = self.db_manager.connect()
        if not conn:
            return
        try: 
            cursor= conn.cursor()
            if modele[:8]=="PPTR018A":
                query="""
                SELECT sp.numero_serie_batterie
                FROM suivi_production AS sp
                JOIN produit_voltr AS pv
                  ON sp.numero_serie_batterie = pv.numero_serie_produit
                WHERE pv.reference_produit_voltr LIKE %s
                AND (sp.picking_tension = 0 or sp.picking_tension is null)
                AND (sp.recyclage = 0 or sp.recyclage is null)
                  """             
                param=(modele[:8]+"%",)
                
                
            else :
                query=self.build_stage_query(stage_act)
                param=(modele,)
            cursor.execute(query, param)
            rows = cursor.fetchall()  
    
            # Transforme en liste simple
            liste_batteries = [str(r[0]) for r in rows]
            
            try:
                first_frac, last_frac = self.listbox_batt.yview()
            except Exception:
                first_frac = 0.0
    
            # Vide la Listbox
            self.listbox_batt.delete(0, tk.END)
            
            # Ajoute chaque batterie dans la Listbox
            for batt in liste_batteries:
                self.listbox_batt.insert(tk.END, batt)
            
            try:
                self.listbox_batt.yview_moveto(first_frac)
            except Exception:
                pass
    
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                cursor.close()
            except:
                pass
            conn.close()
            
    def on_select_batt(self, event=None):
        """Quand on sélectionne une batterie dans la listbox, la mettre dans l'Entry produit."""
        self._fill_entry_from_listbox_selection(self.listbox_batt, self.numero_serie_batt_entry)
                      
    def add_non_conf_batterie(self):
        
        reponse = messagebox.askyesno("Non conformité", "Ouvrir une non-conformité ?")
        if reponse:
            gg_from = NON_CONFORMITE_FORM_URL
            webbrowser.open_new_tab(gg_from) 

        num_batt=str(self.numero_serie_batt_entry.get())
        
        self.verfier_coherence_ref(num_batt)
        
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension_fail = picking_tension_fail + 1 where numero_serie_batterie = %s "
            param = (num_batt,)
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            self.numero_serie_batt_entry.delete(0, tk.END)
            
    def valider_picking(self):
        num_batt=str(self.numero_serie_batt_entry.get())
        if not num_batt:
            return
        if str(self.selected_model)[:8]!='PPTR018A':
            directive=self.verfier_coherence_ref(num_batt)
            if directive=='stop':
                return
        # 1) Pré-requis
        stage_col='picking'
        if not self.verif_etape_act_non_ok(stage_col, num_batt):
            return
        if not self._check_prereqs_and_warn(num_batt, stage_col):
            return
        conn = self.db_manager.connect()
        visa=self.db_manager.user
        if not conn:
            return
        try:
            cursor = conn.cursor()
            query = "UPDATE suivi_production SET picking_tension = 1, date_picking_tension = NOW(), visa_picking_tension = %s where numero_serie_batterie = %s "
            param = (visa,num_batt)  
            cursor.execute(query, param)
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer les données :\n{e}")
        finally:
            try:
                conn.commit()
                cursor.close()
            except:
                pass
            conn.close()
            messagebox.showinfo("Controle OK",f"Batterie {num_batt} controlée")
            self.numero_serie_batt_entry.delete(0, tk.END)
            """
            self.display_model_list()
            self.display_model_list_pack()
            self.display_model_list_nappe()
            self.display_model_list_bms()
            self.display_model_list_wrap()
            self.display_model_list_fermeture()
            self.display_model_list_emballage()
            self.display_model_list_exp()
            """
            
            self._focus_active_tab()
            for f in self.funcs_to_run:
                f()
             
    def tension_defaut(self,event=None):
        defaut=str(self.combobox_default.get())
        if defaut=="Tension":
            self.tension_cell_entry.config(state="normal")
        else :
            self.tension_cell_entry.config(state="disabled")
                       
    def replace_cellule(self):

        self.remplacement_dir = REMPLACEMENT_DIR
        os.makedirs(self.remplacement_dir, exist_ok=True)
        # Template 
        self.template_remplacement_path = TEMPLATE_REMPLACEMENT_PATH
        
        num = self.numero_cell_r_entry.get().strip()  
        defaut = self.combobox_default.get().strip()                  
        tension_txt = self.tension_cell_entry.get().strip()    
        
        if not num:
            messagebox.showwarning("Champ manquant", "Renseigne le N° de série cellule.")
            return
        if not defaut:
            messagebox.showwarning("Champ manquant", "Sélectionne un défaut.")
            return
        if defaut.lower().startswith("tension"):
            if not tension_txt:
                messagebox.showwarning("Champ manquant",
                                       "Défaut = Tension ➜ renseigne la tension NOK.")
                return
            # optionnel: vérifier numérique
            try:
                float(tension_txt.replace(",", "."))
            except ValueError:
                messagebox.showwarning("Valeur invalide",
                                       "La tension doit être un nombre (ex: 3.72).")
                return
    
        projet = None
        conn = self.db_manager.connect()
        if not conn:
            return
        try:
            cur = conn.cursor()
    
            cur.execute("""
                SELECT pv.numero_projet,pv.numero_serie_produit FROM produit_voltr as pv join cellule as c on pv.numero_serie_produit=c.affectation_produit WHERE c.numero_serie_cellule = %s
            """, (num,))
            row = cur.fetchone()
            if row and row[0]:
                projet = str(row[0]).strip()
                produit = str(row[1]).strip()
            else:
                messagebox.showerror("Projet introuvable",
                                     f"Aucun projet trouvé pour la cellule {num}.")
                return
        except Exception as e:
            messagebox.showerror("Erreur SQL", f"Impossible de récupérer le projet :\n{e}")
            return
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass
    
        today_str = datetime.now().strftime("%Y-%m-%d")
        safe_projet = re.sub(r"[^\w\- ]+", "_", projet)  # sécurité nom de fichier
        out_path = os.path.join(self.remplacement_dir, f"{safe_projet}-{today_str}.xlsx")
    
        tension_val = ""
        if defaut.lower().startswith("tension"):
            # conserver le format texte tel que saisi, ou caster en float si tu préfères
            tension_val = tension_txt.replace(",", ".")
    
        new_row = [produit, num, defaut, tension_val]
    
        try:
            if not os.path.exists(out_path):
                if not os.path.exists(self.template_remplacement_path):
                    messagebox.showerror("Template manquant",
                                         f"Template introuvable :\n{self.template_remplacement_path}")
                    return
                shutil.copyfile(self.template_remplacement_path, out_path)
    
            wb = load_workbook(out_path)

            SHEET_CANDIDATES = ["Remplacement", "Remplacements", "Feuil1", "Données"]
            ws = None
            for name in SHEET_CANDIDATES:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:

                expected_headers = ["Date", "N° série", "Défaut", "Tension"]
                for sh in wb.worksheets:
                    headers = [ (sh.cell(row=1, column=i).value or "").strip() for i in range(1, 5) ]
                    if [h.lower() for h in headers] == [h.lower() for h in expected_headers]:
                        ws = sh
                        break
            if ws is None:
              
                ws = wb.active
    
            ws.append(new_row)

            max_row = ws.max_row
            max_col = ws.max_column
            last_col_letter = get_column_letter(max_col)
    
            if hasattr(ws, "tables"):
                for tbl in list(ws.tables.values()):
                    ref = tbl.ref
                    start, end = ref.split(":")
                    start_col = "".join([c for c in start if c.isalpha()])
                    start_row = int("".join([c for c in start if c.isdigit()]))

                    new_ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
                    tbl.ref = new_ref
    
            if ws.auto_filter and ws.auto_filter.ref:
                start, _ = ws.auto_filter.ref.split(":")
                start_col = "".join([c for c in start if c.isalpha()])
                start_row = int("".join([c for c in start if c.isdigit()]))
                ws.auto_filter.ref = f"{start_col}{start_row}:{last_col_letter}{max_row}"
    
            wb.save(out_path)
    
        except Exception as e:
            messagebox.showerror("Erreur Excel", f"Impossible d'écrire dans le fichier :\n{e}")
            return

        try:
            self.numero_cell_r_entry.delete(0, 'end')
            self.combobox_default.set('')
            self.tension_cell_entry.delete(0, 'end')
        except Exception:
            pass
    
        messagebox.showinfo("Remplacement enregistré", f"Ligne ajoutée dans :\n{out_path}")
        
        
    
    def check_entry_length(self, event=None):
        """Recherche la batterie associée dès que le n° de série cellule (12 car.) est saisi."""
        self._lookup_batterie_from_cellule(self.numero_serie_cell_entry, self.numero_serie_batt_entry)
            
            
    #------------------------------ Onglet soudure pack -----------------------------------------------    
    
    #Front
    
